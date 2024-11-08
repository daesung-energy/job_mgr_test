from django.shortcuts import redirect, render, get_object_or_404
from django.http import HttpResponse
#BsPrd 메시지 끌고옴
from .models import BsPrd, CcCdDetail, CcCdHeader, BsJob, BsDept, BsJobDept, BsMbr, BsWorkGrade, MbrJobGrp, MbrJobGrpDetail, JobTask, JobTaskAdj, JobActivity, BsPrd, BsAcnt, BsJobResp, JobSpcfc, BsStdWrkTm, BsWlOvSht, BsPosGrade, BsPosList, BsDeptGrp, BsDeptResp, BsDeptGrpDomain, BsTtlList, BsTtlCnt, BsMbrGrp, BsMbrGrpNm, VJb110F, VJb111
#확인하는 메시지 끌고옴
#from .models import TextConfirm
from datetime import datetime
import datetime as dt
#from pytz import timezone
from django.utils import timezone
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.core.exceptions import ObjectDoesNotExist
import json
import numpy as np
#import mysql.connector
import pandas as pd
from django.contrib import messages, auth
from django.db import IntegrityError
from django.db.models import Case, When, Value, CharField
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from sqlalchemy import create_engine
import traceback
import pymysql
from decimal import Decimal
from django.contrib.auth.hashers import check_password
import os #추가
from pathlib import Path #추가
# from decimal import Decimal, ROUND_HALF_UP

now = dt.datetime.now() #지금 날짜를 가져옴

# Create your views here.
def index(request):
    return HttpResponse("Hello World")


def my_view(request):

    original_rows=JobTask.objects.filter(prd_cd="2022A", dept_cd="DD10")
    data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
            'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_sum': rows.work_lv_sum } for rows in original_rows]

    df1 = pd.DataFrame(data_list)

    # 데이터프레임을 JSON 형식으로 변환하여 전달
    df_json = df1.to_json(orient='records')

    context = {
     'data' : df_json
    }

    # return render(request, 'jobs/AGgrid.html', {'html_table': html_table})
    return render(request, 'jobs/AGgrid.html', context)


def BS101(request): # 회기 관리(회기 등록) 초기화면

    context1 = { #context를 넘겨줌. context는 어떤 type도 가능(?)
        'today_date1' : str(dt.datetime.today()).split()[0],
        'title' : '회기 관리',
        'tab' : 'tab1', # 회기 등록을 선택했다는 키값
        'prd' : BsPrd.objects.all(), # 모든 회기 목록
        'prd_cd_selected' : BsPrd.objects.last().prd_cd, # 회기 선택값
        'modified' : 'n', # 회기 복사나 삭제 작업을 하지 않았다는 키값(메시지용),
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username), # 로그인한 부서가 관리 부서인지 여부. y이면 관리 부서이고, 모든 메뉴를 다 사용할 수 있다. n이면 직무관리/비번변경만 가능
    }

    return render(request, 'jobs/BS101.html', context1) #장고가 context를 meshup해서 html template으로 보내줌


def BS200(request): # 직무 조사 초기화면

    result_object = BsPrd.objects.all().order_by('turn').last()
    prd_cd_selected = result_object.prd_cd

    context2 = { #context를 넘겨줌. context는 어떤 type도 가능(?)
        'title' : '직무 조사',
        'prd' : BsPrd.objects.all(),
        # 'job_srv_str_dt' : str(result_object.job_srv_str_dt).split()[0],
        # 'job_srv_end_dt' : str(result_object.job_srv_end_dt).split()[0],
        'prd_cd_selected' : prd_cd_selected,
        # 'result_object' : result_object,
        'modify' : "n", #수정할 수 없도록 키값 부여
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        # 'prd_selected' : BsPrd.
    }

    return render(request, 'jobs/BS200.html', context2) #장고가 context를 meshup해서 html template으로 보내줌


def BS300(request): # 조직 정보 초기화면 + 회기 선택 화면

    context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
        'prd' : BsPrd.objects.all(),
        'prd_cd_selected' : BsPrd.objects.all().last().prd_cd, # 마지막 회기가 디폴트로 뜰 것임
        'activate' : 'no', #버튼 컨트롤 off
        'title' : '조직 정보', # 제목
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

    if request.method == 'POST':

        prd_cd_selected = request.POST["prd_cd"]

        context = {
            'prd' : BsPrd.objects.all(),
            'activate' : 'yes', #버튼 컨트롤 on
            'title' : '조직 정보', # 제목
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'prd_cd_selected' : prd_cd_selected
        }

    return render(request, 'jobs/BS300.html', context)


def BS103(request): # 회기 마감 초기화면

    # 회기 최종 마감일에 표시되는 날짜와 최종 마감 버튼 control
    if BsPrd.objects.last().prd_done_yn == "Y": # 디폴트로 뜨는 마지막 회기가 마감된 회기이면
        today_date = str(BsPrd.objects.last().prd_end_dt).split()[0] # 회기 최종 마감일에는 마감된 날짜가 표시된다.
        register_act = 'no' # 최종 마감 버튼은 비활성화시킨다.

    else: # 디폴트로 뜨는 마지막 회기가 마감되지 않은 회기이면
        today_date = str(dt.datetime.today()).split()[0] # 회기 최종 마감일에는 오늘 날짜가 표시된다.
        register_act = 'yes' # 최종 마감 버튼을 활성화시킨다.

    context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        'title' : '회기 관리',
        'tab' : 'tab2',
        'prd' : BsPrd.objects.all(),
        'prd_cd_selected' : BsPrd.objects.last().prd_cd,
        'today_date' : today_date, # 디폴트로 뜨는 마지막 회기의 마감일자 혹은 오늘 일자
        'register_act' : register_act # 버튼 컨트롤 키
    }

    return render(request, 'jobs/BS103.html', context) #장고가 context를 meshup해서 html template으로 보내줌


def BS105(request): # 표준 정보 초기화면

    prd_cd_selected = BsPrd.objects.all().last().prd_cd

    try :

        context = {
            'title' : '표준 정보', # 제목
            'prd' : BsPrd.objects.all(),
            # 'activate' : "none", #회기가 설정되었을 때만 input 수정 가능하도록 함.
            'total_dys' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).total_dys,
            'std_wrk_able_dys' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_able_dys,
            'std_wrk_able_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_able_tm,
            'ade_ot_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).ade_ot_tm,
            'std_wrk_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_tm,
            'min_max' : BsWlOvSht.objects.get(prd_cd=prd_cd_selected),
            'g1' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G1"),
            'g2' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G2"),
            'g3' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G3"),
            'g4' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G4"),
            'g5' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G5"),
            's_m_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="수석부장"),
            'd_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="부장"),
            'a_d_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="차장"),
            'm_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="과장"),
            'a_m_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="대리"),
            'c_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="사원"),
            't_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="기능직"),
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : prd_cd_selected,
            'activate' : "activate",
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        return render(request, 'jobs/BS105.html', context)

    except BsStdWrkTm.DoesNotExist as e:

        # error_message = "해당 회기 및 부서에는 데이터가 없습니다."
        messages.error(request, f'에러 발생: {"해당 회기 내 데이터 없음"}')

        context = {
            'prd' : BsPrd.objects.all().order_by('-prd_cd'),
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        return render(request, 'jobs/BS105.html', context)


def BS106(request): # 직무 관리 초기화면 + 회기 선택 화면

    context = {
        'title' : '직무 관리', # 제목
        'prd_list' : BsPrd.objects.all().order_by('prd_cd'), # 회기 리스트
        'prd_selected' : BsPrd.objects.all().order_by('prd_cd').last().prd_cd, # 마지막 회기가 디폴트로 뜰 것임.
        'job_type_selected' : "former", # 직무유형 선택 전
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        # attribute error 처리 필요
    }

    if request.method == 'POST':
        prd_selected = request.POST["prd_selected"]

        context = {
            'title' : '직무 관리', # 제목
            'prd_list' : BsPrd.objects.all().order_by('prd_cd'), # 회기 리스트
            'prd_selected' : prd_selected,
            'job_type_selected' : "former", # 직무유형 선택 전
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

    return render(request, 'jobs/BS106.html', context)


def CC102(request): ## 공통코드관리 초기화면

    context = {
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        'CC_list': CcCdHeader.objects.exclude(domain_cd="A5").all(), #CcCdHeader는 코드에 대한 개략적 정보, A5행을 제외한 모든 값을 가져옴,
        'text' : "초기",
        'title' : '공통 코드' # 제목
    }

    return render(request, 'jobs/CC102.html', context)


def CC105(request): # 비밀번호 변경 초기화면

    context = {
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'title' : '비밀번호 변경', # 제목
        }

    if request.method == "POST":

        user = request.user
        origin_password = request.POST["origin_password"]

        if check_password(origin_password, user.password):

            new_password = request.POST["new_password"]
            confirm_password = request.POST["confirm_password"]

            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                auth.login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                # 성공 메시지 띄워주기
                messages.success(request, '비밀번호를 변경했습니다.')
                return render(request, 'jobs/CC105.html', context)

            else:
                messages.error(request, '새 비밀번호가 일치하지 않습니다.')

        else:
            messages.error(request, '현재 비밀번호가 일치하지 않습니다.')

        return render(request, 'jobs/CC105.html', context)

    else:

        return render(request, 'jobs/CC105.html', context)


def popup(request):
    return render(request, 'jobs/popup.html')


def JB101(request): # 부서 기본정보 초기화면 + 회기 선택 화면

    # 초기화면

    user_name = request.user.username # 아이디(ID___)
    # print('user name', user_name)
    dept_login = get_dept_code(user_name) # 로그인한 부서의 부서코드. 회기 수정해야 함. 아이디 테이블에 없기 때문.
    dept_login_nm = BsDept.objects.get(prd_cd=BsPrd.objects.all().last().prd_cd, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

    context = {
        'title' : '부서 기본정보', # 제목
        'prd_list' : BsPrd.objects.all(),
        'user_name' : user_name,
        'activate' : 'no', #버튼 컨트롤 off
        'prd_cd_selected' : BsPrd.objects.all().last().prd_cd,
        'dept_login_nm' : dept_login_nm,
        'dept_mgr_yn': get_dept_mgr_yn(user_name),
        'status' : 'tab_before'
    }

    # 회기 선택 후 화면

    if request.method == 'POST':

        prd_cd_selected = request.POST["prd_cd_selected"]
        # dept_login = get_dept_code_new(user_name, prd_cd_selected)

        try:
            dept_login_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

            context = {
                'title' : '부서 기본정보', # 제목
                'prd_list' : BsPrd.objects.all(),
                'user_name' : user_name,
                'activate' : 'no', #버튼 컨트롤 off
                'prd_cd_selected' : prd_cd_selected,
                'dept_login_nm' : dept_login_nm,
                'dept_mgr_yn': get_dept_mgr_yn(user_name),
                'status' : 'tab_before'
            }

        except ObjectDoesNotExist as e: # 로그인한 부서가 해당 회기에 없을 때

            # 오류 메시지 띄워주고 탭 선택할 수 없도록 막는다.
            messages.error(request, '해당 회기에 로그인한 부서가 없습니다.')

            context = {
                'title' : '부서 기본정보', # 제목
                'prd_list' : BsPrd.objects.all(),
                'user_name' : user_name,
                'activate' : 'no', #버튼 컨트롤 off
                'prd_cd_selected' : prd_cd_selected,
                'status' : 'tab_before',
                'dept_mgr_yn': get_dept_mgr_yn(user_name),
                'tab_activate' : 'no' # 탭 선택 못하도록 막음
            }

    return render(request, 'jobs/JB101.html', context)


def JB102(request): # 직무 기본정보 초기화면 - 회기 선택은 JB102_1에서 함

    # 경영기획팀이면 team_list 있고 경영기획팀 아니면 team_list 없음

    last_prd_cd = BsPrd.objects.all().last().prd_cd # 가장 최근 회기. default로 띄워줌

    dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
    dept_login_nm = BsDept.objects.get(prd_cd=last_prd_cd, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

    context = {
        'prd_list' : BsPrd.objects.all(),
        'title' : '직무 기본정보', # 제목
        'prd_selected' : last_prd_cd,
        'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
        'job_type_selected' : "former" # 직무유형 선택 전
    }

    if dept_login == "DD06":
        context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd)
        context['dept_login_nm'] = dept_login_nm
        context['dept_selected'] = dept_login
    else:
        context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login)
        context['dept_login_nm'] = dept_login_nm
        context['dept_selected'] = dept_login

    return render(request, 'jobs/JB102.html', context)


def JB103(request): # 직무 상세정보 초기화면(가장 최근 회기와 로그인 부서에 대한 정보를 가져옴)

    last_prd_cd = BsPrd.objects.all().last().prd_cd # 가장 최근 회기. default로 띄워줌

    dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
    dept_login_nm = BsDept.objects.get(prd_cd=last_prd_cd, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

    context = {
        'prd_list' : BsPrd.objects.all(),
        'title' : '직무 상세정보', # 제목
        'prd_selected' : last_prd_cd,
        'prd_done' : BsPrd.objects.get(prd_cd=last_prd_cd).prd_done_yn,
        'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
        # 'dept_selected_key' : "former" # 부서 선택 전
    }

    if dept_login == "DD06":
        context['dept_list'] = BsDept.objects.filter(prd_cd=last_prd_cd) #마지막 회기의 부서 띄워주는게 좋을 듯
        context['dept_login_nm'] = dept_login_nm
        context['dept_cd_selected'] = dept_login
    else:
        context['dept_list'] = BsDept.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login)
        context['dept_login_nm'] = dept_login_nm
        context['dept_cd_selected'] = dept_login

    # 회기, 부서 데이터에 해당하는 JobTask 값에 접근하여, dataframe 생성
    original_rows=JobTask.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함

    data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows]

    df1 = pd.DataFrame(data_list)

    # job_activity 접근
    original_rows_2=JobActivity.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함
    data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq, 'act_prfrm_cnt' : rows.act_prfrm_cnt,
                'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                'dept_rltd': rows.dept_rltd, 'final_rpt_to' : rows.final_rpt_to, 'rpt_nm': rows.rpt_nm,
                'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq, 'act_seq': rows.act_seq} for rows in original_rows_2]

    df2 = pd.DataFrame(data_list_2)
    # print(df2)
    # df2.to_excel('df2.xlsx')

    try:

        df3 = pd.merge(df1, df2).sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq']) # job_task와 job_activity merge, 순서는 job_seq, duty_seq, task_seq, act_seq 순

        # job_nm 찾기
        original_rows_3 = BsJob.objects.filter(prd_cd=last_prd_cd)
        data_list_3 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_3]
        df4 = pd.DataFrame(data_list_3)

        df3 = pd.merge(df3, df4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

        # df3.to_excel('df3.xlsx')
        df_json = df3.to_json(orient='records')

        context.update({'data' : df_json})

        # context['dept_mgr_yn'] = get_dept_mgr_yn(request.user.username),

    except pd.errors.MergeError as e:

        messages.error(request, '해당 회기에 로그인한 부서의 정보가 없습니다.')

        # 버튼 컨트롤 다 막아야 함
        context.update({'button_control' : 'no'})

        context.update({'data' : 'null'})

        # context['dept_mgr_yn'] = get_dept_mgr_yn(request.user.username),

    return render(request, 'jobs/JB103.html', context)


def JB103_1(request): # 직무 상세정보 회기 선택 후 화면(부서 띄워주는 화면).

    if request.method == 'POST':

        #html에서 회기 선택
        prd_selected = request.POST["prd_selected"]

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드

        try:
            dept_login_nm = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

            context = {
                'title' : '직무 상세정보', # 제목
                'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                # 'dept_selected_key' : "former" # 부서 선택 전
            }

            if dept_login == "DD06":
                context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected) #마지막 회기의 부서 띄워주는게 좋을 듯
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login
            else:
                context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login

            # 회기, 부서 데이터에 해당하는 JobTask 값에 접근하여, dataframe 생성
            original_rows=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함

            data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                        'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                        'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                        'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                        'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows]

            df1 = pd.DataFrame(data_list)

            # job_activity 접근
            original_rows_2=JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함
            data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                        'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq, 'act_prfrm_cnt' : rows.act_prfrm_cnt,
                        'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                        'dept_rltd': rows.dept_rltd, 'final_rpt_to' : rows.final_rpt_to, 'rpt_nm': rows.rpt_nm,
                        'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq, 'act_seq': rows.act_seq} for rows in original_rows_2]

            df2 = pd.DataFrame(data_list_2)

            df3 = pd.merge(df1, df2).sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq']) # job_task와 job_activity merge

            # job_nm 찾기
            original_rows_3 = BsJob.objects.filter(prd_cd=prd_selected)
            data_list_3 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_3]
            df4 = pd.DataFrame(data_list_3)

            df3 = pd.merge(df3, df4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

            # df3.to_excel('df3.xlsx')
            df_json = df3.to_json(orient='records')

            context.update({'data' : df_json})

        except pd.errors.MergeError as e: # 바꾼 회기에 그 부서가 없을 때(dept_login_nm)이 없을 때.

            # 메시지 띄움
            messages.error(request, '해당 회기에 로그인한 부서의 정보가 없습니다.')

            # 회기는 그대로 띄워주되, data는 null로 처리하고, button_control 막아준다.
            context = {
                'title' : '직무 상세정보', # 제목
                'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                # 'dept_selected_key' : "former", # 부서 선택 전
                'button_control' : 'no',
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            context.update({'data' : 'null'})

            # 부서 리스트 만들어주기
            if dept_login == "DD06":
                context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected) #마지막 회기의 부서 띄워주는게 좋을 듯
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login
            else:
                context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login


        except ObjectDoesNotExist as e1:

                # 메시지 띄움
                messages.error(request, '해당 회기에 로그인한 부서가 없습니다.')

                # 회기는 그대로 띄워주되, data는 null로 처리하고, button_control 막아준다.
                context = {
                    'title' : '직무 상세정보', # 제목
                    'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    # 'dept_selected_key' : "former", # 부서 선택 전
                    'button_control' : 'no',
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                }

                context.update({'data' : 'null'})


    return render(request, 'jobs/JB103.html', context)


def JB103_2(request): # 직무 상세정보 부서 선택 후 화면(직무 띄워주는 화면)_경영기획팀만 해당

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"]

        # select box에서 팀을 선택함에 따라, 직무(job)칸에는 해당 팀의 직무 목록이 뜨게 된다. select box 형태로.
        dept_cd_selected = request.POST["dept_cd_selected"]

        # 회기, 부서 데이터에 해당하는 JobTask 값에 접근하여, dataframe 생성
        original_rows=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함

        data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                    'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                    'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                    'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                    'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows]

        df1 = pd.DataFrame(data_list)

        # job_activity 접근
        original_rows_2=JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                    'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq, 'act_prfrm_cnt' : rows.act_prfrm_cnt,
                    'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                    'dept_rltd': rows.dept_rltd, 'final_rpt_to' : rows.final_rpt_to, 'rpt_nm': rows.rpt_nm,
                    'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq, 'act_seq': rows.act_seq} for rows in original_rows_2]

        df2 = pd.DataFrame(data_list_2)
        # df2엑셀로 저장
        # df2.to_excel('df2.xlsx')

        try:

            df3 = pd.merge(df1, df2).sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq']) # job_task와 job_activity merge

            # job_nm 찾기
            original_rows_3 = BsJob.objects.filter(prd_cd=prd_selected)
            data_list_3 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_3]
            df4 = pd.DataFrame(data_list_3)

            df3 = pd.merge(df3, df4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

            # df3.to_excel('df3.xlsx')
            df_json = df3.to_json(orient='records')

            context = {
                'title' : '직무 상세정보', # 제목
                'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'dept_list' : BsDept.objects.filter(prd_cd=prd_selected), #부서 목록은 그대로 둔다.
                'dept_cd_selected' : dept_cd_selected,
                'data' : df_json,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                # 'dept_selected_key' : "latter" # 부서 선택 후
                # 데이터프레임을 JSON 형식으로 변환하여 전달
            }

        except pd.errors.MergeError as e: # 경영기획팀이 회기 내에서 부서를 선택했는데, 그 부서가 회기에 존재하긴 하지만 직무 데이터가 없을 경우(신생 부서)
            # data null, button_ctrl 막음

            # 메시지 띄움
            messages.error(request, '해당 회기에 선택한 부서의 직무정보가 없습니다.')

            # 회기는 그대로 띄워주되, data는 null로 처리하고, button_control 막아준다.
            context = {
                'title' : '직무 상세정보', # 제목
                'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'dept_list' : BsDept.objects.filter(prd_cd=prd_selected), #부서 목록은 그대로 둔다.
                'dept_cd_selected' : dept_cd_selected,
                'button_control' : 'no',
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            context.update({'data' : 'null'})

    return render(request, 'jobs/JB103.html', context)


def JB103_3(request): # 직무 상세정보 저장, 취소 버튼 누른 후

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"]
        dept_cd_selected = request.POST["dept_cd_selected"]

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        if 'action' in request.POST: # 저장, 삭제, 추가, 취소 버튼을 눌렀을 때

            action = request.POST["action"]

            if action == "action1" : # 저장 버튼을 눌렀을 때

                # JSON 데이터를 받아옵니다.
                json_data = request.POST.get('jsonData')
                # JSON 문자열을 Python 객체로 변환합니다.
                data = json.loads(json_data)
                # Pandas DataFrame으로 변환합니다.
                df = pd.DataFrame(data).sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq'])
                # df.to_excel('df.xlsx')

                # df에서 job_cd가 JC001 or JC002 or JC004이면, work_lv_imprt, work_lv_dfclt, work_lv_prfcn, work_lv_sum, prfrm_tm_ann을 0으로 바꿔준다.
                # 그리고 act_prfrm_cnt, act_prfrm_cnt_ann, act_prfrm_tm_cs, act_prfrm_tm_ann, act_prfrm_tm_ann을 0으로 바꿔준다.
                df.loc[df['job_cd'].isin(['JC001', 'JC002', 'JC004']), ['work_lv_imprt', 'work_lv_dfclt', 'work_lv_prfcn', 'work_lv_sum', 'prfrm_tm_ann']] = 0
                df.loc[df['job_cd'].isin(['JC001', 'JC002', 'JC004']), ['act_prfrm_cnt', 'act_prfrm_cnt_ann', 'act_prfrm_tm_cs']] = 0
                df.loc[df['job_cd'].isin(['JC001', 'JC002', 'JC004']), ['act_prfrm_tm_ann']] = 0.0

                # print('df', df)

                # job_task와 job_activity로 나누기
                df_task = df[['job_cd', 'duty_nm', 'task_nm', 'task_prsn_chrg', 'work_lv_imprt', 'work_lv_dfclt',
                            'work_lv_prfcn', 'work_lv_sum', 'work_grade', 'work_attrbt', 'prfrm_tm_ann', 'job_seq', 'duty_seq', 'task_seq']] #act_seq 신경쓰기
                df_activity = df[['job_cd', 'duty_nm', 'task_nm', 'act_nm', 'act_prsn_chrg', 'act_prfrm_freq', 'act_prfrm_cnt', 'act_prfrm_cnt_ann', 'act_prfrm_tm_cs', 'act_prfrm_tm_ann',
                                'dept_rltd', 'final_rpt_to', 'rpt_nm', 'job_seq', 'duty_seq', 'task_seq', 'act_seq']]
                df_task = df_task.drop_duplicates() # df_task에서 중복된 것들은 없애준다.
                # df_task.to_excel('df_task.xlsx')
                # df_activity.to_excel('df_activity.xlsx')
                # print('df_task', df_task)
                # print('df_activity', df_activity)


                # DB에 저장하기 - 기존 데이터와 비교하여 추가, 삭제. JobTask부터 먼저 처리하고 JobActivity 처리
                ################################################################ job_task 접근 ################################################################

                # 기존 데이터(해당 회기에 해당 부서의 job_task)에 접근하여 dataframe 생성
                original_rows=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected)
                data_list = [{'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                            'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                            'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                            'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                            'prfrm_tm_ann': rows.prfrm_tm_ann, 'job_seq':rows.job_seq, 'duty_seq':rows.duty_seq, 'task_seq':rows.task_seq } for rows in original_rows]
                df1 = pd.DataFrame(data_list)

                # df_task과 df1의 prfrm_tm_ann은 한번씩 object로 바뀌는 문제가 있음. 이를 해결하기 위해 df_task의 prfrm_tm_ann을 float으로 바꿔줌
                df1['prfrm_tm_ann'] = df1['prfrm_tm_ann'].astype(float)
                # df_task['prfrm_tm_ann'] = df_task['prfrm_tm_ann'].astype(float)
                df_task.loc[:, 'prfrm_tm_ann'] = df_task.loc[:, 'prfrm_tm_ann'].astype(float)

                # 비교 하는 부분 - merge 기능을 이용해 추가된 행, 삭제된 행을 추출할 것이다. 수정은 삭제 후 추가로 볼 것이다.
                # df_left_task : df1(DB)에 있고 df_task(UI)에 없는 것. 즉, 삭제된 것
                df_left_task = pd.merge(df1, df_task, how='outer', indicator=True).query('_merge == "left_only"').drop(columns=['_merge']).reset_index(drop=True)
                # df_right_task : df_task(UI)에 있고 df1(DB)에 없는 것. 즉, 추가된 것
                df_right_task = pd.merge(df1, df_task, how='outer', indicator=True).query('_merge == "right_only"').drop(columns=['_merge']).reset_index(drop=True)

                # print(df_left_task)
                # print(df_right_task)

                # df_left_task를 먼저 다룸. 삭제를 먼저 하고 추가를 나중에 할 것이다. 그래야 중복같은 까다로운 문제를 피할 수 있다. job_cd, duty_nm, task_nm을 보고 삭제할 것이다.

                if df_left_task.empty == False: # df_left_task가 비어있지 않다면, 즉, 삭제할 것이 있다면
                    for i in range(0, len(df_left_task)): # df_left_task의 행 수만큼 반복
                        # df_left_task에 있다는 것은 UI에는 없다는 뜻이다. 즉 삭제해도 되는 값이다.
                        # df_left_task의 i행 1열이 job_cd, 2열이 duty_nm, 3열이 task_nm이다.
                        # DB에서 job_cd가 df_left_task의 i행 1열과 같고, duty_nm이 df_left_task의 i행 2열과 같고, task_nm이 df_left_task의 i행 3열과 같은 것을 찾아서 삭제할 것이다.
                        # 여기서 row_to_delete 따로 정의하고 (get으로) row_to_delete.delete()하면 싹 다 지워짐. prd_cd도 잘 지정해줘야 함.
                        # print('jobTask행 삭제')
                        JobTask.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_cd_selected,
                                                job_cd_id=df_left_task.iloc[i, 0], duty_nm=df_left_task.iloc[i, 1], task_nm=df_left_task.iloc[i, 2]).delete()

                if df_right_task.empty == False: # df_right_task가 비어있지 않다면, 즉 추가할 것이 있다면
                    for i in range(0, len(df_right_task)): # df_right_task의 행 수만큼 반복
                        # df_right_task에 있다는 것은 UI에만 있다는 뜻이다. 즉 추가해야 하는 값이다.
                        # 여기서 .save()쓰면 foreign key 때문에 참조무결성 오류 발생하므로 create를 써준다.
                        # 이미 JobTask에 있는 값(회기, 부서, 직무, 책무, 과업이름이 같은 과업이라면 -- 같은 과업이라는 얘기다.)이면 추가하지 않는다.
                        # if JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected, job_cd=df_right_task.iloc[i, 0],
                        #                         duty_nm=df_right_task.iloc[i, 1], task_nm=df_right_task.iloc[i, 2], task_prsn_chrg=df_right_task.iloc[i, 3],
                        #                         work_lv_imprt=df_right_task.iloc[i, 4], work_lv_dfclt=df_right_task.iloc[i, 5], work_lv_prfcn=df_right_task.iloc[i, 6],
                        #                         work_lv_sum=df_right_task.iloc[i, 7], work_grade_id=df_right_task.iloc[i, 8], work_attrbt=df_right_task.iloc[i, 9],
                        #                         prfrm_tm_ann=df_right_task.iloc[i, 10]).exists():
                        if JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected, job_cd=df_right_task.iloc[i, 0],
                                                   duty_nm=df_right_task.iloc[i, 1], task_nm=df_right_task.iloc[i, 2]).exists():
                            pass
                        else: # 존재하지 않는 값이면 추가한다. 이 때 JC001, JC002, JC004이면 일부 값들이 0이나 0.0으로 저장이 된다.
                            JobTask.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_cd_selected, job_cd_id=df_right_task.iloc[i, 0], duty_nm=df_right_task.iloc[i, 1],
                                                    task_nm=df_right_task.iloc[i, 2], task_prsn_chrg=df_right_task.iloc[i, 3], work_lv_imprt=df_right_task.iloc[i, 4],
                                                    work_lv_dfclt=df_right_task.iloc[i, 5], work_lv_prfcn=df_right_task.iloc[i, 6], work_lv_sum=df_right_task.iloc[i, 7],
                                                    work_grade_id=df_right_task.iloc[i, 8], work_attrbt=df_right_task.iloc[i, 9], prfrm_tm_ann=df_right_task.iloc[i, 10],
                                                    job_seq=df_right_task.iloc[i, 11], duty_seq=df_right_task.iloc[i, 12], task_seq=df_right_task.iloc[i, 13])


                # print('jobTask', JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected))
                ################################################################ job_activity 접근 ################################################################
                # job_activity 접근
                # JobTask가 수정된 후, JobActivity의 정보를 가져와 데이터프레임을 만든다.
                original_rows_2=JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected)
                data_list_2 = [{'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                            'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq, 'act_prfrm_cnt' : rows.act_prfrm_cnt,
                            'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                            'dept_rltd': rows.dept_rltd, 'final_rpt_to' : rows.final_rpt_to, 'rpt_nm': rows.rpt_nm,
                            'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq, 'act_seq': rows.act_seq} for rows in original_rows_2]
                df2 = pd.DataFrame(data_list_2) # job_task를 수정한 다음 상태의 job_activity로 만든 데이터프레임
                # print('df2_act', df2)
                # df_activity와 df2의 자료형 정리
                # print('df2', df2.dtypes)
                # print('df_activity', df_activity.dtypes)

                if df2.empty == False: # df2가 비어있지 않다면 즉, UI에서 수정을 할 때 한 줄이라도 그대로 둔 상태라면 merge이용해 비교를 해서 삭제할 거 삭제하고 추가할 거 추가한다.
                    # print('df2가 비어있지 않다면')

                    df2['act_prfrm_cnt_ann'] = df2['act_prfrm_cnt_ann'].astype(float) # 테이블에서는 int로 되어있는데, df2에서는 float로 되어있어서 맞춰줌
                    df_activity.loc[:, 'act_prfrm_cnt_ann'] = df_activity.loc[:, 'act_prfrm_cnt_ann'].astype(float) # df_activity의 act_prfrm_cnt_ann을 float로 바꿔줌
                    df2['act_prfrm_tm_cs'] = df2['act_prfrm_tm_cs'].astype(float) # 테이블에서는 int로 되어있는데, df2에서는 float로 되어있어서 맞춰줌
                    df_activity.loc[:, 'act_prfrm_tm_cs'] = df_activity.loc[:, 'act_prfrm_tm_cs'].astype(float) # df_activity의 act_prfrm_tm_cs를 float로 바꿔줌
                    df2.loc[:, 'act_prfrm_tm_ann'] = df2.loc[:, 'act_prfrm_tm_ann'].astype(float) # df2의 act_prfrm_tm_ann을 float로 바꿔줌
                    df_activity.loc[:, 'act_prfrm_tm_ann'] = df_activity.loc[:, 'act_prfrm_tm_ann'].astype(float) # df_activity의 act_prfrm_tm_ann을 float로 바꿔줌

                    # 비교 하는 부분 - merge 기능을 이용해 추가된 행, 삭제된 행을 추출할 것이다. 수정은 삭제 후 추가로 볼 것이다.
                    # df2가 비어있지 않다면 비교를 해야 한다.
                    # df2(DB)에 있고 df_activity(UI)에 없는 것. 즉, 삭제된 것
                    df_left_activity = pd.merge(df2, df_activity, how='outer', indicator=True).query('_merge == "left_only"').drop(columns=['_merge']).reset_index(drop=True)
                    # df_activity(UI)에 있고 df2(DB)에 없는 것. 즉, 추가된 것
                    df_right_activity = pd.merge(df2, df_activity, how='outer', indicator=True).query('_merge == "right_only"').drop(columns=['_merge']).reset_index(drop=True)

                    # print('df_left_activity', df_left_activity)
                    # df_right_activity.to_excel('df_right_activity.xlsx')

                    if df_left_activity.empty == False: # df_left_activity가 비어있지 않다면 즉, 삭제할 것이 있다면
                        for i in range(0, len(df_left_activity)): # df_left_activity의 행 수만큼 반복
                            # df_left_activity에 있다는 것은 UI에는 없다는 뜻이다. 즉 삭제해도 되는 값이다.
                            # df_left_activity의 i행 1열이 job_cd, 2열이 duty_nm, 3열이 task_nm, 4열이 act_nm이다.
                            # DB에서 job_cd가 df_left_activity의 i행 1열과 같고, duty_nm이 i행 2열과 같고, task_nm이 i행 3열과 같고, act_nm이 i행 4열과 같은 것을 찾아서 삭제할 것이다.
                            JobActivity.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_cd_selected,
                                                    job_cd_id=df_left_activity.iloc[i, 0], duty_nm_id=df_left_activity.iloc[i, 1],
                                                    task_nm_id=df_left_activity.iloc[i, 2], act_nm=df_left_activity.iloc[i, 3]).delete()

                    if df_right_activity.empty == False: # df_right_activity가 비어있지 않다면, 즉, 추가할 것이 있다면
                        for i in range(0, len(df_right_activity)): # df_right_activity의 행 수만큼 반복
                            # df_right_activity에 있다는 것은 UI에만 있다는 뜻이다. 즉 추가해야 하는 값이다.
                            # 여기서 .save()쓰면 foreign key 때문에 참조무결성 오류 발생하므로 create를 써준다.
                            JobActivity.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_cd_selected, job_cd_id=df_right_activity.iloc[i, 0], duty_nm_id=df_right_activity.iloc[i, 1],
                                                    task_nm_id=df_right_activity.iloc[i, 2], act_nm=df_right_activity.iloc[i, 3], act_prsn_chrg=df_right_activity.iloc[i, 4],
                                                    act_prfrm_freq=df_right_activity.iloc[i, 5], act_prfrm_cnt=df_right_activity.iloc[i, 6], act_prfrm_cnt_ann=df_right_activity.iloc[i, 7], act_prfrm_tm_cs=df_right_activity.iloc[i, 8],
                                                    act_prfrm_tm_ann=df_right_activity.iloc[i, 9], dept_rltd=df_right_activity.iloc[i, 10], final_rpt_to=df_right_activity.iloc[i, 11],
                                                    rpt_nm=df_right_activity.iloc[i, 12], job_seq=df_right_activity.iloc[i, 13], duty_seq=df_right_activity.iloc[i, 14],
                                                    task_seq=df_right_activity.iloc[i, 15], act_seq=df_right_activity.iloc[i, 16])

                    # df_right_task의 job_cd가 JC001, JC002, JC004를 포함할 때, 0으로 바꿔줬던 field들을 다시 Null로 업데이트
                    # JobTask와 JobActivity에 대해서 작업. 다만, act_prfrm_cnt는 예외로 한다.
                    if df_right_task['job_cd'].isin(['JC001', 'JC002', 'JC004']).any():
                        # print('작업한다')
                        JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected, job_cd__in=['JC001', 'JC002', 'JC004']).update(
                            work_lv_imprt=None, work_lv_dfclt=None, work_lv_prfcn=None, work_lv_sum=None, prfrm_tm_ann=None)
                        JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected, job_cd__in=['JC001', 'JC002', 'JC004']).update(
                            act_prfrm_cnt_ann=None, act_prfrm_tm_cs=None, act_prfrm_tm_ann=None)

                # df2가 비어있다면, 즉 JobTask를 수정하고 난 후 JobActivity가 비어있다면(상세정보에서 모든 data에 접근해서 다 수정한 경우)
                else: # 이 경우에는 그냥 df_activity를 이용해 JobActivity를 생성해주면 된다.
                    # print(df_activity)
                    # print('df2가 비어있다')
                    df_activity.loc[:, 'act_prfrm_tm_ann'] = df_activity.loc[:, 'act_prfrm_tm_ann'].astype(float) # df_activity의 act_prfrm_tm_ann을 float로 바꿔줌

                    for i in range(0, len(df_activity)): # df_activity의 행 수만큼 반복
                        # print('생성완료')
                        JobActivity.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_cd_selected, job_cd_id=df_activity.iloc[i, 0], duty_nm_id=df_activity.iloc[i, 1],
                                                task_nm_id=df_activity.iloc[i, 2], act_nm=df_activity.iloc[i, 3], act_prsn_chrg=df_activity.iloc[i, 4],
                                                act_prfrm_freq=df_activity.iloc[i, 5], act_prfrm_cnt=df_activity.iloc[i, 6], act_prfrm_cnt_ann=df_activity.iloc[i, 7], act_prfrm_tm_cs=df_activity.iloc[i, 8],
                                                act_prfrm_tm_ann=df_activity.iloc[i, 9], dept_rltd=df_activity.iloc[i, 10], final_rpt_to=df_activity.iloc[i, 11],
                                                rpt_nm=df_activity.iloc[i, 12], job_seq=df_activity.iloc[i, 13], duty_seq=df_activity.iloc[i, 14],
                                                task_seq=df_activity.iloc[i, 15], act_seq=df_activity.iloc[i, 16])

                    # df_right_task의 job_cd가 JC001, JC002, JC004를 포함할 때, 0으로 바꿔줬던 field들을 다시 Null로 업데이트
                    # JobTask와 JobActivity에 대해서 작업. 다만, act_prfrm_cnt는 예외로 한다.
                    if df_right_task['job_cd'].isin(['JC001', 'JC002', 'JC004']).any():

                        JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected, job_cd__in=['JC001', 'JC002', 'JC004']).update(
                            work_lv_imprt=None, work_lv_dfclt=None, work_lv_prfcn=None, work_lv_sum=None, prfrm_tm_ann=None)
                        JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected, job_cd__in=['JC001', 'JC002', 'JC004']).update(
                            act_prfrm_cnt_ann=None, act_prfrm_tm_cs=None, act_prfrm_tm_ann=None)



                ########################################## DB에 저장 완료. 다시 DB 불러오기 ##########################################
                # DB 다시 접근해서 json 생성
                original_rows_show=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함

                data_list_show = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                            'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                            'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                            'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                            'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows_show]

                df_show = pd.DataFrame(data_list_show)

                # job_activity 접근
                original_rows_show_2=JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함

                data_list_show_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                            'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq, 'act_prfrm_cnt' : rows.act_prfrm_cnt,
                            'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                            'dept_rltd': rows.dept_rltd, 'final_rpt_to' : rows.final_rpt_to, 'rpt_nm': rows.rpt_nm,
                            'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq, 'act_seq': rows.act_seq} for rows in original_rows_show_2]

                df_show_2 = pd.DataFrame(data_list_show_2)

                df_show_3 = pd.merge(df_show, df_show_2).sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq']) # job_task와 job_activity merge

                # job_nm 찾기
                original_rows_show_2 = BsJob.objects.filter(prd_cd=prd_selected)
                data_list_show_4 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_show_2]
                df_show_4 = pd.DataFrame(data_list_show_4)

                df_show_3 = pd.merge(df_show_3, df_show_4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

                # df_show_3.to_excel('df_show_3.xlsx')
                # df3.to_excel('df3.xlsx')
                df_json = df_show_3.to_json(orient='records')

                messages.success(request, "저장되었습니다.")

                context = {
                    'title' : '직무 상세정보', # 제목
                    'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_selected_key' : "former", # 부서 선택 후
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'data' : df_json # 데이터프레임을 JSON 형식으로 변환하여 전달
                }

                if dept_login == "DD06":
                    context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected) #마지막 회기의 부서 띄워주는게 좋을 듯
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_cd_selected'] = dept_cd_selected
                else:
                    context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_cd_selected'] = dept_login

            elif action == "action2" : # 취소 버튼을 눌렀을 때

                # DB 다시 접근해서 json 생성
                original_rows_show=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함

                data_list_show = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                            'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                            'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                            'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                            'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows_show]

                df_show = pd.DataFrame(data_list_show)

                # job_activity 접근
                original_rows_show_2=JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함
                data_list_show_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                            'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq, 'act_prfrm_cnt' : rows.act_prfrm_cnt,
                            'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                            'dept_rltd': rows.dept_rltd, 'final_rpt_to' : rows.final_rpt_to, 'rpt_nm': rows.rpt_nm,
                            'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq, 'act_seq': rows.act_seq} for rows in original_rows_show_2]

                df_show_2 = pd.DataFrame(data_list_show_2)

                df_show_3 = pd.merge(df_show, df_show_2).sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq']) # job_task와 job_activity merge

                # job_nm 찾기
                original_rows_show_2 = BsJob.objects.filter(prd_cd=prd_selected)
                data_list_show_4 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_show_2]
                df_show_4 = pd.DataFrame(data_list_show_4)

                df_show_3 = pd.merge(df_show_3, df_show_4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

                # df3.to_excel('df3.xlsx')
                df_json = df_show_3.to_json(orient='records')

                # messages.error(request, f'수치 데이터를 바르게 입력해 주십시오.')

                context = {
                    'title' : '직무 상세정보', # 제목
                    'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_selected_key' : "former", # 부서 선택 후
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'data' : df_json # 데이터프레임을 JSON 형식으로 변환하여 전달
                }

                if dept_login == "DD06":
                    context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected) #마지막 회기의 부서 띄워주는게 좋을 듯
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_cd_selected'] = dept_cd_selected
                else:
                    context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_cd_selected'] = dept_login

        return render(request, 'jobs/JB103.html', context)
    else:
        return HttpResponse("Invalid request", status=400)


# 직무기술서, 현황표를 위해 추가한 부분###################################################################
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter
from string import ascii_uppercase
from io import BytesIO
import urllib.parse

def JB103_4(request): # 직무 상세정보 직무 현황표, 기술서 print

    if request.method == 'POST':

        action = request.POST["action"] # action1: 직무 현황표, action2: 직무 기술서

        prd_selected = request.POST["prd_selected"]
        dept_cd_selected = request.POST["dept_cd_selected"]
        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        # if action == "action1": # 엑셀 다운로드 버튼을 눌렀을 때

        # pymysql을 사용하여 데이터베이스에 연결
        conn = pymysql.connect(
            host='218.159.103.251', # 데이터베이스 주소
            user='jyj', # 데이터베이스 사용자 이름
            password='1234', # 데이터베이스 비밀번호
            db='jobdb',
            charset='utf8',
            cursorclass=pymysql.cursors.DictCursor
        )

        # 쿼리를 실행하고 DataFrame으로 반환하는 함수
        def get_data(conn, query):
            with conn.cursor() as cursor:
                cursor.execute(query)
                result = cursor.fetchall()
            return pd.DataFrame(result)

        # 직무코드 데이터 가져오기
        query = f"SELECT * FROM bs_job WHERE prd_cd = '{prd_selected}'"
        df_job_code = get_data(conn, query)

        # 수행빈도 데이터 가져오기
        query = "SELECT * FROM cc_cd_detail WHERE domain_cd = 'A5'"
        df_frq_code = get_data(conn, query)

        # 직무-책무-과업-활동 데이터 가져오기
        query = f"SELECT * FROM job_task WHERE prd_cd = '{prd_selected}' AND dept_cd = '{dept_cd_selected}'"
        df1 = get_data(conn, query)

        query = f"SELECT * FROM job_activity WHERE prd_cd = '{prd_selected}' AND dept_cd = '{dept_cd_selected}'"
        df2 = get_data(conn, query)

        # 부서 데이터 가져오기
        query = f"SELECT * FROM bs_dept WHERE prd_cd = '{prd_selected}' AND dept_cd = '{dept_cd_selected}'"
        df_dept = get_data(conn, query)

        # 부서 성과책임 데이터 가져오기
        query = f"SELECT * FROM bs_dept_resp WHERE prd_cd = '{prd_selected}' AND dept_cd = '{dept_cd_selected}'"
        df_resp = get_data(conn, query)

        # 부서 직무 데이터 가져오기
        query = f"SELECT * FROM bs_job_dept WHERE prd_cd = '{prd_selected}' AND dept_cd = '{dept_cd_selected}'"
        df_dept_job = get_data(conn, query)

        # 직무 정보 데이터 가져오기
        query = f"SELECT * FROM bs_job WHERE prd_cd = '{prd_selected}'"
        df_job_info = get_data(conn, query)

        query = f"SELECT * FROM bs_job_resp WHERE prd_cd = '{prd_selected}'"
        df_job_resp = get_data(conn, query)

        # JobTask 테이블을 참고하여 df_dept_job의 순서를 JobTask테이블의 job_seq에 맞춤
        # 해당 회기의 해당 부서의 JobTask테이블 가져옴
        query = f"SELECT * FROM job_task WHERE prd_cd = '{prd_selected}' AND dept_cd = '{dept_cd_selected}'"
        df_job_task = get_data(conn, query)

        # 데이터베이스 연결 닫기
        conn.close()

        # 결합 Key : prd_cd, dept_cd, job_cd, duty_nm, task_nm
        data = pd.merge(df1, df2, how='right', on=['prd_cd', 'dept_cd', 'job_cd', 'duty_nm', 'task_nm'], suffixes=('_left', '_right'))
        # 중복 컬럼 데이터 삭제
        data.drop(data.filter(regex='_left'), axis=1, inplace=True)
        # 직무명 추가
        data = pd.merge(data, df_job_code[['prd_cd', 'job_cd', 'job_nm']], how='left', on=['prd_cd', 'job_cd'])
        # 수행빈도 추가
        data = pd.merge(data, df_frq_code[['cc_code', 'cc_code_nm']], how='left', left_on='act_prfrm_freq', right_on='cc_code')
        # 데이터 순서대로 정렬: 직무-책무-과업-활동
        data.sort_values(by=['job_seq_right', 'duty_seq_right', 'task_seq_right', 'act_seq'], inplace=True)
        # NaN을 None으로 변환
        data = data.replace({np.nan: None})
        # 인덱스 초기화
        data.reset_index(inplace=True)
        assert df2.shape[0] == data.shape[0]    # 전체 데이터 건수가 df2 데이터 건수와 같아야 정상
        # 부서 정보
        dept_nm = df_dept.loc[0].dept_nm
        dept_po = df_dept.loc[0].dept_po

        # 직무 현황표 버튼 눌렀을 때
        if action == "action1":

            wb = Workbook()

            # 테두리 적용
            BORDER_THIN_UP = Border(top=Side(style='thin'))
            BORDER_THIN_ALL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            """
            첫번째 Sheet : 표지
            """
            ws = wb.active  # 현재 활성화된 sheet 가져옴
            ws.title = "부서"

            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 85

            # 회기
            subject1 = ws.cell(row=4, column=1)
            subject1.value = "■ 회기"
            subject1.font = Font(bold=True)
            prd = ws.cell(row=4, column=2)
            prd.value = prd_selected

            # 부서
            subject2 = ws.cell(row=6, column=1)
            subject2.value = "■ 부서"
            subject2.font = Font(bold=True)
            dept = ws.cell(row=6, column=2)
            # dept.value = dept_cd
            dept.value = dept_nm    # 부서명

            # 총 인원(PO)
            subject3 = ws.cell(row=8, column=1)
            subject3.value = "■ 총 인원(PO)"
            subject3.font = Font(bold=True)
            po = ws.cell(row=8, column=2)
            po.value = dept_po
            po.alignment = Alignment(horizontal="left", vertical="center")
            note = ws.cell(row=8, column=3)
            note.value = "(직책자포함)"
            note.font = Font(color="0000FF", size=9)

            # 성과책임
            subject4 = ws.cell(row=10, column=1)
            subject4.value = "■ 성과책임"
            subject4.font = Font(bold=True)
            RESP_START_ROW = 11
            for i, r in df_resp.iterrows():
                row_no = RESP_START_ROW+i
                ws.row_dimensions[row_no].height = 40
                resp_no = ws.cell(row=row_no, column=2)
                resp_no.value = "핵심목표 " + str(r['dept_resp_ordr'])
                resp_no.border = BORDER_THIN_ALL
                resp_no.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                resp_nm = ws.cell(row=row_no, column=3)
                resp_nm.value = r['dept_resp']
                resp_nm.border = BORDER_THIN_ALL
                resp_nm.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

            # 제목
            title = ws.cell(row=1, column=1)
            title.value = "직무표"
            title.font = Font(color="0000FF", size=25, bold=True)
            title.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            ws.row_dimensions[1].height = 50

            # Page Setup
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            ws.print_options.horizontalCentered = True

            """
            두번째 Sheet : 직무 데이터
            """
            ws_data = wb.create_sheet("직무표", 2) # 2번째 index에 sheet 생성

            HEADER_1 = 1     # 항목명 첫째줄
            HEADER_2 = 2     # 항목명 둘째줄

            # TITLE_ROW = 1   # 첫번째 행 "제목"

            DATA_START_ROW = 3
            TASK_START_COL = 3

            COL_WIDTH_1 = 7      # 컬럼 크기
            COL_WIDTH_2 = 10      # 컬럼 크기
            COL_WIDTH_3 = 12      # 컬럼 크기
            COL_WIDTH_4 = 15      # 컬럼 크기
            COL_WIDTH_5 = 25      # 컬럼 크기
            COL_WIDTH_6 = 40      # 컬럼 크기

            # 열 너비 지정(A~Z열)
            alphabet_list = list(ascii_uppercase)
            for c in alphabet_list:
                # if c in ["F", "G", "H", "J", "P"]:
                #     ws_data.column_dimensions[c].width = COL_WIDTH_1
                # elif c in ["I", "K", "N", "Q", "R", "S"]:
                #     ws_data.column_dimensions[c].width = COL_WIDTH_2
                if c in ["F", "G", "H", "J", "P", "Q"]:
                    ws_data.column_dimensions[c].width = COL_WIDTH_1
                elif c in ["I", "K", "N", "R", "S"]:
                    ws_data.column_dimensions[c].width = COL_WIDTH_2
                elif c in ["E"]:
                    ws_data.column_dimensions[c].width = COL_WIDTH_3
                elif c in ["A", "D", "O"]:
                    ws_data.column_dimensions[c].width = COL_WIDTH_4
                elif c in ["B", "C", "M", "T"]:
                    ws_data.column_dimensions[c].width = COL_WIDTH_5
                else:
                    ws_data.column_dimensions[c].width = COL_WIDTH_6

            # 데이터 항목 개수
            DATA_COLS = 20

            # 헤더: 행 높이 지정
            ws_data.row_dimensions[1].height = 20

            # 헤더: 항목 명칭
            # header1 = ["직무\n(Job)", "책무\n(Duty)", "과업\n(Task)", "과업 담당자", "업무 특성", \
            #         "업무 수준 및 등급", "", "", "", "", "과업\n수행시간\n(연간)", \
            #         "활동\n(Activity)", "수행 결과물", "최종 보고대상", "관련 부서", \
            #         "수행\n빈도", "수행시간", "", "", "활동 담당자"]
            header1 = ["직무\n(Job)", "책무\n(Duty)", "과업\n(Task)", "과업 담당자", "업무 특성", \
                    "업무 수준 및 등급", "", "", "", "", "과업\n수행시간\n(연간)", \
                    "활동\n(Activity)", "수행 결과물", "최종 보고대상", "관련 부서", \
                    "수행빈도", "", "수행시간", "", "활동 담당자"]
            header2 = ["", "", "", "", "정형/비정형", \
                    "중요도\n(1~5)", "난이도\n(1~5)", "숙련도\n(1~5)", "업무수준\n(계)", "등급", "", \
                    "", "", "", "", \
                    "빈도", "건수", "건당\n소요시간", "소요시간\n(연간)", ""]
            ws_data.append(header1)
            ws_data.append(header2)

            # 헤더 속성
            for col in range(1, DATA_COLS+1):
                header1 = ws_data.cell(row=HEADER_1, column=col)
                header1.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                header2 = ws_data.cell(row=HEADER_2, column=col)
                header2.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                # 직무 및 책무
                if col >= 1 and col <= 2:
                    header1.fill = PatternFill(fgColor="D0FA58", fill_type="solid")
                    header2.fill = PatternFill(fgColor="D0FA58", fill_type="solid")
                # 과업 관련 데이터 항목
                if col >= 3 and col <= 11:
                    header1.fill = PatternFill(fgColor="2ECCFA", fill_type="solid")
                    header2.fill = PatternFill(fgColor="81DAF5", fill_type="solid")
                # 활동 관련 데이터 항목
                if col >= 12 and col <= 20:
                    header1.fill = PatternFill(fgColor="2EFEC8", fill_type="solid")
                    header2.fill = PatternFill(fgColor="81F7D8", fill_type="solid")

            # 1줄씩 데이터 추가
            prev_job_nm = prev_duty_nm = prev_task_nm = None

            for i, r in data.iterrows():
                row_no = DATA_START_ROW+i
                # 직무명
                job_nm = ws_data.cell(row=row_no, column=1)
                if r['job_nm'] == prev_job_nm:  # 동일 데이터 반복 제거
                    job_nm.value = ""
                else:
                    job_nm.value = prev_job_nm = r['job_nm']
                    job_nm.border = BORDER_THIN_UP
                    job_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                # 책무
                duty_nm = ws_data.cell(row=row_no, column=2)
                if r['duty_nm'] == prev_duty_nm:  # 동일 데이터 반복 제거
                    duty_nm.value = ""
                else:
                    duty_nm.value = prev_duty_nm = r['duty_nm']
                    duty_nm.border = BORDER_THIN_UP
                    duty_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

                """ 과업 데이터 """
                # 과업, 과업 담당자, 과업 업무특성
                task_nm = ws_data.cell(row=row_no, column=3)
                task_prsn_chrg = ws_data.cell(row=row_no, column=4)
                work_attrbt = ws_data.cell(row=row_no, column=5)
                work_lv_imprt = ws_data.cell(row=row_no, column=6)
                work_lv_dfclt = ws_data.cell(row=row_no, column=7)
                work_lv_prfcn = ws_data.cell(row=row_no, column=8)
                work_lv_sum = ws_data.cell(row=row_no, column=9)
                work_grade = ws_data.cell(row=row_no, column=10)
                prfrm_tm_ann = ws_data.cell(row=row_no, column=11)
                if r['task_nm'] == prev_task_nm:  # 동일 데이터 반복 제거
                    task_nm.value = ""
                    task_prsn_chrg.value = work_attrbt = ""
                    work_lv_imprt.value = work_lv_dfclt = work_lv_prfcn = work_lv_sum = work_grade = ""
                    prfrm_tm_ann.value = ""
                else:
                    task_nm.value = prev_task_nm = r['task_nm']
                    task_prsn_chrg.value = r['task_prsn_chrg']
                    work_attrbt.value = r['work_attrbt']
                    work_lv_imprt.value = r['work_lv_imprt']
                    work_lv_dfclt.value = r['work_lv_dfclt']
                    work_lv_prfcn.value = r['work_lv_prfcn']
                    work_lv_sum.value = r['work_lv_sum']
                    work_grade.value = r['work_grade']
                    prfrm_tm_ann.value = r['prfrm_tm_ann']
                    task_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                    task_prsn_chrg.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                    work_attrbt.alignment = Alignment(horizontal="center", vertical="top")
                    work_lv_imprt.alignment = Alignment(horizontal="center", vertical="top")
                    work_lv_dfclt.alignment = Alignment(horizontal="center", vertical="top")
                    work_lv_prfcn.alignment = Alignment(horizontal="center", vertical="top")
                    work_lv_sum.alignment = Alignment(horizontal="center", vertical="top")
                    work_grade.alignment = Alignment(horizontal="center", vertical="top")
                    prfrm_tm_ann.alignment = Alignment(horizontal="center", vertical="top")
                    for c in range(TASK_START_COL, DATA_COLS+1):
                        task_related_cell = ws_data.cell(row=row_no, column=c)
                        task_related_cell.border = BORDER_THIN_UP

                """ 활동 데이터 """
                act_nm = ws_data.cell(row=row_no, column=12)
                rpt_nm = ws_data.cell(row=row_no, column=13)
                final_rpt_to = ws_data.cell(row=row_no, column=14)
                dept_rltd = ws_data.cell(row=row_no, column=15)
                act_prfrm_freq = ws_data.cell(row=row_no, column=16)
                act_prfrm_cnt = ws_data.cell(row=row_no, column=17)
                act_prfrm_tm_cs = ws_data.cell(row=row_no, column=18)
                act_prfrm_tm_ann = ws_data.cell(row=row_no, column=19)
                act_prsn_chrg = ws_data.cell(row=row_no, column=20)
                act_nm.value = r['act_nm']
                rpt_nm.value = r['rpt_nm']
                final_rpt_to.value = r['final_rpt_to']
                dept_rltd.value = r['dept_rltd']
                # act_prfrm_freq.value = r['act_prfrm_freq']
                act_prfrm_freq.value = r['cc_code_nm']
                act_prfrm_freq.value = r['cc_code']
                # act_prfrm_cnt.value = r['act_prfrm_cnt_ann']
                act_prfrm_cnt.value = r['act_prfrm_cnt']
                act_prfrm_tm_cs.value = r['act_prfrm_tm_cs']
                act_prfrm_tm_ann.value = r['act_prfrm_tm_ann']
                act_prsn_chrg.value = r['act_prsn_chrg']
                act_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                rpt_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                final_rpt_to.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                dept_rltd.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                act_prfrm_freq.alignment = Alignment(horizontal="center", vertical="top")
                act_prfrm_cnt.alignment = Alignment(horizontal="center", vertical="top")
                act_prfrm_tm_cs.alignment = Alignment(horizontal="center", vertical="top")
                act_prfrm_tm_ann.alignment = Alignment(horizontal="center", vertical="top")
                act_prsn_chrg.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

            """ 열 병합은 데이터 처리 후에 마지막에 진행 """
            # 열 병합
            ws_data.merge_cells(start_row=HEADER_1, start_column=6, end_row=HEADER_1, end_column=10)   # 업무 수준 및 등급
            # ws_data.merge_cells(start_row=HEADER_1, start_column=17, end_row=HEADER_1, end_column=19)  # 수행시간
            ws_data.merge_cells(start_row=HEADER_1, start_column=16, end_row=HEADER_1, end_column=17)  # 수행빈도(신규)
            ws_data.merge_cells(start_row=HEADER_1, start_column=18, end_row=HEADER_1, end_column=19)  # 수행시간
            # 행 병합: 헤더 1 & 2
            ws_data.merge_cells(start_row=HEADER_1, start_column=1, end_row=HEADER_2, end_column=1)   # 직무
            ws_data.merge_cells(start_row=HEADER_1, start_column=2, end_row=HEADER_2, end_column=2)   # 책무
            ws_data.merge_cells(start_row=HEADER_1, start_column=3, end_row=HEADER_2, end_column=3)   # 과업
            ws_data.merge_cells(start_row=HEADER_1, start_column=4, end_row=HEADER_2, end_column=4)   # 과업 담당자
            ws_data.merge_cells(start_row=HEADER_1, start_column=11, end_row=HEADER_2, end_column=11)   # 과업 수행시간(연간)
            ws_data.merge_cells(start_row=HEADER_1, start_column=12, end_row=HEADER_2, end_column=12)   # 활동
            ws_data.merge_cells(start_row=HEADER_1, start_column=13, end_row=HEADER_2, end_column=13)   # 수행 결과물
            ws_data.merge_cells(start_row=HEADER_1, start_column=14, end_row=HEADER_2, end_column=14)   # 최종 보고대상
            ws_data.merge_cells(start_row=HEADER_1, start_column=15, end_row=HEADER_2, end_column=15)   # 관련 부서
            # ws_data.merge_cells(start_row=HEADER_1, start_column=16, end_row=HEADER_2, end_column=16)   # 수행빈도
            ws_data.merge_cells(start_row=HEADER_1, start_column=20, end_row=HEADER_2, end_column=20)   # 활동 담당자

            # 행 고정
            ws_data.freeze_panes = "A3"

            """
            페이지 설정 및 인쇄 옵션
            """
            ws_data.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws_data.page_setup.orientation = ws_data.ORIENTATION_LANDSCAPE
            ws_data.page_setup.paperSize = ws_data.PAPERSIZE_A3
            ws_data.page_setup.fitToHeight = 0
            ws_data.page_setup.fitToWidth = 1
            ws_data.print_options.horizontalCentered = True
            # ws_data.print_options.verticalCentered = False
            ws_data.print_title_rows = '1:2'  # 헤더 1~2행을 반복해서 인쇄 타이틀로 설정

            # 년월일시분초
            def nowstr():
                now = dt.datetime.now()
                nowstr = now.strftime("%Y%m%d")
                return nowstr

            """
            엑셀 파일 저장
            """
            # 엑셀 파일을 BytesIO 객체에 저장
            excel_buffer = BytesIO()
            excel_file = f"직무현황표_{nowstr()}_{prd_selected}_{dept_nm}.xlsx"
            wb.save(excel_buffer)
            wb.close()
            excel_buffer.seek(0)

            encoded_filename = urllib.parse.quote(excel_file)

            # HttpResponse로 파일 전송
            response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

            return response

        # wb.save('직무기술서.xlsx')

        if action == "action2": # 직무 기술서 버튼 눌렀을 때

            wb = Workbook()

            # 테두리 적용
            BORDER_THIN_UP = Border(top=Side(style='thin'))
            BORDER_THIN_BOTTOM = Border(bottom=Side(style='thin'))
            BORDER_THIN_ALL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            BORDER_THIN_TOP_BOTTOM = Border(top=Side(style='thin'), bottom=Side(style='thin'))
            # border_medium = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
            # border_thick = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

            """
            첫번째 Sheet : 표지
            """
            ws = wb.active  # 현재 활성화된 sheet 가져옴
            ws.title = "부서"

            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 60

            # 회기
            subject1 = ws.cell(row=4, column=1)
            subject1.value = "■ 회기"
            subject1.font = Font(bold=True)
            prd = ws.cell(row=4, column=2)
            prd.value = prd_selected

            # 부서
            subject2 = ws.cell(row=6, column=1)
            subject2.value = "■ 부서"
            subject2.font = Font(bold=True)
            dept = ws.cell(row=6, column=2)
            # dept.value = dept_cd
            dept.value = dept_nm    # 부서명

            # 총 인원(PO)
            subject3 = ws.cell(row=8, column=1)
            subject3.value = "■ 총 인원(PO)"
            subject3.font = Font(bold=True)
            po = ws.cell(row=8, column=2)
            po.value = dept_po
            po.alignment = Alignment(horizontal="left", vertical="center")
            note = ws.cell(row=8, column=3)
            note.value = "(직책자포함)"
            note.font = Font(color="0000FF", size=9)

            # 성과책임
            subject4 = ws.cell(row=10, column=1)
            subject4.value = "■ 성과책임"
            subject4.font = Font(bold=True)
            RESP_START_ROW = 11
            for i, r in df_resp.iterrows():
                row_no = RESP_START_ROW+i
                ws.row_dimensions[row_no].height = 40
                resp_no = ws.cell(row=row_no, column=2)
                resp_no.value = "핵심목표 " + str(r['dept_resp_ordr'])
                resp_no.border = BORDER_THIN_ALL
                resp_no.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                resp_nm = ws.cell(row=row_no, column=3)
                resp_nm.value = r['dept_resp']
                resp_nm.border = BORDER_THIN_ALL
                resp_nm.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

            # 제목
            title = ws.cell(row=1, column=1)
            title.value = "직무기술서"
            title.font = Font(color="0000FF", size=25, bold=True)
            title.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            ws.row_dimensions[1].height = 50

            # Page Setup
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            ws.print_options.horizontalCentered = True

            """
            두번째 Sheet 부터 : 직무기술서 데이터 - 각 직무별로 시트 생성
            """
            SUBTITLE_ROW = 2
            DATA_START_ROW = 5

            # job_cd 중복되는 줄 지움
            df_job_task.drop_duplicates(subset=['job_cd'], keep='first', inplace=True)

            # df_dept_job의 순서를 df_job_task의 job_seq에 맞춤
            for i, r in df_job_task.iterrows():
                job_cd = r['job_cd']
                idx = df_dept_job[df_dept_job.job_cd == job_cd].index[0]
                df_dept_job.loc[idx, 'job_seq'] = r['job_seq']

            df_dept_job.sort_values(by=['job_seq'], inplace=True)

            df_dept_job.reset_index(inplace=True) # 인덱스 초기화

            for i, r in df_dept_job.iterrows():
                # print(i)
                ws_data = wb.create_sheet(str(i), i+1)

                # 컬럼 너비 지정
                ws_data.column_dimensions["A"].width = 20
                ws_data.column_dimensions["B"].width = 30
                ws_data.column_dimensions["C"].width = 40

                # 서식 번호
                tag = ws_data.cell(row=1, column=1)
                tag.value = "[별표 5]"
                tag.alignment = Alignment(horizontal="left", vertical="center")

                # 서브제목
                title = ws_data.cell(row=SUBTITLE_ROW, column=1)
                title.value = "직무기술서"
                title.font = Font(size=20, bold=True)
                title.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                ws_data.merge_cells(start_row=SUBTITLE_ROW, start_column=1, end_row=SUBTITLE_ROW, end_column=3)
                ws_data.row_dimensions[SUBTITLE_ROW].height = 30
                # 서브제목_Job Description
                title_e = ws_data.cell(row=SUBTITLE_ROW+1, column=1)
                title_e.value = "Job Description"
                title_e.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                ws_data.merge_cells(start_row=SUBTITLE_ROW+1, start_column=1, end_row=SUBTITLE_ROW+1, end_column=3)
                ws_data.row_dimensions[SUBTITLE_ROW+1].height = 15

                job_cd = r['job_cd']

                """ Section 1. 직무 기본 정보 """
                section1_row = DATA_START_ROW
                s1_title = ws_data.cell(row=section1_row, column=1)
                s1_title.value = "1. 직무 기본 정보"
                s1_title.font = Font(color="0000FF", size=13, bold=True)
                s1_title.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                ws_data.row_dimensions[section1_row].height = 20

                # 직무명, 직무 개요
                idx = df_job_info[df_job_info.job_cd == job_cd].index[0]
                job_nm, job_descrp = df_job_info.loc[idx][['job_nm', 'job_descrp']]
                # 직무 수행자
                idx = df_dept_job[df_dept_job.job_cd == job_cd].index[0]
                job_by = df_dept_job.loc[idx]['job_by']

                item1 = ws_data.cell(row=section1_row+1, column=1)
                item1.value = "직무명(Job)"
                item1.border = BORDER_THIN_ALL
                item1.alignment = Alignment(horizontal="center", vertical="center")
                item1.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")

                item1_d = ws_data.cell(row=section1_row+1, column=2)
                item1_d.value = job_nm
                item1_d.border = BORDER_THIN_ALL
                item1_d.alignment = Alignment(horizontal="center", vertical="center")
                ws_data.merge_cells(start_row=section1_row+1, start_column=2, end_row=section1_row+1, end_column=3)

                item2 = ws_data.cell(row=section1_row+2, column=1)
                item2.value = "조직"
                item2.border = BORDER_THIN_ALL
                item2.alignment = Alignment(horizontal="center", vertical="center")
                item2.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")

                item2_d = ws_data.cell(row=section1_row+2, column=2)
                item2_d.value = dept_nm
                item2_d.border = BORDER_THIN_ALL
                item2_d.alignment = Alignment(horizontal="center", vertical="center")
                ws_data.merge_cells(start_row=section1_row+2, start_column=2, end_row=section1_row+2, end_column=3)

                item3 = ws_data.cell(row=section1_row+3, column=1)
                item3.value = "직무 수행자"
                item3.border = BORDER_THIN_ALL
                item3.alignment = Alignment(horizontal="center", vertical="center")
                item3.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")

                item3_d = ws_data.cell(row=section1_row+3, column=2)
                item3_d.value = job_by
                item3_d.border = BORDER_THIN_ALL
                item3_d.alignment = Alignment(horizontal="center", vertical="center")
                ws_data.merge_cells(start_row=section1_row+3, start_column=2, end_row=section1_row+3, end_column=3)

                """ Section 2. 직무 개요 """
                section2_row = section1_row + 5
                s2_title = ws_data.cell(row=section2_row, column=1)
                s2_title.value = "2. 직무 개요"
                s2_title.font = Font(color="0000FF", size=13, bold=True)
                s2_title.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                ws_data.row_dimensions[section2_row].height = 20

                # 직무 개요
                job_desc = ws_data.cell(row=section2_row+1, column=1)
                job_desc.value = job_descrp
                job_desc.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                job_desc.border = BORDER_THIN_ALL
                ws_data.row_dimensions[section2_row+1].height = 40
                ws_data.merge_cells(start_row=section2_row+1, start_column=1, end_row=section2_row+1, end_column=3)

                """ Section 3. 직무 성과책임 """
                section3_row = section2_row + 3
                s3_title = ws_data.cell(row=section3_row, column=1)
                s3_title.value = "3. 직무 성과책임"
                s3_title.font = Font(color="0000FF", size=13, bold=True)
                s3_title.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                ws_data.row_dimensions[section3_row].height = 20

                # 직무 성과책임
                df_job_resp_job_by = df_job_resp[df_job_resp.job_cd == job_cd].sort_values(by='job_resp_ordr')
                cnt_job_resp = df_job_resp_job_by.shape[0]     # 직무 성과책임 개수

                if not df_job_resp_job_by.empty:
                    df_job_resp_job_by.reset_index(inplace=True)
                    for i, r in df_job_resp_job_by.iterrows():
                        row_no = section3_row+i+1
                        resp = ws_data.cell(row=row_no, column=1)
                        resp.value = str(r['job_resp_ordr']) + ". " + r['job_resp']
                        resp.alignment = Alignment(horizontal="left", vertical="center")
                        resp.border = BORDER_THIN_ALL
                        ws_data.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=3)
                else:
                    # NO DATA
                    no_data = ws_data.cell(row=section3_row+1, column=1)
                    no_data.value = "NO DATA"
                    no_data.alignment = Alignment(horizontal="center", vertical="center")
                    ws_data.merge_cells(start_row=section3_row+1, start_column=1, end_row=section3_row+1, end_column=3)
                    cnt_job_resp = 1


                """ Section 4. 직무 내용 """
                section4_row = section3_row + cnt_job_resp + 2  # 행 위치 가변적
                s4_title = ws_data.cell(row=section4_row, column=1)
                s4_title.value = "4. 직무 내용"
                s4_title.font = Font(color="0000FF", size=13, bold=True)
                s4_title.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                ws_data.row_dimensions[section4_row].height = 20

                # 직무-책무-과업-활동 Data
                df_job_data = data[data.job_cd == job_cd]
                cnt_job_data = df_job_data.shape[0]     # 활동 Data 개수

                # 해당 직무의 세부 데이터: 1줄씩 데이터 추가
                if not df_job_data.empty:
                    # 데이터 표 헤더(컬럼명)
                    header1 = ws_data.cell(row=section4_row+1, column=1)
                    header1.value = "책무(Duty)"
                    header1.border = BORDER_THIN_TOP_BOTTOM
                    header1.alignment = Alignment(horizontal="center", vertical="center")
                    header1.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")
                    header2 = ws_data.cell(row=section4_row+1, column=2)
                    header2.value = "과업(Task)"
                    header2.border = BORDER_THIN_TOP_BOTTOM
                    header2.alignment = Alignment(horizontal="center", vertical="center")
                    header2.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")
                    header3 = ws_data.cell(row=section4_row+1, column=3)
                    header3.value = "활동(Activity)"
                    header3.border = BORDER_THIN_TOP_BOTTOM
                    header3.alignment = Alignment(horizontal="center", vertical="center")
                    header3.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")

                    df_job_data.reset_index(inplace=True)
                    prev_duty_nm = prev_task_nm = None
                    for i, r in df_job_data.iterrows():
                        row_no = section4_row+2+i
                        # 책무
                        duty_nm = ws_data.cell(row=row_no, column=1)
                        if r['duty_nm'] == prev_duty_nm:  # 동일 데이터 반복 제거
                            duty_nm.value = ""
                        else:
                            duty_nm.value = prev_duty_nm = r['duty_nm']
                            duty_nm.border = BORDER_THIN_UP
                            duty_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                        # 과업
                        task_nm = ws_data.cell(row=row_no, column=2)
                        if r['task_nm'] == prev_task_nm:  # 동일 데이터 반복 제거
                            task_nm.value = ""
                        else:
                            task_nm.value = prev_task_nm = r['task_nm']
                            task_nm.border = BORDER_THIN_UP
                            task_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                        # 활동
                        act_nm = ws_data.cell(row=row_no, column=3)
                        act_nm.value = r['act_nm']
                        act_nm.border = BORDER_THIN_UP
                        act_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                    # 마지막 Line
                    ws_data.cell(row=section4_row+cnt_job_data+1, column=1).border = BORDER_THIN_BOTTOM
                    ws_data.cell(row=section4_row+cnt_job_data+1, column=2).border = BORDER_THIN_BOTTOM
                    ws_data.cell(row=section4_row+cnt_job_data+1, column=3).border = BORDER_THIN_BOTTOM
                else:
                    # NO DATA
                    no_data = ws_data.cell(row=section4_row+1, column=1)
                    no_data.value = "NO DATA"
                    no_data.alignment = Alignment(horizontal="center", vertical="center")
                    ws_data.merge_cells(start_row=section4_row+1, start_column=1, end_row=section4_row+1, end_column=3)

                """ 열 병합은 데이터 처리 후에 마지막에 진행 """

            # Page Setup
            ws_data.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws_data.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws_data.page_setup.paperSize = ws.PAPERSIZE_A4
            ws_data.page_setup.fitToHeight = 0
            ws_data.page_setup.fitToWidth = 1
            ws_data.print_options.horizontalCentered = True

            # 년월일시분초
            def nowstr():
                now = dt.datetime.now()
                nowstr = now.strftime("%Y%m%d")
                return nowstr

            # 엑셀 파일이 저장될 위치
            # file_root = "D:/PythonProject/job-mgr/output/"

            """
            엑셀 파일 저장
            """
            # 파일명은 full path로 지정
            # excel_file = "dept_job_" + nowstr() + ".xlsx"
            # excel_file_path = os.path.join(file_root, excel_file)
            # wb.save(excel_file_path)

            # download_folder = str(Path.home() / "Downloads")
            # excel_file = "직무기술서_" + nowstr() + ".xlsx"
            # excel_file_path = os.path.join(download_folder, excel_file)
            # wb.save(excel_file_path)
            # wb.close() # 엑셀 파일 닫기

            # 엑셀 파일을 BytesIO 객체에 저장
            excel_buffer = BytesIO()
            excel_file = f"직무기술서_{nowstr()}_{prd_selected}_{dept_nm}.xlsx"
            wb.save(excel_buffer)
            wb.close()
            excel_buffer.seek(0)

            encoded_filename = urllib.parse.quote(excel_file)

            # HttpResponse로 파일 전송
            response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # response['Content-Disposition'] = f'attachment; filename={excel_file}'

            # filename_header = f"filename*=UTF-8''{excel_file}"
            # response['Content-Disposition'] = f'attachment; {filename_header}'
            # response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{excel_file.encode("utf-8").decode("latin1")}'
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

            return response


        # DB 다시 접근해서 json 생성
        original_rows_show=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함

        data_list_show = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                    'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                    'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                    'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                    'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows_show]

        df_show = pd.DataFrame(data_list_show)

        # job_activity 접근
        original_rows_show_2=JobActivity.objects.filter(prd_cd=prd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list_show_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                    'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq, 'act_prfrm_cnt' : rows.act_prfrm_cnt,
                    'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                    'dept_rltd': rows.dept_rltd, 'final_rpt_to' : rows.final_rpt_to, 'rpt_nm': rows.rpt_nm,
                    'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq, 'act_seq': rows.act_seq} for rows in original_rows_show_2]

        df_show_2 = pd.DataFrame(data_list_show_2)

        df_show_3 = pd.merge(df_show, df_show_2).sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq']) # job_task와 job_activity merge

        # job_nm 찾기
        original_rows_show_2 = BsJob.objects.filter(prd_cd=prd_selected)
        data_list_show_4 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_show_2]
        df_show_4 = pd.DataFrame(data_list_show_4)

        df_show_3 = pd.merge(df_show_3, df_show_4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

        # df3.to_excel('df3.xlsx')
        df_json = df_show_3.to_json(orient='records')


        context = {
            'prd_selected' : prd_selected,
            'title' : "직무 상세정보",
            'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
            'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
            'dept_selected_key' : "former", # 부서 선택 후
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'data' : df_json, # 데이터프레임을 JSON 형식으로 변환하여 전달
        }

        if dept_login == "DD06":
            context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected) #마지막 회기의 부서 띄워주는게 좋을 듯
            context['dept_login_nm'] = dept_login_nm
            context['dept_cd_selected'] = dept_cd_selected
        else:
            context['dept_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
            context['dept_login_nm'] = dept_login_nm
            context['dept_cd_selected'] = dept_login

    return render(request, 'jobs/JB103.html', context)


def JB103_grid(request): # 직무정보 조회 초기화면

    last_prd_cd = BsPrd.objects.all().last().prd_cd # 가장 최근 회기. default로 띄워줌

    dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
    dept_login_nm = BsDept.objects.get(prd_cd=last_prd_cd, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

    # 회기, 부서 데이터에 해당하는 JobTask 값에 접근하여, dataframe 생성
    original_rows=JobTask.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함

    data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows]

    df1 = pd.DataFrame(data_list)

    # job_activity 접근
    original_rows_2=JobActivity.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함
    data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq,
                'act_prfrm_cnt': rows.act_prfrm_cnt, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                'rpt_nm': rows.rpt_nm, 'job_seq':rows.job_seq, 'duty_seq':rows.duty_seq, 'task_seq':rows.task_seq, 'act_seq':rows.act_seq } for rows in original_rows_2]

    df2 = pd.DataFrame(data_list_2)

    try:

        df3 = pd.merge(df1, df2)

        # dataframe의 index를 열로 만들어줌
        df3.reset_index(inplace=True)

        # # df3의 index 열을 복사하여, 새로운 열인 index_pos를 만들어줌. 이 값은 변하지 않는 값이며, grid에서 추가가 되면 999가 되는 값이다.
        # df3['index_pos'] = df3['index']

        # job_nm 열을 추가
        df3['job_nm'] = df3['job_cd'].apply(lambda x: BsJob.objects.get(prd_cd=last_prd_cd, job_cd=x).job_nm)

        # job_cd 열 삭제
        df3.drop('job_cd', axis=1, inplace=True)

        # job_seq, duty_seq, task_seq, act_seq 순으로 정렬
        df3 = df3.sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq'])

        # 데이터프레임을 JSON 형식으로 변환하여 전달
        df_json = df3.to_json(orient='records')

        context = {
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : last_prd_cd,
            'data' : df_json,
            'activate' : 'no', #버튼 컨트롤 off
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if dept_login == "DD06":
            context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd) #마지막 회기의 부서 띄워주는게 좋을 듯
            context['dept_login_nm'] = dept_login_nm
            context['dept_cd_selected'] = dept_login
        else:
            context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login)
            context['dept_login_nm'] = dept_login_nm
            context['dept_cd_selected'] = dept_login

        return render(request, 'jobs/JB103_grid.html', context)

    except pd.errors.MergeError as e:

        # messages.error(request, f'에러 발생: {"해당 회기 및 부서 내 데이터 없음"}')
        messages.error(request, '해당 회기에 로그인한 부서의 정보가 없습니다.')

        context = {
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : last_prd_cd,
            # 'error_message' : "해당 회기 및 부서에는 데이터가 없습니다.",
            'my_value' : "에러",
            'activate' : 'no', #버튼 컨트롤 off
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if dept_login == "DD06":
            context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd) #마지막 회기의 부서 띄워주는게 좋을 듯
            context['dept_login_nm'] = dept_login_nm
            context['dept_cd_selected'] = dept_login
        else:
            context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login)
            context['dept_login_nm'] = dept_login_nm
            context['dept_cd_selected'] = dept_login

        return render(request, 'jobs/JB103_grid.html', context)


def create_bs_prd(request): #BS101에서 submit했을 때 request에 대한 반응 - 회기 복사 및 삭제

    if request.method == 'POST':

        action = request.POST['action']

        if action == 'action1': # 회기 생성
            period_old = request.POST['prd_cd'] # 복사 대상 회기

            last_prd_cd = BsPrd.objects.last().prd_cd #BsPrd모델 중 마지막 줄의 prd_cd를 last_prd_cd 문자열 변수에 입력.
            last_year = last_prd_cd.strip()[0:4] #BsPrd 모델의 마지막 prd_cd 중 연도 정보
            last_char = last_prd_cd.strip()[-1] #char에는 마지막 prd_cd 중 연도 정보 중 글자만 입력한다.

            if last_year == str(now.year): # 마지막 회기의 연도가 현재 년도와 같다면, 즉 현재 년도의 회기가 이미 있으면
                period_new = str(now.year) + chr(ord(last_char)+1) # 현재 년도의 회기에다가 마지막 회기의 마지막 글자에 +1을 해서 새로운 회기를 만든다.

            else: # 마지막 회기의 연도가 현재 년도와 다르다면, 즉 현재 년도의 회기가 없다면
                period_new = str(now.year) + "A" # 새로운 회기를 만들 때 현재 년도와 그 뒤에 A를 붙여 해당 년도의 첫 번째 회기를 생성한다.

            messages = copy_period_data(period_old, period_new)

            last_bs_prd = BsPrd.objects.last() # 마지막 회기의 정보를 last_bs_prd에 저장. 새로운 회기를 뜻할 것이다.

            if last_bs_prd:
                # 회기를 만들어줬으니 그 회기에 대한 정보들을 입력해준다.
                last_bs_prd.year = int(now.year) # 복사한 회기의 year는 지금 현재 년도
                last_bs_prd.turn = BsPrd.objects.filter(year=now.year).count() + 1 # 새로 만든 회기의 turn은 그 해당하는 연도에 있는 회기 데이터 개수에 1을 더한 같다.
                last_bs_prd.prd_done_yn = 'N'
                last_bs_prd.prd_str_dt = dt.datetime.today()
                last_bs_prd.job_srv_str_dt = None
                last_bs_prd.job_srv_end_dt = None
                last_bs_prd.prd_end_dt = None
                last_bs_prd.save()

            # 새로운 회기에 대한 BsDept 테이블에 접근하여 필요한 데이터들을 업데이트한다.
            update_target = BsDept.objects.filter(prd_cd_id=last_bs_prd.prd_cd)
            update_target.update(job_details_submit_yn="N", job_details_submit_dttm=None)

            # 새로운 회기에 대한 BsJobDept 테이블에 접근하여 필요한 데이터들을 업데이트한다.
            update_target_2 = BsJobDept.objects.filter(prd_cd_id=last_bs_prd.prd_cd)
            update_target_2.update(create_dttm=dt.datetime.today(), alter_dttm=dt.datetime.today())

            context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
                'title' : '회기 관리',
                'user_name' : 'inu',
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'today_date' : str(dt.datetime.today()).split()[0],
                'messages': messages,  # 메시지 리스트를 context에 추가
                'prd' : BsPrd.objects.all(),
                'tab' : 'tab1',
                'modified' : 'y' # 회기 복사나 삭제 작업을 했다는 키값(메시지용)
            }

        elif action == 'action2': # 회기 삭제

            period_del = request.POST['prd_cd'] # 삭제 대상 회기

            messages = delete_period_data(period_del)

            context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
                'title' : '회기 관리',
                'user_name' : 'inu',
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'today_date' : str(dt.datetime.today()).split()[0],
                'messages': messages,  # 메시지 리스트를 context에 추가
                'prd' : BsPrd.objects.all(),
                'tab' : 'tab1',
                'modified' : 'y' # 회기 복사나 삭제 작업을 했다는 키값(메시지용)
            }

    return render(request, 'jobs/BS101.html', context) #장고가 context를 meshup해서 html template으로 보내줌
    #return redirect('BS101') #BS101로 돌아감.


def JB108(request): # 직무현황 제출 초기화면

    last_prd_cd = BsPrd.objects.all().last().prd_cd # 가장 최근 회기. default로 띄워줌
    prd_done_yn = BsPrd.objects.get(prd_cd=last_prd_cd).prd_done_yn

    dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
    submit_yn = BsDept.objects.get(prd_cd=last_prd_cd, dept_cd=dept_login).job_details_submit_yn

    if prd_done_yn == 'N':
            if submit_yn == 'N':
                confirm_text = "직무현황을 제출하지 않았습니다."
            else:
                confirm_text = "직무현황을 제출한 상태입니다."
    else:
        confirm_text = "마감된 회기입니다."

    context = {
        'prd_list' : BsPrd.objects.all(),
        'title' : '직무 현황제출', # 제목
        'prd_selected' : last_prd_cd,
        'prd_done_yn' : prd_done_yn,
        'modified' : "n",
        'confirm_text' : confirm_text,
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
    }

    if dept_login == "DD06":
        context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd)
        context['dept_selected'] = dept_login
    else:
        context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login)
        context['dept_selected'] = dept_login

    return render(request, 'jobs/JB108.html', context)


def JB109(request): # 업무량 분석화면 - 회기선택화면

    last_prd_cd = BsPrd.objects.all().last().prd_cd # 가장 최근 회기. default로 띄워줌

    context = {
        'prd_list' : BsPrd.objects.all(),
        'title' : '업무량 분석', # 제목
        'prd_cd_selected' : last_prd_cd,
        'dept_selected_key' : 'former',
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
    }

    return render(request, 'jobs/JB109.html', context)


def JB110(request): # 부서 업무량 분석화면 초기 화면 + 회기 선택 화면

    # 초기화면

    user_name = request.user.username # 아이디(ID___)
    # print('user name', user_name)
    dept_login = get_dept_code(user_name) # 로그인한 부서의 부서코드. 회기 수정해야 함. 아이디 테이블에 없기 때문.
    dept_login_nm = BsDept.objects.get(prd_cd=BsPrd.objects.all().last().prd_cd, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

    context = {
        'title' : '부서 업무량 분석', # 제목
        'prd_list' : BsPrd.objects.all(),
        'user_name' : user_name,
        # 'activate' : 'no', #버튼 컨트롤 off
        'prd_cd_selected' : BsPrd.objects.all().last().prd_cd,
        'dept_login_nm' : dept_login_nm,
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        'status' : 'tab_before'
    }

    # 회기 선택 후 화면

    if request.method == 'POST':
        prd_cd_selected = request.POST["prd_cd_selected"]

        try:
            dept_login_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

            context = {
                'title' : '부서 업무량 분석', # 제목
                'prd_list' : BsPrd.objects.all(),
                'user_name' : user_name,
                'activate' : 'no', #버튼 컨트롤 off
                'prd_cd_selected' : prd_cd_selected,
                'dept_login_nm' : dept_login_nm,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'status' : 'tab_before'
            }

        except ObjectDoesNotExist as e: # 회기 선택을 했는데 로그인한 부서가 그 회기에 없는 경우

            # 탭 선택 비활성화 시켜줌.

            messages.error(request, '해당 회기에 로그인한 부서가 없습니다.')

            context = {
                'title' : '부서 업무량 분석', # 제목
                'prd_list' : BsPrd.objects.all(),
                'user_name' : user_name,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'activate' : 'no', #버튼 컨트롤 off
                'prd_cd_selected' : prd_cd_selected,
                'status' : 'tab_before',
                'tab_activate' : 'no' # 탭 선택 비활성화
            }

    return render(request, 'jobs/JB110.html', context)


def JB200(request): # 업무량 분석 기초 자료 화면. 이 화면은 경영기획팀만 사용한다. 회기 선택화면이다.

    last_prd_cd = BsPrd.objects.all().last().prd_cd # 가장 최근 회기. default로 띄워줌

    dept_list = BsDept.objects.filter(prd_cd=last_prd_cd) # 마지막 회기의 부서 리스트

    dept_selected = get_dept_code(request.user.username) # 로그인한 부서의 부서코드(경영기획팀)

    # std_wrk_tm = BsStdWrkTm.objects.get(prd_cd=last_prd_cd).std_wrk_tm
    std_wrk_tm = float(BsStdWrkTm.objects.get(prd_cd=last_prd_cd).std_wrk_tm)

    # job_task 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
    original_rows=JobTask.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
    data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                   'work_lv_imprt': rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                     'work_grade': rows.work_grade_id, 'prfrm_tm_ann': rows.prfrm_tm_ann, 'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq } for rows in original_rows]
    df1 = pd.DataFrame(data_list)

    # df1에 prfrm_tm_ann_cal 열을 추가해준다. 초기값은 null이다.
    df1['prfrm_tm_ann_cal'] = None

    # df1에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
    # work_grade가 null일 경우 prfrm_tm_ann_cal열도 null값으로 한다.
    # df1의 각 행에 대하여 수행해준다. for문을 활용한다.
    for i in range(len(df1)):
        if df1.loc[i, 'work_grade'] == 'G1':
            if df1['prfrm_tm_ann'][i] == None:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G1').workload_wt)
        elif df1.loc[i, 'work_grade'] == 'G2':
            if df1['prfrm_tm_ann'][i] == None:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G2').workload_wt)
        elif df1.loc[i, 'work_grade'] == 'G3':
            if df1['prfrm_tm_ann'][i] == None:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G3').workload_wt)
        elif df1.loc[i, 'work_grade'] == 'G4':
            if df1['prfrm_tm_ann'][i] == None:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G4').workload_wt)
        elif df1.loc[i, 'work_grade'] == 'G5':
            if df1['prfrm_tm_ann'][i] == None:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G5').workload_wt)
        else:
            df1.loc[i, 'prfrm_tm_ann_cal'] = None

    # df1['prfrm_tm_ann_cal']의 자료형을 float으로 변경
    df1['prfrm_tm_ann_cal'] = df1['prfrm_tm_ann_cal'].astype(float)
    # print(df1)

    # job_task_adj 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
    original_rows_2=JobTaskAdj.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
    data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                     'work_lv_imprt_adj': rows.work_lv_imprt, 'work_lv_dfclt_adj': rows.work_lv_dfclt, 'work_lv_prfcn_adj': rows.work_lv_prfcn, 'work_lv_sum_adj': rows.work_lv_sum,
                       'work_grade_adj': rows.work_grade_id, 'prfrm_tm_ann_adj': rows.prfrm_tm_ann, 'adj_yn': rows.adj_yn } for rows in original_rows_2]
    df2 = pd.DataFrame(data_list_2)

    # df2에 prfrm_tm_ann_cal_adj열을 추가해준다. 초기값은 null이다.
    df2['prfrm_tm_ann_cal_adj'] = None

    # df2에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal_adj)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade_adj가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
    # work_grade_adj가 null일 경우 prfrm_tm_ann_cal_adj열도 null값으로 한다.
    # df2의 각 행에 대하여 수행해준다. for문을 활용한다.
    for i in range(len(df2)):
        if df2['work_grade_adj'][i] == 'G1':
            if df2['prfrm_tm_ann_adj'][i] == None:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G1').workload_wt)
        elif df2['work_grade_adj'][i] == 'G2':
            if df2['prfrm_tm_ann_adj'][i] == None:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G2').workload_wt)
        elif df2['work_grade_adj'][i] == 'G3':
            if df2['prfrm_tm_ann_adj'][i] == None:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G3').workload_wt)
        elif df2['work_grade_adj'][i] == 'G4':
            if df2['prfrm_tm_ann_adj'][i] == None:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G4').workload_wt)
        elif df2['work_grade_adj'][i] == 'G5':
            if df2['prfrm_tm_ann_adj'][i] == None:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=last_prd_cd, work_grade='G5').workload_wt)
        else:
            df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None

    # df2['prfrm_tm_ann_cal_adj']의 자료형을 float으로 변경
    df2['prfrm_tm_ann_cal_adj'] = df2['prfrm_tm_ann_cal_adj'].astype(float)

    try:
        df3 = pd.merge(df1, df2).sort_values(['job_seq', 'duty_seq', 'task_seq']) # job_task와 job_activity merge, 순서는 job_seq, duty_seq, task_seq, act_seq 순

        # job_nm 찾기
        original_rows_3 = BsJob.objects.filter(prd_cd=last_prd_cd)
        data_list_3 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_3]
        df4 = pd.DataFrame(data_list_3)

        df3 = pd.merge(df3, df4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

        # df3.to_excel('df3.xlsx')
        df_json = df3.to_json(orient='records')

        # BsWorkGrade에서 데이터를 가져와서, json으로 만들어서 넘겨준다.
        original_rows_4 = BsWorkGrade.objects.filter(prd_cd=last_prd_cd)
        data_list_4 = [{'work_grade': rows.work_grade, 'work_lv_min': rows.work_lv_min, 'work_lv_max': rows.work_lv_max, 'workload_wt': rows.workload_wt} for rows in original_rows_4]
        df5 = pd.DataFrame(data_list_4)
        df_json_2 = df5.to_json(orient='records')

        # 환산 업무량 합 구하기 #####################
        work_grade_list = ['G1', 'G2', 'G3', 'G4', 'G5'] # 업무등급 리스트. 이 리스트는 변하지 않는다.
        mbr_of_dept = BsMbr.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_selected) # 해당 부서의 부서원 목록
        mbr_of_dept = mbr_of_dept.exclude(ttl_nm='팀장')

        analysis = pd.DataFrame({'work_grade':work_grade_list}) # 분석 결과를 담을 데이터프레임
        prfrm_tm_ann = [] # 각 업무등급별 총 업무량을 담을 리스트
        prfrm_tm_ann_adj = [] # 각 업무등급별 총 조정업무량을 담을 리스트

        for grade in work_grade_list: # 각 업무 등급에 대해서

            df6 = df3[df3['work_grade'] == grade] # JobTask 혹은 JobTaskAdj을 이용해 만든 df1에서 그 grade에 해당하는 task만 추출하여 df2 만들기
            df7 = df3[df3['work_grade_adj'] == grade]

            prfrm_tm_ann.append(float(df6['prfrm_tm_ann'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann의 합을 구한다.
            prfrm_tm_ann_adj.append(float(df7['prfrm_tm_ann_adj'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann_adj의 합을 구한다.

        analysis['prfrm_tm_ann'] = prfrm_tm_ann # 분석 결과 데이터프레임에 prfrm_tm_ann 열 추가
        analysis['prfrm_tm_ann_adj'] = prfrm_tm_ann_adj # 분석 결과 데이터프레임에 prfrm_tm_ann_adj 열 추가

        workload = BsWorkGrade.objects.filter(prd_cd=last_prd_cd) # 업무량 가중치 object
        workload_wt = [float(n.workload_wt) for n in workload ] # 업무량 가중치 list(float)

        analysis['workload_wt'] = workload_wt # 업무량 가중치

        analysis['prfrm_tm_ann_cal'] = analysis['prfrm_tm_ann'] * analysis['workload_wt']
        analysis['prfrm_tm_ann_cal_adj'] = analysis['prfrm_tm_ann_adj'] * analysis['workload_wt']

        sum_prfrm_tm_ann_cal = round(analysis['prfrm_tm_ann_cal'].sum(), 1)
        sum_prfrm_tm_ann_cal_adj = round(analysis['prfrm_tm_ann_cal_adj'].sum(), 1)

        analysis['po_right'] = round(analysis['prfrm_tm_ann_cal']/std_wrk_tm, 1) # 적정 인력 산정
        analysis['po_right_adj'] = round(analysis['prfrm_tm_ann_cal_adj']/std_wrk_tm, 1) # 조정 적정 인력 산정

        po_right = round(analysis['po_right'].sum(), 1)
        po_right_adj = round(analysis['po_right_adj'].sum(), 1)

        # df3의 prfrm_tm_ann_cal 열의 합을 구한다. None일 경우 0으로 처리한다. 소숫점 1자리까지 표시한다.
        # sum_prfrm_tm_ann_cal = round(df3['prfrm_tm_ann_cal'].sum(), 1)
        # sum_prfrm_tm_ann_cal_adj = round(df3['prfrm_tm_ann_cal_adj'].sum(), 1)

        # sum_prfrm_tm_ann_cal을 BsStdWrkTm 테이블에서 가져온 std_wrk_tm으로 나누어서, 소숫점 1자리까지 표시한다. 이 값이 적정 인원이다.
        # po_right = round(float(sum_prfrm_tm_ann_cal) / float(std_wrk_tm), 1)
        # po_right_adj = round(float(sum_prfrm_tm_ann_cal_adj) / float(std_wrk_tm), 1)

        context = {
            'prd_list' : BsPrd.objects.all(),
            'title' : '분석 기초자료', # 제목
            'prd_selected' : last_prd_cd,
            'prd_done' : BsPrd.objects.get(prd_cd=last_prd_cd).prd_done_yn,
            'dept_list' : dept_list,
            'dept_cd_selected' : dept_selected,
            'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
            'data' : df_json,
            'standard': df_json_2,
            'std_wrk_tm': std_wrk_tm,
            'sum_prfrm_tm_ann_cal': sum_prfrm_tm_ann_cal,
            'po_right': po_right,
            'sum_prfrm_tm_ann_cal_adj': sum_prfrm_tm_ann_cal_adj,
            'po_right_adj': po_right_adj,
        }
    except pd.errors.MergeError as e:

        messages.error(request, '해당 회기에 분석 정보 기초자료 데이터가 없습니다.')

        # 버튼 컨트롤 다 막아야 함
        context.update({'data' : 'null'})

        context = {
            'prd_list' : BsPrd.objects.all(),
            'title' : '분석 기초자료', # 제목
            'prd_selected' : last_prd_cd,
            'prd_done' : BsPrd.objects.get(prd_cd=last_prd_cd).prd_done_yn,
            'dept_list' : dept_list,
            'dept_cd_selected' : dept_selected,
            'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
            'data' : 'null',
        }

    if request.method == 'POST': # 회기 변경 시

        prd_selected = request.POST["prd_selected"]
        dept_list = BsDept.objects.filter(prd_cd=prd_selected) # 변경한 회기의 부서 리스트

        dept_selected = get_dept_code(request.user.username) # 로그인한 부서의 부서코드(경영기획팀)
        std_wrk_tm = BsStdWrkTm.objects.get(prd_cd=prd_selected).std_wrk_tm

        # job_task 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
        original_rows=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                    'work_lv_imprt': rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                        'work_grade': rows.work_grade_id, 'prfrm_tm_ann': rows.prfrm_tm_ann,
                          'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq } for rows in original_rows]
        df1 = pd.DataFrame(data_list)

        # df1에 prfrm_tm_ann_cal 열을 추가해준다. 환산 업무량이지. 초기값은 null이다.
        df1['prfrm_tm_ann_cal'] = None

        # df1에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
        # work_grade가 null일 경우 prfrm_tm_ann_cal열도 null값으로 한다.
        # df1의 각 행에 대하여 수행해준다. for문을 활용한다.
        for i in range(len(df1)):
            if df1['work_grade'][i] == 'G1':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G1').workload_wt)
            elif df1['work_grade'][i] == 'G2':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G2').workload_wt)
            elif df1['work_grade'][i] == 'G3':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G3').workload_wt)
            elif df1['work_grade'][i] == 'G4':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G4').workload_wt)
            elif df1['work_grade'][i] == 'G5':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G5').workload_wt)
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None

        # df1['prfrm_tm_ann_cal']의 자료형을 float으로 변경
        df1['prfrm_tm_ann_cal'] = df1['prfrm_tm_ann_cal'].astype(float)

        # job_task_adj 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
        original_rows_2=JobTaskAdj.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                        'work_lv_imprt_adj': rows.work_lv_imprt, 'work_lv_dfclt_adj': rows.work_lv_dfclt, 'work_lv_prfcn_adj': rows.work_lv_prfcn, 'work_lv_sum_adj': rows.work_lv_sum,
                        'work_grade_adj': rows.work_grade_id, 'prfrm_tm_ann_adj': rows.prfrm_tm_ann, 'adj_yn': rows.adj_yn } for rows in original_rows_2]
        df2 = pd.DataFrame(data_list_2)

        # df2에 prfrm_tm_ann_cal_adj열을 추가해준다. 환산 업무량의 adj인 것이다. 초기값은 null이다.
        df2['prfrm_tm_ann_cal_adj'] = None

        # df2에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal_adj)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade_adj가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
        # work_grade_adj가 null일 경우 prfrm_tm_ann_cal_adj열도 null값으로 한다.
        # df2의 각 행에 대하여 수행해준다. for문을 활용한다.
        for i in range(len(df2)):
            if df2['work_grade_adj'][i] == 'G1':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G1').workload_wt)
            elif df2['work_grade_adj'][i] == 'G2':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G2').workload_wt)
            elif df2['work_grade_adj'][i] == 'G3':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G3').workload_wt)
            elif df2['work_grade_adj'][i] == 'G4':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G4').workload_wt)
            elif df2['work_grade_adj'][i] == 'G5':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G5').workload_wt)
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None


        # df2['prfrm_tm_ann_cal_adj']의 자료형을 float으로 변경
        df2['prfrm_tm_ann_cal_adj'] = df2['prfrm_tm_ann_cal_adj'].astype(float)

        try:
            df3 = pd.merge(df1, df2).sort_values(['job_seq', 'duty_seq', 'task_seq']) # job_task와 job_activity merge, 순서는 job_seq, duty_seq, task_seq, act_seq 순

            # job_nm 찾기
            original_rows_3 = BsJob.objects.filter(prd_cd=prd_selected)
            data_list_3 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_3]
            df4 = pd.DataFrame(data_list_3)

            df3 = pd.merge(df3, df4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

            # df3.to_excel('df3.xlsx')
            df_json = df3.to_json(orient='records')

            # BsWorkGrade에서 데이터를 가져와서, json으로 만들어서 넘겨준다.
            original_rows_4 = BsWorkGrade.objects.filter(prd_cd=prd_selected)
            data_list_4 = [{'work_grade': rows.work_grade, 'work_lv_min': rows.work_lv_min, 'work_lv_max': rows.work_lv_max, 'workload_wt': rows.workload_wt} for rows in original_rows_4]
            df5 = pd.DataFrame(data_list_4)
            df_json_2 = df5.to_json(orient='records')

            # 환산 업무량 합 구하기 #####################
            work_grade_list = ['G1', 'G2', 'G3', 'G4', 'G5'] # 업무등급 리스트. 이 리스트는 변하지 않는다.
            mbr_of_dept = BsMbr.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 해당 부서의 부서원 목록
            mbr_of_dept = mbr_of_dept.exclude(ttl_nm='팀장')

            analysis = pd.DataFrame({'work_grade':work_grade_list}) # 분석 결과를 담을 데이터프레임
            prfrm_tm_ann = [] # 각 업무등급별 총 업무량을 담을 리스트
            prfrm_tm_ann_adj = [] # 각 업무등급별 총 조정업무량을 담을 리스트

            for grade in work_grade_list: # 각 업무 등급에 대해서

                df6 = df3[df3['work_grade'] == grade] # JobTask 혹은 JobTaskAdj을 이용해 만든 df1에서 그 grade에 해당하는 task만 추출하여 df2 만들기
                df7 = df3[df3['work_grade_adj'] == grade]

                prfrm_tm_ann.append(float(df6['prfrm_tm_ann'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann의 합을 구한다.
                prfrm_tm_ann_adj.append(float(df7['prfrm_tm_ann_adj'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann_adj의 합을 구한다.

            analysis['prfrm_tm_ann'] = prfrm_tm_ann # 분석 결과 데이터프레임에 prfrm_tm_ann 열 추가
            analysis['prfrm_tm_ann_adj'] = prfrm_tm_ann_adj # 분석 결과 데이터프레임에 prfrm_tm_ann_adj 열 추가

            workload = BsWorkGrade.objects.filter(prd_cd=prd_selected) # 업무량 가중치 object
            workload_wt = [float(n.workload_wt) for n in workload ] # 업무량 가중치 list(float)

            analysis['workload_wt'] = workload_wt # 업무량 가중치

            analysis['prfrm_tm_ann_cal'] = analysis['prfrm_tm_ann'] * analysis['workload_wt']
            analysis['prfrm_tm_ann_cal_adj'] = analysis['prfrm_tm_ann_adj'] * analysis['workload_wt']

            sum_prfrm_tm_ann_cal = round(analysis['prfrm_tm_ann_cal'].sum(), 1)
            sum_prfrm_tm_ann_cal_adj = round(analysis['prfrm_tm_ann_cal_adj'].sum(), 1)

            analysis['po_right'] = round(analysis['prfrm_tm_ann_cal']/std_wrk_tm, 1) # 적정 인력 산정
            analysis['po_right_adj'] = round(analysis['prfrm_tm_ann_cal_adj']/std_wrk_tm, 1) # 조정 적정 인력 산정

            po_right = round(analysis['po_right'].sum(), 1)
            po_right_adj = round(analysis['po_right_adj'].sum(), 1)

            context = {
                'prd_list' : BsPrd.objects.all(),
                'title' : '분석 기초자료', # 제목
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'dept_list' : dept_list,
                'dept_cd_selected' : dept_selected,
                'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
                'data' : df_json,
                'standard': df_json_2,
                'std_wrk_tm': std_wrk_tm,
                'sum_prfrm_tm_ann_cal': sum_prfrm_tm_ann_cal,
                'po_right': po_right,
                'sum_prfrm_tm_ann_cal_adj': sum_prfrm_tm_ann_cal_adj,
                'po_right_adj': po_right_adj,
            }
        except pd.errors.MergeError as e:

            messages.error(request, '해당 회기에 분석 정보 기초자료 데이터가 없습니다.')

            context = {
                'prd_list' : BsPrd.objects.all(),
                'title' : '분석 기초자료', # 제목
                'prd_selected' : last_prd_cd,
                'prd_done' : BsPrd.objects.get(prd_cd=last_prd_cd).prd_done_yn,
                'dept_list' : dept_list,
                'dept_cd_selected' : dept_selected,
                'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
                'data' : 'null',
            }

    return render(request, 'jobs/JB200.html', context)


def JB300(request): # 직무 분류 체계 초기화면

    prd_cd_selected = BsPrd.objects.all().last().prd_cd

    dept_domain_list = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected).values_list('dept_domain', flat=True).distinct()

    # dept_domain_list의 마지막 값을 dept_domain_selected로 지정한다.
    dept_domain_selected = dept_domain_list.last()

    context = {
        'title' : '직무 분류 체계', # 제목
        'prd' : BsPrd.objects.all(),
        'prd_cd_selected' : prd_cd_selected,
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        'dept_domain_list' : dept_domain_list,
        'dept_domain_selected' : dept_domain_selected,
    }

    if request.method == 'POST': # 회기 변경 조직 도메인을 선택할 수 있는 화면을 띄워준다.

        prd_cd_selected = request.POST["prd_cd_selected"]


        # BsDeptGrpDomain 테이블에서 해당 회기의 object 중 dept_domain을 중복을 제거해서 리스트로 만든다.
        dept_domain_list = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected).values_list('dept_domain', flat=True).distinct()

        dept_domain_selected = dept_domain_list.last()

        context = {
            'title' : '직무 분류 체계', # 제목
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : prd_cd_selected,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'dept_domain_list' : dept_domain_list,
            'dept_domain_selected' : dept_domain_selected,
        }

    return render(request, 'jobs/JB300.html', context)


def BS200_1(request): #BS200에서 탭 선택 후 display

    if request.method == 'POST':

        # span을 어느 것을 선택하느냐에 따라 다르게 연산을 수행할 것임. 일단 그 span이 뭔지 알아낼 것임.
        span_name = request.POST.get('span_name', '')
        prd_cd_selected = request.POST['prd_cd'] # 선택한 회기를 input으로 받아옴.
        result_object = BsPrd.objects.get(prd_cd=prd_cd_selected) #선택한 회기에 대한 BsPrd object를 호출

        if span_name == 'span1': # 직무 조사 기간 탭을 선택했을 경우
            # 두 가지 케이스로 나눔. 회기가 확정되지 않은 경우만 수정할 수 있음. 회기가 확정된 경우는 그 회기에 관련한 정보를 출력
            if result_object.prd_done_yn == "Y": #회기 확정된 경우
                context = { #context를 넘겨줌.
                'title' : '직무 조사',
                'prd' : BsPrd.objects.all(),
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'prd_cd_selected' : prd_cd_selected,
                'job_srv_str_dt' : str(result_object.job_srv_str_dt).split()[0],
                'job_srv_end_dt' : str(result_object.job_srv_end_dt).split()[0],
                'result_object' : result_object,
                'modify' : "n", #수정할 수 없도록 키값 부여
                'tab' : 'tab1'
            }
            else: #회기 확정되지 않은 경우
                if result_object.job_srv_end_dt == None:
                    job_srv_str_dt = str(dt.datetime.today()).split()[0]
                    job_srv_end_dt = ""
                else:
                    job_srv_str_dt = str(result_object.job_srv_str_dt).split()[0]
                    job_srv_end_dt = str(result_object.job_srv_end_dt).split()[0]

                context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
                    'title' : '직무 조사',
                    'prd' : BsPrd.objects.all(),
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'job_srv_str_dt' : job_srv_str_dt,
                    'job_srv_end_dt' : job_srv_end_dt,
                    'prd_cd_selected' : prd_cd_selected,
                    'result_object' : result_object,
                    'modify' : "y", #수정할 수 있도록 키값 부여
                    'confirm_text_1' : "해당 회기의 직무 조사 시작일과 종료일을 설정하여 저장해 주세요.",
                    'confirm_text_2' : "직무 조사 기간 변경 시 종료일을 재선택한 후 저장해 주세요.",
                    'prd_str_dt' : str(result_object.prd_str_dt).split()[0], # 회기가 확정되지 않은 경우 그 회기의 회기 시작일을 받아온다(hidden). 직무 조사 시작일 input과 비교할 것이다.
                    'tab' : 'tab1'
                }

        elif span_name == 'span2': # 직무 조사 기간 탭을 선택했을 경우

            original_rows=BsDept.objects.filter(pk=prd_cd_selected)
            data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'job_details_submit_yn': rows.job_details_submit_yn} for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            df1 = df1.replace('Y', '제출')
            df1 = df1.replace('N', '미제출')
            df1['job_details_submit_yn'].fillna('미제출', inplace=True)

            dept_all_cnt = len(df1) # 전체 부서 수
            dept_y_cnt = len(df1.loc[df1['job_details_submit_yn'] == "제출"]) # 제출 부서 수
            dept_n_cnt = len(df1.loc[df1['job_details_submit_yn'] == "미제출"]) # 미제출 부서 수

            context={
                'title' : '직무 조사',
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'tab' : 'tab2',
                'dept_list' : df1,
                'dept_all_cnt' : dept_all_cnt,
                'dept_y_cnt' : dept_y_cnt,
                'dept_n_cnt' : dept_n_cnt,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

        if 'action' in request.POST: # 직무 제출현황 요약에서 조회 버튼 눌렀을 때

            action = request.POST['action']

            original_rows=BsDept.objects.filter(pk=prd_cd_selected)
            data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'job_details_submit_yn': rows.job_details_submit_yn} for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            df1 = df1.replace('Y', '제출')
            df1 = df1.replace('N', '미제출')
            df1['job_details_submit_yn'].fillna('미제출', inplace=True)

            dept_all_cnt = len(df1) # 전체 부서 수
            dept_y_cnt = len(df1.loc[df1['job_details_submit_yn'] == "제출"]) # 제출 부서 수
            dept_n_cnt = len(df1.loc[df1['job_details_submit_yn'] == "미제출"]) # 미제출 부서 수

            if action == 'action1': # 제출 부서 조회

                df2 = df1[df1['job_details_submit_yn'] == '제출']

            elif action == 'action2': # 미제출 부서 조회

                df2 = df1[df1['job_details_submit_yn'] == '미제출']

            context={
                'title' : '직무 조사',
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'tab' : 'tab2',
                'dept_list' : df1,
                'dept_all_cnt' : dept_all_cnt,
                'dept_y_cnt' : dept_y_cnt,
                'dept_n_cnt' : dept_n_cnt,
                'dept_list_selected' : df2,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

    return render(request, 'jobs/BS200.html', context) #장고가 context를 meshup해서 html template으로 보내줌


def BS200_2(request): ## BS200 직무 조사일탭 직무조사 시작일, 종료일 지정(저장 버튼 눌렀을 때와 취소 버튼 눌렀을 때를 구분)

    if request.method == 'POST':

        action = request.POST['action']

        if action == 'action1': #저장 버튼 눌렀을 때

            prd_cd_selected = request.POST['prd_cd_selected']
            bs_prd_update = BsPrd.objects.get(prd_cd = prd_cd_selected) # BsPrd의 마지막것을 가져옴
            bs_prd_update.job_srv_str_dt = request.POST['job_srv_str_dt'] #html에서 받은 시작일을 BsPrd 마지막 데이터의 시작일로 지정
            bs_prd_update.job_srv_end_dt = request.POST['job_srv_end_dt'] #html에서 받은 종료일을 BsPrd 마지막 데이터의 시작일로 지정
            bs_prd_update.save()

            context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
                'title' : '직무 조사',
                'prd' : BsPrd.objects.all(),
                'job_srv_str_dt' : str(bs_prd_update.job_srv_str_dt).split()[0],
                'job_srv_end_dt' : str(bs_prd_update.job_srv_end_dt).split()[0],
                'prd_cd_selected' : prd_cd_selected,
                'result_object' : bs_prd_update,
                'modify' : "y", #수정할 수 있도록 키값 부여
                'confirm_text_1' : "해당 회기의 직무 조사 기간이 저장되었습니다.",
                'modified' : "y", #저장되었을 경우 키값 부여. 인포 메시지 띄우기 위함.
                'prd_str_dt' : str(bs_prd_update.prd_str_dt).split()[0],
                'tab' : 'tab1',
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

        elif action == 'action2': #취소 버튼 눌렀을 때

            prd_cd_selected = request.POST['prd_cd_selected'] # 선택한 회기를 input으로 받아옴.
            result_object = BsPrd.objects.get(prd_cd=prd_cd_selected) #선택한 회기에 대한 BsPrd object를 호출

            print(result_object.job_srv_str_dt)

            # 두 가지 케이스로 나눔. 회기가 확정되지 않은 경우만 수정할 수 있음. 회기가 확정된 경우는 그 회기에 관련한 정보를 출력
            if result_object.prd_done_yn == "Y": #회기 확정된 경우 - 수정할 수 없음

                context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
                    'title' : '직무 조사',
                    'prd' : BsPrd.objects.all(),
                    'job_srv_str_dt' : str(result_object.job_srv_str_dt).split()[0],
                    'job_srv_end_dt' : str(result_object.job_srv_end_dt).split()[0],
                    'prd_cd_selected' : prd_cd_selected,
                    'result_object' : result_object,
                    'modify' : "n", #수정할 수 없도록 키값 부여
                    'tab' : 'tab1',
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                }

            elif result_object.prd_done_yn == "N": #회기 확정되지 않은 경우 - 수정 가능

                if result_object.job_srv_end_dt == None:
                    job_srv_str_dt = str(dt.datetime.today()).split()[0]
                    job_srv_end_dt = ""
                else:
                    job_srv_str_dt = str(result_object.job_srv_str_dt).split()[0]
                    job_srv_end_dt = str(result_object.job_srv_end_dt).split()[0]

                context = { #context를 넘겨줌. context는 어떤 type도 가능(?)
                    'title' : '직무 조사',
                    'prd' : BsPrd.objects.all(),
                    'job_srv_str_dt' : job_srv_str_dt,
                    'job_srv_end_dt' : job_srv_end_dt,
                    'prd_cd_selected' : prd_cd_selected,
                    'result_object' : result_object,
                    'modify' : "y", #수정할 수 있도록 키값 부여
                    'confirm_text_1' : "해당 회기의 직무 조사 시작일과 종료일을 설정하여 저장해 주세요.",
                    'confirm_text_2' : "직무 조사 기간 변경 시 종료일을 재선택한 후 저장해 주세요.",
                    'tab' : 'tab1',
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                }

        return render(request, 'jobs/BS200.html', context)


def BS300_1(request): #BS300 회기 및 탭 선택 후 display

    if request.method == 'POST':

        # 선택한 회기를 input으로 받아옴.
        prd_cd_selected = request.POST['prd_cd_selected']

        # span을 어느 것을 선택하느냐에 따라 다르게 연산을 수행할 것임. 일단 그 span이 뭔지 알아낼 것임.
        span_name = request.POST.get('span_name', '')

        if span_name == 'span1': # 부서 관리 탭을 선택했을 경우
            # span1에 대한 처리
            original_rows=BsDept.objects.filter(pk=prd_cd_selected)
            data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'dept_to': rows.dept_to, 'dept_po': rows.dept_po} for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                'dept_list' : df1,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'activate' : 'yes', #버튼 컨트롤 on
                'tab' : "tab1",
                'save_activate' : 'no', # 저장 버튼 activate
                'del_activate' : 'no', # 삭제 버튼 activate
                'title' : '조직 정보' # 제목
            }

        elif span_name == 'span2': # 조직 그룹 탭을 선택했을 경우
            # 아직 아무일도 일어나지 않아야 함.
            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                'tab' : "tab2",
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'title' : '조직 정보' # 제목
            }

            #################### 여기서부터 새로 바꿔준 부분 ####################
            # BsDeptGrp를 가져와서 json으로 변경시켜준다. JB103이랑 비슷하게 한다.
            original_rows = BsDeptGrp.objects.filter(prd_cd_id=prd_cd_selected)
            data_list = [{'dept_domain': rows.dept_domain_id, 'dept_grp_nm' : rows.dept_grp_nm_id, 'dept_cd' : rows.dept_cd_id, 'dept_seq': rows.dept_seq} for rows in original_rows]
            # df1 = pd.DataFrame(data_list).sort_values(by=['domain_seq', 'grp_seq', 'dept_seq']).reset_index(drop=True)
            df1 = pd.DataFrame(data_list)
            # dataframe을 json으로 변경시켜서 html로 보낸다. JB103이랑 비슷하게 한다.
            df_json = df1.to_json(orient='records')
            # json데이터 엑셀 저장
            # df1.to_excel('dept_grp.xlsx', index=False)

            original_rows2 = BsDeptGrpDomain.objects.filter(prd_cd_id=prd_cd_selected)
            data_list2 = [{'dept_domain': rows.dept_domain, 'dept_grp_nm' : rows.dept_grp_nm, 'domain_seq': rows.domain_seq,
                           'grp_seq': rows.grp_seq } for rows in original_rows2]
            df2 = pd.DataFrame(data_list2).sort_values(by=['domain_seq', 'grp_seq']).reset_index(drop=True)
            df_json2 = df2.to_json(orient='records')
            # 엑셀 저장
            # df2.to_excel('dept_grp_domain.xlsx', index=False)

            original_rows3 = BsDept.objects.filter(prd_cd_id=prd_cd_selected)
            data_list3 = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm} for rows in original_rows3]
            df3 = pd.DataFrame(data_list3).sort_values(by=['dept_nm']).reset_index(drop=True)
            df_json3 = df3.to_json(orient='records')
            # 엑셀 저장
            # df3.to_excel('dept.xlsx', index=False)

            context.update({'data_bs_dept_grp' : df_json})
            context.update({'data_bs_dept_grp_domain' : df_json2})
            context.update({'data_bs_dept' : df_json3})
            #################### 새로 바꿔준 부분 끝 ####################

        elif span_name == 'span3': # 직위 관리 탭을 선택했을 경우

            try:
                # BsPosGrade 와 BsPosList를 merge하여, pos_ordr에 따라 pos_nm을 정렬시킨 dataframe df1을 만들고 넘겨준다.
                original_rows1 = BsPosGrade.objects.filter(prd_cd_id=prd_cd_selected)
                original_rows2 = BsPosList.objects.filter(prd_cd_id=prd_cd_selected)

                data_list1 = [{'pos_nm' : rows.pos_nm, 'work_grade' : rows.work_grade_id} for rows in original_rows1]
                df1 = pd.DataFrame(data_list1)

                data_list2 = [{'pos_nm' : rows.pos_nm, 'pos_ordr' : rows.pos_ordr} for rows in original_rows2]
                df2 = pd.DataFrame(data_list2)

                df1 = df1.merge(df2, how='inner', on='pos_nm').sort_values(by=['pos_ordr']).reset_index(drop=True)

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                    'pos_grade_list' : df1,
                    'pos_list' : BsPosGrade.objects.filter(prd_cd_id = prd_cd_selected),
                    'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab3",
                    'work_grade_list' : BsWorkGrade.objects.filter(prd_cd_id=prd_cd_selected),
                    'title' : '조직 정보' # 제목
                    }

            except KeyError as e:

                messages.error(request, f'데이터가 존재하지 않습니다.')

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    # 'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab3",
                    'title' : '조직 정보' # 제목
                    }

        elif span_name == 'span4': # 직책 관리 탭을 선택했을 경우
            # BsTtlList 테이블을 ttl_ordr로 arrange한 dataframe df1을 만들고 넘겨준다.

            try:
                original_rows = BsTtlList.objects.filter(prd_cd_id=prd_cd_selected)
                data_list = [{'ttl_nm' : rows.ttl_nm, 'ttl_ordr' : rows.ttl_ordr} for rows in original_rows]
                df1 = pd.DataFrame(data_list).sort_values(by=['ttl_ordr']).reset_index(drop=True)

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                    'ttl_list' : df1,
                    'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab4",
                    'title' : '조직 정보' # 제목
                    }

            except KeyError as e:
                messages.error(request, f'데이터가 존재하지 않습니다.')

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    # 'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab4",
                    'title' : '조직 정보' # 제목
                    }

    return render(request, 'jobs/BS300.html', context)


def BS300_2(request): #BS300 편집(행추가, 저장, 취소, 삭제 후)

    if request.method == 'POST':

        action = request.POST['action']
        # 선택한 회기와 탭 정보를 가져옴
        prd_cd_selected = request.POST['prd_cd_selected']
        tab = request.POST['tab_selected']

        # 조직 그룹 탭 관련한 것은 tab == 'tab2'로 해서 필요함.
        if tab == 'tab3': # 직위 관리 탭

            # UI의 직위명과 직위의 등급을 가져와서 list를 만든다. 화면의 순서대로 만든다.
            pos_nm = request.POST.getlist('pos_nm')
            work_grade = request.POST.getlist('work_grade')
            pos_ordr_list = list(range(1, len(pos_nm)+1))
            result = zip(pos_nm, work_grade, pos_ordr_list)


            # 저장 버튼을 눌렀을 때
            if action == 'action1':

                BsPosGrade.objects.filter(prd_cd_id=prd_cd_selected).delete()
                BsPosList.objects.filter(prd_cd_id=prd_cd_selected).delete()

                for i, j, k in result:
                    BsPosList.objects.create(prd_cd_id = prd_cd_selected, pos_nm=i, pos_ordr=k)
                    BsPosGrade.objects.create(prd_cd_id = prd_cd_selected, pos_nm=i, work_grade_id=j)

                # df1이라는 dataframe을 만들기 위한 것임. df1은 선택한 회기의 work_grade, pos_nm 값으로 구성된 df이다.
                original_rows1 = BsPosGrade.objects.filter(prd_cd_id=prd_cd_selected)
                original_rows2 = BsPosList.objects.filter(prd_cd_id=prd_cd_selected)

                data_list1 = [{'pos_nm' : rows.pos_nm, 'work_grade' : rows.work_grade_id} for rows in original_rows1]
                df1 = pd.DataFrame(data_list1)

                data_list2 = [{'pos_nm' : rows.pos_nm, 'pos_ordr' : rows.pos_ordr} for rows in original_rows2]
                df2 = pd.DataFrame(data_list2)

                df1 = df1.merge(df2, how='inner', on='pos_nm').sort_values(by=['pos_ordr']).reset_index(drop=True)

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                    'pos_grade_list' : df1,
                    'pos_list' : BsPosGrade.objects.filter(prd_cd_id = prd_cd_selected),
                    'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab3",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'work_grade_list' : BsWorkGrade.objects.filter(prd_cd_id=prd_cd_selected),
                    'title' : '조직 정보' # 제목
                    }

            elif action == 'action2':

                original_rows1 = BsPosGrade.objects.filter(prd_cd_id=prd_cd_selected)
                original_rows2 = BsPosList.objects.filter(prd_cd_id=prd_cd_selected)

                data_list1 = [{'pos_nm' : rows.pos_nm, 'work_grade' : rows.work_grade_id} for rows in original_rows1]
                df1 = pd.DataFrame(data_list1)

                data_list2 = [{'pos_nm' : rows.pos_nm, 'pos_ordr' : rows.pos_ordr} for rows in original_rows2]
                df2 = pd.DataFrame(data_list2)

                df1 = df1.merge(df2, how='inner', on='pos_nm').sort_values(by=['pos_ordr']).reset_index(drop=True)

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                    'pos_grade_list' : df1,
                    'pos_list' : BsPosGrade.objects.filter(prd_cd_id = prd_cd_selected),
                    'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab3",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'work_grade_list' : BsWorkGrade.objects.filter(prd_cd_id=prd_cd_selected),
                    'title' : '조직 정보' # 제목
                }

        elif tab == 'tab4': # 직책 관리 탭

            if action == 'action1': # 저장 버튼 누를 때

                # UI의 직책 list를 만든다. 화면의 순서대로 만든다.
                ttl_nm = request.POST.getlist('ttl_nm')
                ttl_ordr_list = list(range(1, len(ttl_nm)+1))

                # 해당 회기의 부서 이름 목록을 가져와 리스트를 만든다.
                original_rows2 = BsDept.objects.filter(prd_cd_id=prd_cd_selected)
                dept_cd_list = [row.dept_cd for row in original_rows2]

                result = zip(ttl_nm, ttl_ordr_list)

                BsTtlList.objects.filter(prd_cd_id=prd_cd_selected).delete()
                BsTtlCnt.objects.filter(prd_cd_id=prd_cd_selected).delete()

                for i, j in result:
                    BsTtlList.objects.create(prd_cd_id = prd_cd_selected, ttl_nm=i, ttl_ordr=j)

                for k in dept_cd_list:
                    for a in ttl_nm:
                        BsTtlCnt.objects.create(prd_cd_id = prd_cd_selected, dept_cd_id=k, ttl_nm=a, ttl_cnt=0)

                original_rows = BsTtlList.objects.filter(prd_cd_id=prd_cd_selected)
                data_list = [{'ttl_nm' : rows.ttl_nm, 'ttl_ordr' : rows.ttl_ordr} for rows in original_rows]
                df1 = pd.DataFrame(data_list).sort_values(by=['ttl_ordr']).reset_index(drop=True)

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                    'ttl_list' : df1,
                    'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab4",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'title' : '조직 정보' # 제목
                    }

            elif action == 'action2': # 취소 버튼 누를 때

                original_rows = BsTtlList.objects.filter(prd_cd_id=prd_cd_selected)
                data_list = [{'ttl_nm' : rows.ttl_nm, 'ttl_ordr' : rows.ttl_ordr} for rows in original_rows]
                df1 = pd.DataFrame(data_list).sort_values(by=['ttl_ordr']).reset_index(drop=True)

                context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                    'ttl_list' : df1,
                    'activate' : 'yes', #버튼 컨트롤 on
                    'tab' : "tab4",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'title' : '조직 정보' # 제목
                    }

    return render(request, 'jobs/BS300.html', context)


def BS300_3(request): # BS300 부서관리 탭에서 부서 선택했을 때, 그에 해당하는 부서 정보들과 control 버튼들을 띄워준다.

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd_selected']
        tab = request.POST['tab_selected']
        dept_selected = request.POST['dept_selected']

        # 부서 목록 만들어주기
        original_rows=BsDept.objects.filter(pk=prd_cd_selected)
        data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'dept_to': rows.dept_to, 'dept_po': rows.dept_po} for rows in original_rows]
        df1 = pd.DataFrame(data_list)

        # 선택한 부서 띄워주기 위해 선언
        dept_selected_row = BsDept.objects.filter(pk=prd_cd_selected, dept_cd=dept_selected)

        # 선택한 회기와 부서에 대한 직책별 to dataframe 만들기
        original_rows2 = BsTtlList.objects.filter(prd_cd_id=prd_cd_selected)
        data_list2 = [{'ttl_nm' : rows.ttl_nm, 'ttl_ordr' : rows.ttl_ordr} for rows in original_rows2]
        df2 = pd.DataFrame(data_list2).sort_values(by=['ttl_ordr']).reset_index(drop=True) # 해당 회기의 직책 DataFrame(ttl_ordr로 order_by된)

        original_rows3 = BsTtlCnt.objects.filter(prd_cd_id=prd_cd_selected, dept_cd=dept_selected)
        data_list3 = [{'ttl_nm' : rows.ttl_nm, 'ttl_cnt' : rows.ttl_cnt} for rows in original_rows3]
        df3 = pd.DataFrame(data_list3) # 해당 회기 및 해당 부서의 직책별 cnt DataFrame

        df2 = df2.merge(df3, how='left', on='ttl_nm').sort_values(by=['ttl_ordr']).reset_index(drop=True)

        context = {
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : prd_cd_selected,
            'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'dept_list' : df1,
            'activate' : 'yes', #버튼 컨트롤 on
            'tab' : tab,
            'dept_selected' : dept_selected,
            'dept_selected_row' : dept_selected_row,
            'dept_ttl' : BsTtlCnt.objects.filter(pk=prd_cd_selected, dept_cd=dept_selected),
            'to_info' : df2,
            'save_activate' : 'save_activate', # 저장 버튼 activate
            'del_activate' : 'del_activate', # 삭제 버튼 activate
            'title' : '조직 정보' # 제목
        }

    return render(request, 'jobs/BS300.html', context)


def BS300_4(request): # 부서 관리 탭에서 부서를 선택한 후 편집하고 저장하거나 추가, 삭제 버튼 후 동작 수행__추가 시 추가할 수 있는 화면을 띄워준다.

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd_selected']
        tab = request.POST['tab_selected']
        dept_selected = request.POST['dept_selected']

        dept_ttl_nm = request.POST.getlist('dept_ttl_nm') # 부서의 직책 이름
        dept_ttl_cnt = request.POST.getlist('dept_ttl_cnt') # 부서의 직책별 TO
        action = request.POST['action']

        if action == 'action1': # 저장 눌렀을 때

            dept_selected_nm = request.POST['dept_selected_nm']

            # 부서 이름 업데이트
            BsDept.objects.filter(pk=prd_cd_selected, dept_cd=dept_selected).update(dept_nm=dept_selected_nm)

            # 직책 이름에 따라 TO 업데이트
            for nm, cnt in zip(dept_ttl_nm ,dept_ttl_cnt):
                BsTtlCnt.objects.filter(prd_cd_id=prd_cd_selected, dept_cd=dept_selected, ttl_nm=nm).update(ttl_cnt=cnt)

            # 부서 목록 만들어주기
            original_rows=BsDept.objects.filter(pk=prd_cd_selected)
            data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'dept_to': rows.dept_to, 'dept_po': rows.dept_po} for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            # 선택한 부서 띄워주기 위해 선언
            dept_selected_row = BsDept.objects.filter(pk=prd_cd_selected, dept_cd=dept_selected)

            # 선택한 회기와 부서에 대한 직책별 to dataframe 만들기
            original_rows2 = BsTtlList.objects.filter(prd_cd_id=prd_cd_selected)
            data_list2 = [{'ttl_nm' : rows.ttl_nm, 'ttl_ordr' : rows.ttl_ordr} for rows in original_rows2]
            df2 = pd.DataFrame(data_list2).sort_values(by=['ttl_ordr']).reset_index(drop=True) # 해당 회기의 직책 DataFrame(ttl_ordr로 order_by된)

            original_rows3 = BsTtlCnt.objects.filter(prd_cd_id=prd_cd_selected, dept_cd=dept_selected)
            data_list3 = [{'ttl_nm' : rows.ttl_nm, 'ttl_cnt' : rows.ttl_cnt} for rows in original_rows3]
            df3 = pd.DataFrame(data_list3) # 해당 회기 및 해당 부서의 직책별 cnt DataFrame

            df2 = df2.merge(df3, how='left', on='ttl_nm').sort_values(by=['ttl_ordr']).reset_index(drop=True)

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'dept_list' : df1,
                'activate' : 'yes', #버튼 컨트롤 on
                'tab' : tab,
                'dept_selected' : dept_selected,
                'dept_selected_row' : dept_selected_row,
                'dept_ttl' : BsTtlCnt.objects.filter(pk=prd_cd_selected, dept_cd=dept_selected),
                'to_info' : df2,
                'save_activate' : 'save_activate', # 저장 버튼 activate
                'del_activate' : 'del_activate', # 삭제 버튼 activate
                'title' : '조직 정보' # 제목
            }

        if action == 'action2': # 추가 눌렀을 때 - 추가할 수 있는 입력칸을 띄워주도록 한다.

            original_rows=BsDept.objects.filter(pk=prd_cd_selected)
            data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'dept_to': rows.dept_to, 'dept_po': rows.dept_po} for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'dept_list' : df1,
                'activate' : 'yes', # 버튼 컨트롤 on
                'tab' : tab,
                'new_dept' : 'yes', # 새 부서 추가할 수 있는 버튼 contrl key
                'ttl_list' : BsTtlList.objects.filter(prd_cd_id=prd_cd_selected).order_by('ttl_ordr'),
                'del_activate' : 'no', # 삭제 버튼 activate
                'title' : '조직 정보' # 제목
            }

        if action == 'action3': # 삭제 눌렀을 때 - 해당 부서 코드의 부서를 삭제한다.

            BsDept.objects.filter(pk=prd_cd_selected, dept_cd=dept_selected).delete()

            # span1에 대한 처리
            original_rows=BsDept.objects.filter(pk=prd_cd_selected)
            data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'dept_to': rows.dept_to, 'dept_po': rows.dept_po} for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'dept_list' : df1,
                'activate' : 'yes', #버튼 컨트롤 on
                'tab' : "tab1",
                'save_activate' : 'no', # 저장 버튼 activate
                'del_activate' : 'no', # 삭제 버튼 activate
                'title' : '조직 정보' # 제목
            }

        if action == 'action4': # 취소 눌렀을 때 - 해당 부서 코드의 부서를 삭제한다.

            # span1에 대한 처리
            original_rows=BsDept.objects.filter(pk=prd_cd_selected)
            data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'dept_to': rows.dept_to, 'dept_po': rows.dept_po} for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'dept_list' : df1,
                'activate' : 'yes', #버튼 컨트롤 on
                'tab' : "tab1",
                'save_activate' : 'no', # 저장 버튼 activate
                'del_activate' : 'no', # 삭제 버튼 activate
                'title' : '조직 정보' # 제목
            }

    return render(request, 'jobs/BS300.html', context)


def BS300_5(request): # 부서 관리 탭에서 부서 추가

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd_selected']
        tab = request.POST['tab_selected']
        action = request.POST['action']

        if action == 'action1' : # 저장 버튼 누르면

            new_dept_nm = request.POST['new_dept_nm'] # 새로운 부서 이름
            dept_ttl_nm = request.POST.getlist('dept_ttl_nm') # 새로운 부서의 직책 이름
            dept_ttl_cnt = request.POST.getlist('dept_ttl_cnt') # 새로운 부서의 직책별 TO

            # 새로운 object의 cc_code
            # 직무 유형에 따른 새로운 직무코드의 설정 값 정의(고유일 때와 공통일 때 구분)
            code_prefix = "DD"

            # 새로운 코드 생성 로직
            char = BsDept.objects.filter(prd_cd_id=prd_cd_selected).order_by('dept_cd').last().dept_cd
            new_dept_cd = code_prefix + f"{(int(char[2:5]) + 1):02d}"  # char의 마지막 두 숫자를 기준으로 새 코드 생성

            BsDept.objects.create(prd_cd_id=prd_cd_selected, dept_cd=new_dept_cd, dept_nm=new_dept_nm, dept_po=0, dept_to=0, job_details_submit_yn='N')

            for i, j in zip(dept_ttl_nm, dept_ttl_cnt):
                BsTtlCnt.objects.create(prd_cd_id=prd_cd_selected, dept_cd_id=new_dept_cd, ttl_nm=i, ttl_cnt=j)

        original_rows=BsDept.objects.filter(pk=prd_cd_selected)
        data_list = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm, 'dept_to': rows.dept_to, 'dept_po': rows.dept_po} for rows in original_rows]
        df1 = pd.DataFrame(data_list)

        context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'dept_list' : df1,
                'activate' : 'yes', # 버튼 컨트롤 on
                'tab' : tab,
                'save_activate' : 'no', # 저장 버튼 activate
                'del_activate' : 'no', # 삭제 버튼 activate
                'title' : '조직 정보' # 제목
            }

    return render(request, 'jobs/BS300.html', context)

######################################### 조직 그룹 탭 #########################################
def BS300_6(request): # 조직 그룹 탭에서 저장, 취소

    if 'action' in request.POST:

        # 선택한 회기를 input으로 받아옴.
        prd_cd_selected = request.POST['prd_cd_selected']
        tab = request.POST['tab_selected']
        action = request.POST['action']

        # 공통 context
        context = {
                    'prd' : BsPrd.objects.all(),
                    'prd_cd_selected' : prd_cd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                    'tab' : "tab2",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'title' : '조직 정보' # 제목
        }

        if action == 'action1': # 저장 버튼 눌렀을 때

            # json 데이터들을 받아와서 dataframe으로 변환
            json_data = request.POST.get('jsonData')
            data = json.loads(json_data) # JSON 문자열을 Python 객체로 변환
            df_domain_grp = pd.DataFrame(data) # Pandas DataFrame으로 변환. 도메인 테이블과 비교

            deptGrpData = request.POST.get('deptGrpData')
            data2 = json.loads(deptGrpData)
            df_dept = pd.DataFrame(data2) # 그룹별 부서 테이블과 비교
            # print(df_dept)

            # 먼저 BsDeptGrpDomain 테이블을 df_domain_grp와 비교하여 업데이트. JB103_3과 비슷한 로직
            # 기존 데이터(해당 회기의 데이터)를 가져와서 비교한다.
            original_rows1 = BsDeptGrpDomain.objects.filter(prd_cd_id=prd_cd_selected)
            data_list1 = [{'dept_domain': rows.dept_domain, 'dept_grp_nm' : rows.dept_grp_nm, 'domain_seq': rows.domain_seq,
                           'grp_seq': rows.grp_seq } for rows in original_rows1]
            df_1 = pd.DataFrame(data_list1).sort_values(by=['domain_seq', 'grp_seq']).reset_index(drop=True)

            # 비교 하는 부분 - merge 기능을 이용해 추가된 행, 삭제된 행을 추출할 것이다. 수정은 삭제 후 추가로 볼 것이다.
            # df_left_domain : df1(DB)에 있고 df_domain_grp(UI)에 없는 것. 즉, 삭제된 것
            df_left_domain = pd.merge(df_1, df_domain_grp, how='outer', indicator=True).query('_merge == "left_only"').drop(columns=['_merge']).reset_index(drop=True)
            # df_right_domain : df_task(UI)에 있고 df1(DB)에 없는 것. 즉, 추가된 것
            df_right_domain = pd.merge(df_1, df_domain_grp, how='outer', indicator=True).query('_merge == "right_only"').drop(columns=['_merge']).reset_index(drop=True)

            # df_left_domain부터 다룬다. df_left_domain에 있는 것들은 삭제해야 하는 것들이다. DB에 접근해서 지워줌
            for i in range(0, len(df_left_domain)):
                BsDeptGrpDomain.objects.filter(prd_cd_id=prd_cd_selected, dept_domain=df_left_domain.iloc[i, 0], dept_grp_nm=df_left_domain.iloc[i, 1]).delete()

            # df_right_domain 다룬다. df_right_domain에 있는 것들은 추가해야 하는 것들이다. DB에 접근해서 추가해줌
            for i in range(0, len(df_right_domain)):
                BsDeptGrpDomain.objects.create(prd_cd_id=prd_cd_selected, dept_domain=df_right_domain.iloc[i, 0], dept_grp_nm=df_right_domain.iloc[i, 1],
                                                domain_seq=df_right_domain.iloc[i, 2], grp_seq=df_right_domain.iloc[i, 3])

            # 그 다음은 BsDeptGrp 테이블을 df_dept와 비교하여 업데이트. JB103_3과 비슷한 로직
            # 기존 데이터(해당 회기의 데이터)를 가져와서 비교한다.
            original_rows2 = BsDeptGrp.objects.filter(prd_cd_id=prd_cd_selected)
            data_list2 = [{'dept_domain': rows.dept_domain_id, 'dept_grp_nm' : rows.dept_grp_nm_id, 'dept_cd':rows.dept_cd_id, 'dept_seq': rows.dept_seq } for rows in original_rows2]
            df_2 = pd.DataFrame(data_list2)

            # df_2가 비어있지 않은 경우(BsDeptGrp에 데이터 있는 경우)에만 비교를 진행한다.
            if not df_2.empty:

                if not df_dept.empty: # df_dept가 비어있지 않은 경우, 즉 화면에서 어느 한 부서라도 체크를 한 경우에는 비교를 진행한다.

                    # 비교 하는 부분 - merge 기능을 이용해 추가된 행, 삭제된 행을 추출할 것이다. 수정은 삭제 후 추가로 볼 것이다.
                    # df_left_dept : df2(DB)에 있고 df_dept(UI)에 없는 것. 즉, 삭제된 것
                    df_left_dept = pd.merge(df_2, df_dept, how='outer', indicator=True).query('_merge == "left_only"').drop(columns=['_merge']).reset_index(drop=True)
                    # df_right_dept : df_dept(UI)에 있고 df2(DB)에 없는 것. 즉, 추가된 것
                    df_right_dept = pd.merge(df_2, df_dept, how='outer', indicator=True).query('_merge == "right_only"').drop(columns=['_merge']).reset_index(drop=True)
                    print(df_right_dept)

                    # df_left_dept부터 다룬다. df_left_dept에 있는 것들은 삭제해야 하는 것들이다. DB에 접근해서 지워줌
                    for i in range(0, len(df_left_dept)):
                        BsDeptGrp.objects.filter(prd_cd_id=prd_cd_selected, dept_domain_id=df_left_dept.iloc[i, 0], dept_grp_nm_id=df_left_dept.iloc[i, 1], dept_cd_id=df_left_dept.iloc[i, 2]).delete()

                    # df_right_dept 다룬다. df_right_dept에 있는 것들은 추가해야 하는 것들이다. DB에 접근해서 추가해줌
                    for i in range(0, len(df_right_dept)):
                        BsDeptGrp.objects.create(prd_cd_id=prd_cd_selected, dept_domain_id=df_right_dept.iloc[i, 0], dept_grp_nm_id=df_right_dept.iloc[i, 1], dept_cd_id=df_right_dept.iloc[i, 2],
                                                dept_seq=df_right_dept.iloc[i, 3])

                else: # df_dept가 비어있는 경우, 즉 화면에서 어느 한 부서도 체크를 하지 않은 경우에는 BsDeptGrp 테이블을 다 없애준다.
                    BsDeptGrp.objects.filter(prd_cd_id=prd_cd_selected).delete()

            # df2가 비어있는 경우, 즉 BsDeptGrp가 비어있는 경우는 df_dept 이용해서 BsDeptGrp 테이블에 추가해준다.
            else:
                for i in range(0, len(df_dept)):
                    BsDeptGrp.objects.create(prd_cd_id=prd_cd_selected, dept_domain_id=df_dept.iloc[i, 0], dept_grp_nm_id=df_dept.iloc[i, 1], dept_cd_id=df_dept.iloc[i, 2],
                                            dept_seq=df_dept.iloc[i, 3])

            # 편집을 다 끝내고 나면 다시 화면을 띄워준다.
            # BsDeptGrp를 가져와서 json으로 변경시켜준다. JB103이랑 비슷하게 한다.
            original_rows = BsDeptGrp.objects.filter(prd_cd_id=prd_cd_selected)
            data_list = [{'dept_domain': rows.dept_domain_id, 'dept_grp_nm' : rows.dept_grp_nm_id, 'dept_cd' : rows.dept_cd_id, 'dept_seq': rows.dept_seq} for rows in original_rows]
            # df1 = pd.DataFrame(data_list).sort_values(by=['domain_seq', 'grp_seq', 'dept_seq']).reset_index(drop=True)
            df1 = pd.DataFrame(data_list)
            # dataframe을 json으로 변경시켜서 html로 보낸다. JB103이랑 비슷하게 한다.
            df_json = df1.to_json(orient='records')
            # json데이터 엑셀 저장
            # df1.to_excel('dept_grp.xlsx', index=False)

            original_rows2 = BsDeptGrpDomain.objects.filter(prd_cd_id=prd_cd_selected)
            data_list2 = [{'dept_domain': rows.dept_domain, 'dept_grp_nm' : rows.dept_grp_nm, 'domain_seq': rows.domain_seq,
                        'grp_seq': rows.grp_seq } for rows in original_rows2]
            df2 = pd.DataFrame(data_list2).sort_values(by=['domain_seq', 'grp_seq']).reset_index(drop=True)
            df_json2 = df2.to_json(orient='records')
            # 엑셀 저장
            # df2.to_excel('dept_grp_domain.xlsx', index=False)

            original_rows3 = BsDept.objects.filter(prd_cd_id=prd_cd_selected)
            data_list3 = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm} for rows in original_rows3]
            df3 = pd.DataFrame(data_list3).sort_values(by=['dept_nm']).reset_index(drop=True)
            df_json3 = df3.to_json(orient='records')
            # 엑셀 저장
            # df3.to_excel('dept.xlsx', index=False)

            context.update({'data_bs_dept_grp' : df_json})
            context.update({'data_bs_dept_grp_domain' : df_json2})
            context.update({'data_bs_dept' : df_json3})


        if action == 'action2': # 취소 버튼 눌렀을 때

            # 그냥 다시 화면을 띄워준다.
            # BsDeptGrp를 가져와서 json으로 변경시켜준다. JB103이랑 비슷하게 한다.
            original_rows = BsDeptGrp.objects.filter(prd_cd_id=prd_cd_selected)
            data_list = [{'dept_domain': rows.dept_domain_id, 'dept_grp_nm' : rows.dept_grp_nm_id, 'dept_cd' : rows.dept_cd_id, 'dept_seq': rows.dept_seq} for rows in original_rows]
            # df1 = pd.DataFrame(data_list).sort_values(by=['domain_seq', 'grp_seq', 'dept_seq']).reset_index(drop=True)
            df1 = pd.DataFrame(data_list)
            # dataframe을 json으로 변경시켜서 html로 보낸다. JB103이랑 비슷하게 한다.
            df_json = df1.to_json(orient='records')
            # json데이터 엑셀 저장
            # df1.to_excel('dept_grp.xlsx', index=False)

            original_rows2 = BsDeptGrpDomain.objects.filter(prd_cd_id=prd_cd_selected)
            data_list2 = [{'dept_domain': rows.dept_domain, 'dept_grp_nm' : rows.dept_grp_nm, 'domain_seq': rows.domain_seq,
                           'grp_seq': rows.grp_seq } for rows in original_rows2]
            df2 = pd.DataFrame(data_list2).sort_values(by=['domain_seq', 'grp_seq']).reset_index(drop=True)
            df_json2 = df2.to_json(orient='records')
            # 엑셀 저장
            # df2.to_excel('dept_grp_domain.xlsx', index=False)

            original_rows3 = BsDept.objects.filter(prd_cd_id=prd_cd_selected)
            data_list3 = [{'dept_cd' : rows.dept_cd, 'dept_nm' : rows.dept_nm} for rows in original_rows3]
            df3 = pd.DataFrame(data_list3).sort_values(by=['dept_nm']).reset_index(drop=True)
            df_json3 = df3.to_json(orient='records')
            # 엑셀 저장
            # df3.to_excel('dept.xlsx', index=False)

            context.update({'data_bs_dept_grp' : df_json})
            context.update({'data_bs_dept_grp_domain' : df_json2})
            context.update({'data_bs_dept' : df_json3})

    return render(request, 'jobs/BS300.html', context)


def BS103_1(request): ## 회기 선택 후 화면

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd']

        # 회기 최종 마감일에 표시되는 날짜와 최종 마감 버튼 control
        if BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn == "Y": # 선택된 회기가 마감된 회기면
            today_date = str(BsPrd.objects.get(prd_cd=prd_cd_selected).prd_end_dt).split()[0] # 회기 최종 마감일에는 마감된 날짜가 표시된다.
            register_act = 'no' # 최종 마감 버튼은 비활성화시킨다.

        else: # 선택된 회기가 마감되지 않은 회기이면
            today_date = str(dt.datetime.today()).split()[0] # 회기 최종 마감일에는 오늘 날짜가 표시된다.
            register_act = 'yes' # 최종 마감 버튼을 활성화시킨다.

        context = {
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'title' : '회기 관리',
            'tab' : 'tab2',
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : prd_cd_selected,
            'today_date' : today_date, # 디폴트로 뜨는 마지막 회기의 마감일자 혹은 오늘 일자
            'register_act' : register_act # 버튼 컨트롤 키
        }

    return render(request, 'jobs/BS103.html', context) #장고가 context를 meshup해서 html template으로 보내줌


def BS103_2(request): ## 회기 확정일 지정

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd']

        bs_prd_update = BsPrd.objects.get(prd_cd=prd_cd_selected) #BsPrd의 마지막 것 선택
        bs_prd_update.prd_end_dt = request.POST['prd_end_dt'] #html에서 받은 확정일을 BsPrd 마지막 데이터의 확정일로 지정
        bs_prd_update.prd_done_yn = "Y" #BsPrd의 마지막 것의 prd_done 여부를 Y로 지정하여 확정 처리함.
        bs_prd_update.save()

        context = {
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'title' : '회기 관리',
            'today_date' : str(dt.datetime.today()).split()[0],
            'tab' : 'tab2',
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : BsPrd.objects.last().prd_cd,
            'register_act' : 'no' # 버튼 컨트롤 키
        }

    return render(request, 'jobs/BS103.html', context) #장고가 context를 meshup해서 html template으로 보내줌


def BS105_1(request): #회기 표준정보에서 회기 선택할 시 그 회기에 해당하는 데이터 표시

    if request.method == 'POST':

        #회기 선택 값을 받아온다.
        prd_cd_selected = request.POST["prd_cd_selected"]

        try:

            context = {
            'title' : '표준 정보', # 제목
            'total_dys' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).total_dys,
            'std_wrk_able_dys' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_able_dys,
            'std_wrk_able_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_able_tm,
            'ade_ot_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).ade_ot_tm,
            'std_wrk_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_tm,
            'min_max' : BsWlOvSht.objects.get(prd_cd=prd_cd_selected),
            'g1' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G1"),
            'g2' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G2"),
            'g3' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G3"),
            'g4' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G4"),
            'g5' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G5"),
            's_m_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="수석부장"),
            'd_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="부장"),
            'a_d_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="차장"),
            'm_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="과장"),
            'a_m_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="대리"),
            'c_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="사원"),
            't_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="기능직"),
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : prd_cd_selected,
            'activate' : "activate",
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            return render(request, 'jobs/BS105.html', context)

        except BsStdWrkTm.DoesNotExist as e :

            # error_message = "해당 회기 및 부서에는 데이터가 없습니다."
            messages.error(request, f'에러 발생: {"해당 회기 내 데이터 없음"}')

            context = {
                'prd' : BsPrd.objects.all().order_by('-prd_cd'),
                'error_message' : "해당 회기 내 데이터 없음",
                'my_value' : "에러",
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),

            }

            return render(request, 'jobs/BS105.html', context)


def BS105_2(request): #회기 표준정보에서 등록 버튼 누를 시 그 회기에 대한 내용을 update

    if request.method == 'POST':
        prd_cd_selected = request.POST["prd_cd_selected"]
        total_dys_new = request.POST["total_dys_new"]
        std_wrk_able_dys_new = request.POST["std_wrk_able_dys_new"]
        std_wrk_able_tm_new = request.POST["std_wrk_able_tm_new"]
        ade_ot_tm_new = request.POST["ade_ot_tm_new"]
        std_wrk_tm_new = request.POST["std_wrk_tm_new"]

        item = BsStdWrkTm.objects.get(prd_cd=prd_cd_selected)
        item.total_dys = total_dys_new
        item.std_wrk_able_dys = std_wrk_able_dys_new
        item.std_wrk_able_tm = std_wrk_able_tm_new
        item.ade_ot_tm = ade_ot_tm_new
        item.std_wrk_tm = std_wrk_tm_new
        item.save()

        min_new = request.POST["min"]
        max_new = request.POST["max"]

        item2 = BsWlOvSht.objects.get(prd_cd=prd_cd_selected)
        item2.ov_sht_min = min_new
        item2.ov_sht_max = max_new
        item2.save()

        g1_wt = request.POST["g1_wt"]
        g2_wt = request.POST["g2_wt"]
        g3_wt = request.POST["g3_wt"]
        g4_wt = request.POST["g4_wt"]
        g5_wt = request.POST["g5_wt"]

        g1_min = request.POST["g1_min"]
        g2_min = request.POST["g2_min"]
        g3_min = request.POST["g3_min"]
        g4_min = request.POST["g4_min"]
        g5_min = request.POST["g5_min"]
        g1_max = request.POST["g1_max"]
        g2_max = request.POST["g2_max"]
        g3_max = request.POST["g3_max"]
        g4_max = request.POST["g4_max"]
        g5_max = request.POST["g5_max"]

        s_m_grade = request.POST["senior_manager"]
        d_grade = request.POST["director"]
        a_d_grade = request.POST["associate_director"]
        m_grade = request.POST["manager"]
        a_m_grade = request.POST["assistant_manager"]
        c_grade = request.POST["clerk"]
        t_grade = request.POST["technical"]

        g1 = BsWorkGrade.objects.filter(prd_cd=prd_cd_selected, work_grade="G1")
        g1.update(workload_wt=g1_wt)
        g1.update(work_lv_min=g1_min)
        g1.update(work_lv_max=g1_max)
        g2 = BsWorkGrade.objects.filter(prd_cd=prd_cd_selected, work_grade="G2")
        g2.update(workload_wt=g2_wt)
        g2.update(work_lv_min=g2_min)
        g2.update(work_lv_max=g2_max)
        g3 = BsWorkGrade.objects.filter(prd_cd=prd_cd_selected, work_grade="G3")
        g3.update(workload_wt=g3_wt)
        g3.update(work_lv_min=g3_min)
        g3.update(work_lv_max=g3_max)
        g4 = BsWorkGrade.objects.filter(prd_cd=prd_cd_selected, work_grade="G4")
        g4.update(workload_wt=g4_wt)
        g4.update(work_lv_min=g4_min)
        g4.update(work_lv_max=g4_max)
        g5 = BsWorkGrade.objects.filter(prd_cd=prd_cd_selected, work_grade="G5")
        g5.update(workload_wt=g5_wt)
        g5.update(work_lv_min=g5_min)
        g5.update(work_lv_max=g5_max)

        pos_grade1 = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, pos_nm="수석부장")
        pos_grade1.update(work_grade = s_m_grade)
        pos_grade2 = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, pos_nm="부장")
        pos_grade2.update(work_grade = d_grade)
        pos_grade3 = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, pos_nm="차장")
        pos_grade3.update(work_grade = a_d_grade)
        pos_grade4 = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, pos_nm="과장")
        pos_grade4.update(work_grade = m_grade)
        pos_grade5 = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, pos_nm="대리")
        pos_grade5.update(work_grade = a_m_grade)
        pos_grade6 = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, pos_nm="사원")
        pos_grade6.update(work_grade = c_grade)
        pos_grade7 = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, pos_nm="기능직")
        pos_grade7.update(work_grade = t_grade)

        context = {
            'title' : '표준 정보', # 제목
            'total_dys' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).total_dys,
            'std_wrk_able_dys' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_able_dys,
            'std_wrk_able_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_able_tm,
            'ade_ot_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).ade_ot_tm,
            'std_wrk_tm' : BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_tm,
            'min_max' : BsWlOvSht.objects.get(prd_cd=prd_cd_selected),
            'g1' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G1"),
            'g2' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G2"),
            'g3' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G3"),
            'g4' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G4"),
            'g5' : BsWorkGrade.objects.get(prd_cd=prd_cd_selected, work_grade="G5"),
            's_m_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="수석부장"),
            'd_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="부장"),
            'a_d_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="차장"),
            'm_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="과장"),
            'a_m_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="대리"),
            'c_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="사원"),
            't_grade' : BsPosGrade.objects.get(prd_cd=prd_cd_selected, pos_nm="기능직"),
            'prd_cd_selected' : prd_cd_selected,
            'prd' : BsPrd.objects.all(),
            'activate' : "activate",
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

    return render(request, 'jobs/BS105.html', context)


def BS106_1(request): # 직무 관리에서 회기 및 직무 유형을 선택하면 해당하는 직무 리스트를 띄워준다.

    context = {
        'title' : '직무 관리', # 제목
        'prd_list': BsPrd.objects.all().order_by('-prd_cd'),  # 회기 리스트
        'activate': "activate",  # 라디오 버튼 작동 시 사용
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
    }

    if request.method == 'POST':
        prd_selected = request.POST['prd_selected']  # 회기값
        job_type = request.POST['job_type']  # 직무 유형

        # 초기 job_list 설정
        job_list = BsJob.objects.filter(prd_cd_id=prd_selected).order_by('job_nm')

        # job_type에 따라 조건 분기
        if job_type == 'all':
            save = "no"  # 저장 버튼 deactivate
        elif job_type == 'common':
            job_list = job_list.filter(job_type="공통")
            save = "yes"  # 저장 버튼 activate
            job_type = "common"
        elif job_type == 'unique':
            job_list = job_list.filter(job_type="고유")
            save = "yes"  # 저장 버튼 activate
            job_type = 'unique'

        # 공통 context 업데이트
        context.update({
            'prd_selected': prd_selected,
            'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
            'job_list': job_list,
            'job_type': job_type,
            'save': save,
            'job_type_selected' : "latter" # 직무유형 선택 후
        })

    return render(request, 'jobs/BS106.html', context)


def BS106_2(request): # 직무 선택하면 아래에 직무 성과책임과 직무 사용부서 목록을 띄워준다.

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"] # 회기
        job_type = request.POST['job_type'] # 직무 유형

        if 'job_radio_102' in request.POST: # 직무 선택 시 (라디오 버튼 선택)

            radio_selected = request.POST['job_radio_102'] # 라디오 버튼(직무코드) value 가져옴
            job_list = BsJob.objects.filter(prd_cd_id=prd_selected).order_by('job_nm')

            if job_type == 'all': # job_type이 all일 경우 해당 dept_cd에 해당하는 bs_job 데이터를 모두 가져옴
                # BsJobDept 테이블로부터 해당 직무를 사용하는 오브젝트 목록을 가져온다.
                job_dept_list = BsJobDept.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected)
                # 가져온 목록으로부터 부서 코드 목록을 추출한다.
                job_dept_list = [obj.dept_cd_id for obj in job_dept_list]
                # 부서 코드 목록을 활용하여 BsDept 테이블을 이용해 부서명 목록을 만든다.
                job_dept_list = BsDept.objects.filter(prd_cd_id=prd_selected, dept_cd__in=job_dept_list)
                # 부서명 목록을 추출한다.
                job_dept_list = [obj.dept_nm for obj in job_dept_list]

                context = {
                    'title' : '직무 관리', # 제목
                    'prd_list' : BsPrd.objects.all().order_by('-prd_cd'), # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'job_list' : job_list, # job_list는 html에 띄워줄 결과값으로, 전체를 다 가져옴
                    'job_type' : "all",
                    'activate' : "activate", # 직무유형 라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                    'save' : "no", #저장 버튼 deactivate
                    'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                    'radio_selected' : radio_selected,
                    'job_dept_list' : job_dept_list,
                    'act_del' : "yes",
                    'job_type_selected' : "latter" # 직무유형 선택 전
                }

            elif job_type == 'common': #job_type이 common일 경우

                # BsJobDept 테이블로부터 해당 직무를 사용하는 오브젝트 목록을 가져온다.
                job_dept_list = BsJobDept.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected)
                # 가져온 목록으로부터 부서 코드 목록을 추출한다.
                job_dept_list = [obj.dept_cd_id for obj in job_dept_list]
                # 부서 코드 목록을 활용하여 BsDept 테이블을 이용해 부서명 목록을 만든다.
                job_dept_list = BsDept.objects.filter(prd_cd_id=prd_selected, dept_cd__in=job_dept_list)
                # 부서명 목록을 추출한다.
                job_dept_list = [obj.dept_nm for obj in job_dept_list]

                context = {
                    'title' : '직무 관리', # 제목
                    'prd_list' : BsPrd.objects.all().order_by('-prd_cd'), # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'job_list' : job_list.filter(job_type="공통"), # job_list는 html에 띄워줄 결과값으로, 전체를 다 가져옴
                    'activate' : "activate", #라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                    'job_type' : "common",
                    'save' : "yes", #저장 버튼 activate
                    'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                    'radio_selected' : radio_selected,
                    'job_dept_list' : job_dept_list,
                    'job_type_selected' : "latter", # 직무유형 선택 전
                    'act_del' : "yes" # 삭제 버튼 activate
                }

            #job_type이 spec일 경우 해당 dept_cd에서 job_type이 고유인 bs_job 데이터만 모두 가져옴
            elif job_type == 'unique':

                # BsJobDept 테이블로부터 해당 직무를 사용하는 오브젝트 목록을 가져온다.
                job_dept_list = BsJobDept.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected)
                # 가져온 목록으로부터 부서 코드 목록을 추출한다.
                job_dept_list = [obj.dept_cd_id for obj in job_dept_list]
                # 부서 코드 목록을 활용하여 BsDept 테이블을 이용해 부서명 목록을 만든다.
                job_dept_list = BsDept.objects.filter(prd_cd_id=prd_selected, dept_cd__in=job_dept_list)
                # 부서명 목록을 추출한다.
                job_dept_list = [obj.dept_nm for obj in job_dept_list]

                context = {
                    'title' : '직무 관리', # 제목
                    'prd_list' : BsPrd.objects.all().order_by('-prd_cd'), # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'job_list' : job_list.filter(job_type="고유"), # job_list는 html에 띄워줄 결과값으로, 전체를 다 가져옴
                    'activate' : "activate", #라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                    'job_type' : 'unique',
                    'save' : "yes", #저장 버튼 activate
                    'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                    'radio_selected' : radio_selected,
                    'job_dept_list' : job_dept_list,
                    'act_del' : "yes",
                    'job_type_selected' : "latter" # 직무유형 선택 전
                }

        if 'action' in request.POST:

            #html에서 띄워진 결과값들을 수정한 결과를 가져옴.
            input_cd = request.POST.getlist('job_cd_hidden')
            input_nm = request.POST.getlist('job_nm_102')
            input_descrp = request.POST.getlist('job_desc_102')

            action = request.POST['action']

            if action == 'action1': # 저장 - 조회할 때 직무 유형에 따라 조회하므로 유형별 update코드는 동일한 코드 사용. 재조회시 조건따라 구분.

                # 직무 정보 업데이트
                for value, code, descrp in zip(input_nm, input_cd, input_descrp):
                    BsJob.objects.filter(prd_cd_id=prd_selected, job_cd=code).update(job_nm=value, job_descrp=descrp)

                # 업데이트된 직무 목록(job_list) 가져오기
                job_list = BsJob.objects.filter(prd_cd_id=prd_selected).order_by('job_nm')

                # 고유 직무와 공통 직무에 따라 job_list 필터링
                if job_type == "unique":
                    job_list = job_list.filter(job_type="고유")
                elif job_type == "common":
                    job_list = job_list.filter(job_type="공통")

                # 공통 컨텍스트 설정
                context = {
                    'title': '직무 관리',
                    'prd_list': BsPrd.objects.all(),
                    'prd_selected': prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'job_list': job_list,
                    'job_type': job_type,
                    'activate': "activate",
                    'save': "yes",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'job_type_selected' : "latter" # 직무유형 선택 전
                }

            elif action == 'action2': #action이 삭제버튼 눌렀을 때(삭제) 라디오 버튼 값을 받는다(job_cd), job_cd와 해당 팀 정보를 이용해 삭제할 행을 삭제한다.

                # 삭제할 라디오 버튼 값(직무코드)을 받는다.
                radio_value = request.POST['job_radio_102']

                # 해당 직무코드를 가진 행 삭제
                BsJob.objects.filter(prd_cd_id=prd_selected, job_cd=radio_value).delete()

                # 업데이트된 직무 목록(job_list) 가져오기
                job_list = BsJob.objects.filter(prd_cd_id=prd_selected).order_by('job_nm')

                # 고유 직무와 공통 직무에 따라 job_list 필터링
                if job_type == "unique":
                    job_list = job_list.filter(job_type="고유")
                elif job_type == "common":
                    job_list = job_list.filter(job_type="공통")

                context = {
                    'title' : '직무 관리', # 제목
                    'prd_list' : BsPrd.objects.all(), #회기 목록
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'job_list': job_list,
                    'job_type' : job_type,
                    'activate' : "activate",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'save' : "yes", #저장 버튼 activate
                    'job_type_selected' : "latter" # 직무유형 선택 전
                }

            elif action == 'action3': # action이 추가버튼 눌렀을 때(추가) - 입력할 수 있는 칸을 늘려주는 것. 고유와 공통일 때만 가능.

                try:
                    # 직무 유형에 따른 새로운 직무코드의 설정 값 정의(고유일 때와 공통일 때 구분)
                    job_type_kr = "고유" if job_type == "unique" else "공통"
                    code_prefix = "JU" if job_type == "unique" else "JC"

                    # 새로운 코드 생성 로직
                    # char = BsJob.objects.filter(prd_cd_id=prd_selected, job_type=job_type_kr).order_by('job_cd').last().job_cd
                    # 회기와 상관없이 가장 큰 문자를 가져와서 새로운 코드 생성
                    char = BsJob.objects.filter(job_type=job_type_kr).order_by('job_cd').last().job_cd
                    new_code = code_prefix + f"{(int(char[2:6]) + 1):03d}"  # char의 마지막 세 글자(숫자)를 기준으로 새 코드 생성

                    # 직무 목록 필터링
                    job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type=job_type_kr).order_by('job_nm')

                    # 공통 컨텍스트 설정
                    context = {
                        'title': '직무 관리',  # 제목
                        'prd_list': BsPrd.objects.all(),  # 회기 목록
                        'prd_selected': prd_selected,
                        'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                        'job_list': job_list,
                        'job_type': job_type,
                        'activate': "activate",
                        'save': "yes",  # 저장 버튼 활성화
                        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                        'new_key': "activate",  # 새 칸 추가 활성화
                        'new_code': new_code,  # 생성된 새 코드
                        'job_type_selected' : "latter" # 직무유형 선택 전
                    }

                except AttributeError:
                    job_type_kr = "고유" if job_type == "unique" else "공통"
                    code_prefix = "JU" if job_type == "unique" else "JC"
                    new_code = code_prefix + "001"  # char의 마지막 세 글자(숫자)를 기준으로 새 코드 생성

                    # 직무 목록 필터링
                    job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type=job_type_kr).order_by('job_nm')

                    # 공통 컨텍스트 설정
                    context = {
                        'title': '직무 관리',  # 제목
                        'prd_list': BsPrd.objects.all(),  # 회기 목록
                        'prd_selected': prd_selected,
                        'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                        'job_list': job_list,
                        'job_type': job_type,
                        'activate': "activate",
                        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                        'save': "yes",  # 저장 버튼 활성화
                        'new_key': "activate",  # 새 칸 추가 활성화
                        'new_code': new_code  # 생성된 새 코드
                    }

            #action이 취소버튼 눌렀을 때
            elif action == 'action4':

                if job_type == "unique":
                    job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유").order_by('job_nm')
                elif job_type == "common":
                    job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통").order_by('job_nm')

                context = {
                    'title' : '직무 관리', # 제목
                    'prd_list' : BsPrd.objects.all(), #회기 목록
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'job_list' : job_list,
                    'job_type' : job_type,
                    'activate' : "activate",
                    'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    'save' : "yes", #저장 버튼 activate
                    'job_type_selected' : "latter" # 직무유형 선택 전
                }

    return render(request, 'jobs/BS106.html', context)


def BS106_3(request): # 추가 후 저장 혹은 취소 버튼 누르기

    context = {}

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"]
        job_type = request.POST['job_type']
        job_type_kr = "고유" if job_type == "unique" else "공통"
        action = request.POST['action']

        # 공통으로 사용되는 job_list 설정(동작 후 조회)
        job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type=job_type_kr).order_by('job_nm')

        # 직무 추가 저장 로직
        if action == 'action1' and request.POST.get("new_y") == "new_member_yes":
            BsJob.objects.create(
                pk=prd_selected,
                job_cd=request.POST['job_cd_102_new'],
                job_nm=request.POST['job_nm_102_new'],
                job_type=job_type_kr,
                job_descrp=request.POST['job_desc_102_new']
            )

        # 공통 context 설정
        context.update({
            'title': '직무 관리',
            'prd_list': BsPrd.objects.all(),
            'prd_selected': prd_selected,
            'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
            'job_list': job_list,
            'job_type': job_type,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'save': "yes",  # "저장 버튼 activate" 의미를 갖는지 확인 필요
            'activate': "activate",  # 상황에 따라 필요한지 검토
            'job_type_selected' : "latter" # 직무유형 선택 전
        })

    return render(request, 'jobs/BS106.html', context)


def BS106_4(request): # 직무 성과책임 저장 혹은 취소

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"]
        job_type = request.POST['job_type']
        radio_selected = request.POST['radio_selected'] # 라디오 버튼(직무코드) value 가져옴
        action = request.POST['action']

        if action == 'action1': # 저장 버튼 눌렀을 때

            # input value들의 값을 가져온다.
            job_resp = request.POST.getlist('job_resp')
            columns = ['job_resp']
            df1 = pd.DataFrame(job_resp, columns=columns)
            df1.reset_index(inplace=True)

            # BsJobResp 테이블에서 해당 회기, 직무 데이터를 삭제하고 새로 create해준다.
            BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).delete()

            for i in range(0, len(df1)):
                BsJobResp.objects.create(pk=prd_selected, job_cd_id=radio_selected, job_resp_ordr=i+1, job_resp=df1.iloc[i, 1])

            if job_type == "unique":
                job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유").order_by('job_nm')
            elif job_type == "common":
                job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통").order_by('job_nm')

            context = {
                'title' : '직무 기본정보', # 제목
                'prd_list' : BsPrd.objects.all(),
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'job_list' : job_list,
                'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                'activate' : "activate", #라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                'job_type' : job_type,
                'radio_selected' : radio_selected,
                'act_del' : "yes",
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'job_type_selected' : "latter" # 직무유형 선택 전
            }

        elif action == 'action2': # 취소 버튼 눌렀을 때 - 그냥 원상복구

            if job_type == "unique":
                job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유").order_by('job_nm')
            elif job_type == "common":
                job_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통").order_by('job_nm')

            context = {
                'title' : '직무 기본정보', # 제목
                'prd_list' : BsPrd.objects.all(),
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'job_list' : job_list,
                'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                'activate' : "activate", #라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                'job_type' : job_type,
                'radio_selected' : radio_selected,
                'act_del' : "yes",
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'job_type_selected' : "latter" # 직무유형 선택 전
            }

    return render(request, 'jobs/BS106.html', context)


def CC102_a(request): ## 공통코드관리 초기화면
    if request.method == 'POST':
        common_code = request.POST["common_code"] #html에서 선택한 값(CcCdHeader 모델의 domain_cd)를 common_code라는 변수에 지정(str)
        context = {
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'title' : '공통 코드', # 제목
            'CC_list': CcCdHeader.objects.exclude(domain_cd="A5").all(), #A1은 직무 유형이라서 필요없고 다 가져옴
            'new_value' : "ready", #신규추가 버튼을 나타나게 해줌
            'text' : '조회',
            'common_code' : common_code, #선택한 공통 코드의 domain_cd를 넘겨준다.
            'radio_list' : CcCdDetail.objects.filter(domain_cd=common_code) #중요; common_code를 domain_cd로 갖고 있는 CcCodeDetail모델의 모든 테이블을 가져와 html에 보냄. 거기서 라디오 버튼 목록 생성
        }

    return render(request, 'jobs/CC102.html', context)


def CC102_b(request): ## 공통코드 관리 수정

    if request.method == 'POST':
        common_code = request.POST["common_code"] #선택한 공통 코드 정보를 계속 그대로 받아온다.
        selected_domain_cd = request.POST.getlist('selected_domain_cd') #선택한 공통 코드에 해당하는 cc_code 리스트를 받아온다.
        selected_cc_code = request.POST.getlist('selected_cc_code') #선택한 공통 코드에 해당하는 cc_code_nm 리스트를 받아온다.
        cc_code_nm_change = request.POST.getlist('cc_code_nm_change')
        action = request.POST['action'] #어떤 버튼을 누르냐에 따라 다른 행동을 취한다.

        #저장 버튼을 누를 때. 기존 DB 삭제하고 새로운 값들로 채워넣을 것이다. 그리고 추가 칸이 있을때와 없을 때 로직이 달라져야 한다.
        if action == 'action1':
            context = {
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'title' : '공통 코드', # 제목
                'CC_list': CcCdHeader.objects.exclude(domain_cd="A5").all(),
                'radio_list' : CcCdDetail.objects.filter(domain_cd=common_code),
                'text' : "저장",
                'new_value' : "ready", #신규추가 버튼을 나타나게 해줌
                'common_code' : common_code
            }

            CcCdDetail.objects.filter(domain_cd_id=common_code).delete()

            for i, j in zip(selected_domain_cd, cc_code_nm_change):
                CcCdDetail.objects.create(domain_cd_id=common_code, cc_code=i, cc_code_nm=j)

        #삭제 버튼을 누를 시
        elif action == 'action2':
            del_target_nm = request.POST["radanswer"] #삭제 버튼 누르면, 라디오 버튼의 선택값(cc_code_nm 즉, 공통코드이름)을 넘겨받는다.
            context = {
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'title' : '공통 코드', # 제목
                'CC_list': CcCdHeader.objects.exclude(domain_cd="A5").all(),
                'radio_list' : CcCdDetail.objects.filter(domain_cd=common_code),
                'text' : "삭제",
                'new_value' : "ready", #신규추가 버튼을 나타나게 해줌
                'common_code' : common_code
            }

            CcCdDetail.objects.filter(domain_cd_id=common_code, cc_code_nm=del_target_nm).delete()

        #추가 버튼을 누르면 실제로는 추가하기 위한 key값만 넘겨주고 html에서 추가를 위한 입력란을 형성한다. 그 input은 저장 버튼을 누르면 저장된다.
        elif action == 'action3' :

            last_number = CcCdDetail.objects.filter(domain_cd_id=common_code).order_by('cc_code').last().cc_code
            new_number = f"{(int(last_number)+1):02}"
            # print(last_number)
            # print(new_number)

            context = {
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'title' : '공통 코드', # 제목
                'CC_list': CcCdHeader.objects.exclude(domain_cd="A5").all(),
                'radio_list' : CcCdDetail.objects.filter(domain_cd=common_code),
                'text' : "추가", # 이 값이 추가하기 위한 값이다. 중요!
                'new_value' : "ready", #신규추가 버튼을 나타나게 해줌
                'common_code' : common_code,
                'last_number' : last_number,
                'new_number' : new_number,
                'new_cd_number' : common_code + new_number
            }
        # 취소 버튼을 누르면 빽도
        elif action == 'action4' :

            context = {
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'title' : '공통 코드', # 제목
                'CC_list': CcCdHeader.objects.exclude(domain_cd="A5").all(),
                'radio_list' : CcCdDetail.objects.filter(domain_cd=common_code),
                'text' : "저장",
                'new_value' : "ready", #신규추가 버튼을 나타나게 해줌
                'common_code' : common_code
            }

    return render(request, 'jobs/CC102.html', context)


def CC102_c(request): ## 공통코드 관리 수정

    if request.method == 'POST':

        # 추가해줄 대상의 domain_cd_id를 받아온다.
        target_domain_cd = request.POST['common_code']
        new_nm = request.POST['new_cc_code_nm']

        # 새로운 object의 cc_code
        last_number = CcCdDetail.objects.filter(domain_cd_id=target_domain_cd).order_by('cc_code').last().cc_code
        new_number = f"{(int(last_number)+1):02}"

        if request.POST["new_code"] == "new_code" :
            CcCdDetail.objects.create(domain_cd_id=target_domain_cd, cc_code=new_number, cc_code_nm=new_nm)

        else :
            text = "check"

        context = {
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'title' : '공통 코드', # 제목
                'CC_list': CcCdHeader.objects.exclude(domain_cd="A5").all(),
                'radio_list' : CcCdDetail.objects.filter(domain_cd=target_domain_cd),
                'text' : "저장",
                'new_value' : "ready", #신규추가 버튼을 나타나게 해줌
                'common_code' : target_domain_cd
            }

    return render(request, 'jobs/CC102.html', context)


def jb101_1(request): # JB101에서 회기를 선택한 후 탭을 선택했을 때. 바로 탭 선택에 대한 결과를 띄워줌

    context = {}

    if request.method == 'POST':
        # 선택한 회기를 input으로 받아옴.
        prd_cd_selected = request.POST['prd_cd_selected']
        # span을 탭을 어느 것을 선택하느냐에 따라 다르게 연산을 수행할 것임. 일단 그 span이 뭔지 알아낼 것임.
        span_name = request.POST.get('span_name', '')

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        # 공통으로 사용하는 context 설정
        context = {
            'title': '부서 기본정보',  # 제목
            'prd_list': BsPrd.objects.all(),
            # 'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            'prd_cd_selected': prd_cd_selected,
            'activate': 'no',  # 버튼 컨트롤 off
            'status': 'tab_after',
            'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
            'dept_login_nm' : dept_login_nm,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if dept_login == 'DD06': # 경영기획팀이 로그인했을 경우 다른 팀들도 선택할 수 있도록 함.
            context.update({
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            })
        else:
            context.update({
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login),
            })

        # 선택된 탭에 따라 tab 값 설정
        if span_name == 'span1':
            context['tab'] = "tab1"

            # 부서 성과책임 리스트를 가져옴
            rows = BsDeptResp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_login)

            context.update({
                'dept_resp_list' : rows,
                'dept_selected': dept_login,
                'activate': 'yes', # 버튼 컨트롤 on
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn
            })

        elif span_name == 'span2':
            context['tab'] = "tab2"

            last_item = BsPosList.objects.filter(prd_cd_id=prd_cd_selected).order_by('pos_ordr').last()
            last_pos_ordr = last_item.pos_ordr
            pos_list = [str(item.pos_nm) for item in BsPosList.objects.filter(prd_cd_id=prd_cd_selected).order_by('-pos_ordr').exclude(pos_ordr = last_pos_ordr)]
            # 직책 리스트를 가져옴
            title_list = [str(item.ttl_nm) for item in BsTtlList.objects.filter(prd_cd_id=prd_cd_selected).order_by('-ttl_ordr')]

            context.update({
                'mbr_list': BsMbrArrange(prd_cd_selected, dept_login),
                'mbr_count': BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_login).count(),
                'pos_list': pos_list,
                'ttl_list': title_list,
                'activate': 'yes', # 버튼 컨트롤 on
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                'dept_selected' : dept_login
            })

        elif span_name == 'span3': # 부서원 탭 선택한 상태일 시

            context['tab'] = "tab3"

            # 해당 회기와 해당 부서의 BsMbrGrp와 BsMbrGrpNm을 가져옴
            mbr_grp_list = BsMbrGrp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_login)
            # mbr_grp_list로 dataframe 생성
            data_list = [{'mbr_grp_nm': rows.mbr_grp_nm_id, 'mbr_nm': rows.mbr_nm_id, 'work_ratio': rows.work_ratio} for rows in mbr_grp_list]
            df_mbr_grp = pd.DataFrame(data_list)
            mbr_grp_nm_list = BsMbrGrpNm.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_login)

            # 부서원 그룹 테이블을 만들기 위한 데이터프레임 생성
            columns = list(mbr_grp_nm_list.values_list('mbr_grp_nm', flat=True))  # 부서원 그룹 이름
            # print('columns', columns)

            df_mbr = pd.DataFrame()
            df_mbr['부서원'] = [mbr.mbr_nm for mbr in BsMbrArrange(prd_cd_selected, dept_login)]

            # df_mbr 에 columns 열 추가. 데이터는 비워둠
            for column in columns:
                df_mbr[column] = None

            # df_mbr_grp의 work_ratio를 df_mbr에 넣기
            # df_mbr_grp의 각 행에 대해서 mbr_grp_nm은 df_mbr의 column에 있다.
            # df_mbr_grp의 mbr_nm은 df_mbr의 bs_mbr_nm에 있다.
            # df_mbr_grp의 work_ratio는 df_mbr의 해당 cell에 들어가야 한다.
            for i in range(len(df_mbr_grp)): # df_mbr_grp의 행 수만큼 반복한다. df_mbr_grp는 BsMbrGrp 테이블을 dataframe으로 변환한 것이다.
                mbr_grp_nm = df_mbr_grp.iloc[i, 0] # mbr_grp_nm 데이터를 저장
                mbr_nm = df_mbr_grp.iloc[i, 1] # mbr_nm 데이터를 저장
                work_ratio = df_mbr_grp.iloc[i, 2] # work_ratio 데이터를 저장

                # df_mbr은 부서원 이름과 부서원 그룹을 컬럼으로 가지는 dataframe으로, json으로 변환하여 UI로 보내줄 데이터이다.
                # df_mbr의 bs_mbr_nm 컬럼 데이터가 df_mbr_grp의 mbr_nm이고, 열 이름이 df_mbr_grp의 mbr_grp_nm인 cell에 work_ratio를 넣어준다.
                df_mbr.loc[df_mbr['부서원'] == mbr_nm, mbr_grp_nm] = work_ratio

            # print(df_mbr_grp)
            # print(df_mbr)
            # 예시 데이터 생성
            data = {
                'group1': [0.1, 0.2, 0.3, 0.1],
                'group2': [0.15, 0.25, 0.35, 0.2],
                'group3': [0.1, 0.3, 0.4, 0.03],
                'group4': [0.1, None, 0.4, 0.2]
            }
            df = pd.DataFrame(data)
            df['부서원'] = ['Employee1', 'Employee2', 'Employee3', 'Employee4']

            df = df[['부서원', 'group1', 'group2', 'group3', 'group4']]

            groups = ['group1', 'group2', 'group3' , 'group4']

            # print(df)
            # 데이터프레임을 JSON으로 변환
            df_json = df_mbr.to_json(orient='records')
            groups_json = json.dumps(groups)  # jobs 리스트를 JSON으로 변환

            # 템플릿에 전달할 컨텍스트 딕셔너리 생성
            context.update({
                'df' : df,
                'data' : df_json,
                'groups_json' : groups_json,
                'activate': 'yes', # 버튼 컨트롤 on
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                'dept_selected' : dept_login,
            })

    return render(request, 'jobs/JB101.html', context)


def jb101_2(request): # 탭이 선택된 상태에서 부서를 선택했을 때임. 관련 정보를 가져와서 보여줄 것임. 경영기획팀만 해당함

    if request.method == 'POST':
        prd_cd_selected = request.POST['prd_cd_selected']
        dept_selected = request.POST['dept_selected']

        tab = request.POST['tab']  # 탭 정보

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드

        # 공통 context 설정
        context = {
            'title': '부서 기본정보',
            'prd_list': BsPrd.objects.all(),
            'prd_cd_selected': prd_cd_selected,
            'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            'dept_selected': dept_selected,
            'tab': tab,
            'activate': 'yes', # 버튼 컨트롤 on
            'status': 'tab_after',
            'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if tab == "tab1": # 부서 정보 탭 선택한 상태일 시 - 부서 성과책임 표시
            # context['dept_resp_list'] = BsDeptResp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)

            rows = BsDeptResp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)

            context.update({
                'dept_resp_list' : rows,
            })

        elif tab == "tab2": # 부서원 탭 선택한 상태일 시 - 부서원 목록 표시
            # 직위 리스트를 가져옴, 마지막 글자(기능직) 제거
            last_item = BsPosList.objects.filter(prd_cd_id=prd_cd_selected).order_by('pos_ordr').last()
            last_pos_ordr = last_item.pos_ordr
            pos_list = [str(item.pos_nm) for item in BsPosList.objects.filter(prd_cd_id=prd_cd_selected).order_by('-pos_ordr').exclude(pos_ordr = last_pos_ordr)]
            # 직책 리스트를 가져옴
            title_list = [str(item.ttl_nm) for item in BsTtlList.objects.filter(prd_cd_id=prd_cd_selected).order_by('-ttl_ordr')]

            context.update({
                'mbr_list': BsMbrArrange(prd_cd_selected, dept_selected),
                'mbr_count': BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count(),
                'pos_list': pos_list,
                'ttl_list': title_list,
            })

        elif tab == "tab3": # 부서원 그룹 탭 선택한 상태일 시 - 부서원 그룹 테이블 표시

            # 해당 회기와 해당 부서의 BsMbrGrp와 BsMbrGrpNm을 가져옴
            mbr_grp_list = BsMbrGrp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)
            # mbr_grp_list로 dataframe 생성
            data_list = [{'mbr_grp_nm': rows.mbr_grp_nm_id, 'mbr_nm': rows.mbr_nm_id, 'work_ratio': rows.work_ratio} for rows in mbr_grp_list]
            df_mbr_grp = pd.DataFrame(data_list)
            mbr_grp_nm_list = BsMbrGrpNm.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)

            # 부서원 그룹 테이블을 만들기 위한 데이터프레임 생성
            columns = list(mbr_grp_nm_list.values_list('mbr_grp_nm', flat=True))  # 부서원 그룹 이름
            # print('columns', columns)

            df_mbr = pd.DataFrame()
            df_mbr['부서원'] = [mbr.mbr_nm for mbr in BsMbrArrange(prd_cd_selected, dept_selected)]

            # df_mbr 에 columns 열 추가. 데이터는 비워둠
            for column in columns:
                df_mbr[column] = None

            # df_mbr_grp의 work_ratio를 df_mbr에 넣기
            # df_mbr_grp의 각 행에 대해서 mbr_grp_nm은 df_mbr의 column에 있다.
            # df_mbr_grp의 mbr_nm은 df_mbr의 bs_mbr_nm에 있다.
            # df_mbr_grp의 work_ratio는 df_mbr의 해당 cell에 들어가야 한다.
            for i in range(len(df_mbr_grp)): # df_mbr_grp의 행 수만큼 반복한다. df_mbr_grp는 BsMbrGrp 테이블을 dataframe으로 변환한 것이다.
                mbr_grp_nm = df_mbr_grp.iloc[i, 0] # mbr_grp_nm 데이터를 저장
                mbr_nm = df_mbr_grp.iloc[i, 1] # mbr_nm 데이터를 저장
                work_ratio = df_mbr_grp.iloc[i, 2] # work_ratio 데이터를 저장

                # df_mbr은 부서원 이름과 부서원 그룹을 컬럼으로 가지는 dataframe으로, json으로 변환하여 UI로 보내줄 데이터이다.
                # df_mbr의 bs_mbr_nm 컬럼 데이터가 df_mbr_grp의 mbr_nm이고, 열 이름이 df_mbr_grp의 mbr_grp_nm인 cell에 work_ratio를 넣어준다.
                df_mbr.loc[df_mbr['부서원'] == mbr_nm, mbr_grp_nm] = work_ratio

            # print(df_mbr_grp)
            # print(df_mbr)
            # 예시 데이터 생성
            data = {
                'group1': [0.1, 0.2, 0.3, 0.1],
                'group2': [0.15, 0.25, 0.35, 0.2],
                'group3': [0.1, 0.3, 0.4, 0.03],
                'group4': [0.1, None, 0.4, 0.2]
            }
            df = pd.DataFrame(data)
            df['부서원'] = ['Employee1', 'Employee2', 'Employee3', 'Employee4']

            df = df[['부서원', 'group1', 'group2', 'group3', 'group4']]

            groups = ['group1', 'group2', 'group3' , 'group4']

            # print(df)
            # 데이터프레임을 JSON으로 변환
            df_json = df_mbr.to_json(orient='records')
            groups_json = json.dumps(groups)  # jobs 리스트를 JSON으로 변환

            # 템플릿에 전달할 컨텍스트 딕셔너리 생성
            context.update({
                'df' : df,
                'data' : df_json,
                'groups_json' : groups_json,

            })

    return render(request, 'jobs/JB101.html', context)


def jb101_3(request): # 저장 및 취소 버튼을 눌렀을 때(부서정보, 부서원)

    if request.method == 'POST':
        prd_cd_selected = request.POST.get('prd_cd_selected')
        dept_selected = request.POST.get('dept_selected')
        tab = request.POST.get('tab')  # 탭 정보
        action = request.POST.get('action')

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        # 공통 context 설정
        common_context = {
            'title': '부서 기본정보',
            'prd_list': BsPrd.objects.all(),
            'prd_cd_selected': prd_cd_selected,
            'dept_selected': dept_selected,
            'tab': tab,
            'activate': 'yes',  # 버튼 컨트롤 on
            'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
            'status': 'tab_after',
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        # 로그인 부서가 경영기획팀일 때와 아닐 때 구분
        if dept_login == "DD06":
            common_context.update({
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            })
        else:
            common_context.update({
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login),
            })

        if tab == 'tab1': # 부서 정보 탭일때

            if action == 'action1':  # 저장 버튼 눌렀을 때

                # input value들의 값을 가져온다.
                dept_resp_values = request.POST.getlist('dept_resp')  # input 태그의 name 속성 값이 'column2'
                columns = ['dept_resp']
                df1 = pd.DataFrame(dept_resp_values, columns=columns)
                df1.reset_index(inplace=True)

                # BsJobResp 테이블에서 해당 회기, 직무 데이터를 삭제하고 새로 create해준다.
                BsDeptResp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).delete()

                for i in range(0, len(df1)):
                    BsDeptResp.objects.create(pk=prd_cd_selected, dept_cd_id=dept_selected, dept_resp_ordr=i+1, dept_resp=df1.iloc[i, 1])

                common_context['dept_resp_list'] = BsDeptResp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)

            else:  # 취소 버튼 눌렀을 때
                common_context['dept_resp_list'] = BsDeptResp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)

        elif tab == 'tab2': # 부서원 탭일때

            # 직위 리스트
            last_item = BsPosList.objects.filter(prd_cd_id=prd_cd_selected).order_by('pos_ordr').last()
            last_pos_ordr = last_item.pos_ordr
            column_values_list = [str(item.pos_nm) for item in BsPosList.objects.filter(prd_cd_id=prd_cd_selected).order_by('-pos_ordr').exclude(pos_ordr = last_pos_ordr)]
            # 직책 리스트
            title_list = [str(item.ttl_nm) for item in BsTtlList.objects.filter(prd_cd_id=prd_cd_selected).order_by('-ttl_ordr')]

            if action == 'action1':  # 저장 버튼 눌렀을 때

                if dept_login == "DD06":

                    try: # 이름 중복 제한을 위해서 try-except 문 사용
                        # print('dept_selected', dept_selected)
                        # input value들의 값을 가져온다.
                        input_values_name = request.POST.getlist('team_member_name')
                        input_values_job_level = request.POST.getlist('job_level')
                        input_values_job_title = request.POST.getlist('job_title')

                        # 기존 BsMbr 데이터 삭제
                        BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd=dept_selected).delete()

                        # input value들의 값에 따라 추가.
                        for i, j, k in zip(input_values_name, input_values_job_level, input_values_job_title):
                            BsMbr.objects.create(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected, mbr_nm=i, pos_nm=j, ttl_nm=k)

                        # BsDept 테이블의 dept_po 값 업데이트. 해당 회기의 해당 부서의 BsMbr 테이블의 오브젝트 개수를 가져와서 업데이트
                        BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected).update(dept_po=BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count())


                        common_context.update({
                            'mbr_list': BsMbrArrange(prd_cd_selected, dept_selected),
                            'pos_list': column_values_list,
                            'ttl_list': title_list,
                            'mbr_count': BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count(),
                        })
                    except IntegrityError as e: # 부서원 이름의 중복을 제한함.

                        messages.error(request, '부서원 이름이 중복되었습니다.')

                        common_context.update({
                            'mbr_list': BsMbrArrange(prd_cd_selected, dept_selected),
                            'pos_list': column_values_list,
                            'ttl_list': title_list,
                            'mbr_count': BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count(),
                        })

                else:
                    try: # 이름 중복 제한을 위해서 try-except 문 사용

                        # input value들의 값을 가져온다.
                        input_values_name = request.POST.getlist('team_member_name')
                        input_values_job_level = request.POST.getlist('job_level')
                        input_values_job_title = request.POST.getlist('job_title')

                        # 기존 BsMbr 데이터 삭제
                        BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd=dept_login).delete()

                        # input value들의 값에 따라 추가.
                        for i, j, k in zip(input_values_name, input_values_job_level, input_values_job_title):
                            BsMbr.objects.create(prd_cd_id=prd_cd_selected, dept_cd_id=dept_login, mbr_nm=i, pos_nm=j, ttl_nm=k)

                        # BsDept 테이블의 dept_po 값 업데이트. 해당 회기의 해당 부서의 BsMbr 테이블의 오브젝트 개수를 가져와서 업데이트
                        BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected).update(dept_po=BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count())

                        common_context.update({
                            'mbr_list': BsMbrArrange(prd_cd_selected, dept_login),
                            'pos_list': column_values_list,
                            'ttl_list': title_list,
                            'mbr_count': BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count(),
                        })

                    except IntegrityError as e: # 부서원 이름의 중복을 제한함.

                        messages.error(request, '부서원 이름이 중복되었습니다.')

                        common_context.update({
                            'mbr_list': BsMbrArrange(prd_cd_selected, dept_selected),
                            'pos_list': column_values_list,
                            'ttl_list': title_list,
                            'mbr_count': BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count(),
                        })

            else:  # 취소 버튼 눌렀을 때
                common_context.update({
                    'mbr_list': BsMbrArrange(prd_cd_selected, dept_selected),
                    'pos_list': column_values_list,
                    'ttl_list': title_list,
                    'mbr_count': BsMbr.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).count(),
                })

        return render(request, 'jobs/JB101.html', common_context)


def jb101_4(request): # 부서원 그룹 탭에서 저장 및 취소 눌렀을 때

    if request.method == 'POST':

        prd_cd_selected = request.POST.get('prd_cd_selected')
        dept_selected = request.POST.get('dept_selected')
        tab = request.POST.get('tab')
        action = request.POST["action"]

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        if action == 'action1': # 저장 버튼 눌렀을 때

            json_data = request.POST.get('jsonData')
            data = json.loads(json_data)
            df = pd.DataFrame(data)
            # DataFrame `df`를 사용하여 필요한 처리 수행

            columns = list(df.columns)

            columns.pop()
            columns.insert(0, '부서원')
            df.columns = columns

            # df의 마지막 행을 없애줌
            df = df[:-1]

            # df의 첫째 열을 제외하고 나머지 열들의 자료형을 float으로 바꿔줌
            for column in df.columns[1:]:
                # 해당 열 중에서 빈칸이 있으면 그 빈칸은 제외하고 float으로 바꿔줌
                df[column] = df[column].apply(lambda x: float(x) if x != '' else 0)

            # 나중에 DB에 저장할 때 0.0인 것은 저장하지 않을 것임.

            # print(df)

            # 해당 회기, 해당 부서의 BsMbrGrp와 BsMbrGrpNm을 삭제함
            BsMbrGrp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).delete()
            BsMbrGrpNm.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected).delete()

            # df를 활용하여 BsMbrGrpNm과 BsMbrGrp를 생성함
            # BsMbrGrpNm은 df의 column 이름을 가져와서 생성함
            # df column 이름 리스트로 만들기
            column_list = list(df.columns)
            column_list.pop(0)  # 첫번째 열인 '부서원'은 제외함
            # column_list로 BsMbrGrpNm 생성
            for column in column_list:
                BsMbrGrpNm.objects.create(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected, mbr_grp_nm=column)

            # df의 데이터를 BsMbrGrp에 넣어줌
            for i in range(len(df)):
                mbr_nm = df.iloc[i, 0]
                for j in range(1, len(df.columns)):
                    mbr_grp_nm = column_list[j-1]
                    work_ratio = df.iloc[i, j]
                    if work_ratio != 0:
                        BsMbrGrp.objects.create(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected, mbr_grp_nm_id=mbr_grp_nm, mbr_nm_id=mbr_nm, work_ratio=work_ratio)

            ############################# 다시 띄워줄 context 설정
            # 해당 회기와 해당 부서의 BsMbrGrp와 BsMbrGrpNm을 가져옴
            mbr_grp_list = BsMbrGrp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)
            # mbr_grp_list로 dataframe 생성
            data_list = [{'mbr_grp_nm': rows.mbr_grp_nm_id, 'mbr_nm': rows.mbr_nm_id, 'work_ratio': rows.work_ratio} for rows in mbr_grp_list]
            df_mbr_grp = pd.DataFrame(data_list)
            mbr_grp_nm_list = BsMbrGrpNm.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)

            # 부서원 그룹 테이블을 만들기 위한 데이터프레임 생성
            columns_2 = list(mbr_grp_nm_list.values_list('mbr_grp_nm', flat=True))  # 부서원 그룹 이름
            # print('columns', columns_2)

            df_mbr = pd.DataFrame()
            df_mbr['부서원'] = [mbr.mbr_nm for mbr in BsMbrArrange(prd_cd_selected, dept_selected)]

            # df_mbr 에 columns 열 추가. 데이터는 비워둠
            for column in columns_2:
                df_mbr[column] = None

            # df_mbr_grp의 work_ratio를 df_mbr에 넣기
            # df_mbr_grp의 각 행에 대해서 mbr_grp_nm은 df_mbr의 column에 있다.
            # df_mbr_grp의 mbr_nm은 df_mbr의 bs_mbr_nm에 있다.
            # df_mbr_grp의 work_ratio는 df_mbr의 해당 cell에 들어가야 한다.
            for i in range(len(df_mbr_grp)): # df_mbr_grp의 행 수만큼 반복한다. df_mbr_grp는 BsMbrGrp 테이블을 dataframe으로 변환한 것이다.
                mbr_grp_nm = df_mbr_grp.iloc[i, 0] # mbr_grp_nm 데이터를 저장
                mbr_nm = df_mbr_grp.iloc[i, 1] # mbr_nm 데이터를 저장
                work_ratio = df_mbr_grp.iloc[i, 2] # work_ratio 데이터를 저장

                # df_mbr은 부서원 이름과 부서원 그룹을 컬럼으로 가지는 dataframe으로, json으로 변환하여 UI로 보내줄 데이터이다.
                # df_mbr의 bs_mbr_nm 컬럼 데이터가 df_mbr_grp의 mbr_nm이고, 열 이름이 df_mbr_grp의 mbr_grp_nm인 cell에 work_ratio를 넣어준다.
                df_mbr.loc[df_mbr['부서원'] == mbr_nm, mbr_grp_nm] = work_ratio

            # print(df_mbr_grp)
            # print(df_mbr)

            df_json = df_mbr.to_json(orient='records')

            context = {
                'title': '부서 기본정보',
                'prd_list': BsPrd.objects.all(),
                'prd_cd_selected': prd_cd_selected,
                'dept_selected': dept_selected,
                'tab': tab,
                'activate': 'yes',  # 버튼 컨트롤 on
                'status': 'tab_after',
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                'data' : df_json,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            # 로그인 부서가 경영기획팀일 때와 아닐 때 구분
            if dept_login == "DD06":
                context.update({
                    'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
                })
            else:
                context.update({
                    'status': 'tab_after',
                    'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login),
                })


        elif action == 'action2': # 취소 버튼 눌렀을 때

            ############################# 다시 띄워줄 context 설정 ##### 저장 코드 참고
            mbr_grp_list = BsMbrGrp.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)
            data_list = [{'mbr_grp_nm': rows.mbr_grp_nm_id, 'mbr_nm': rows.mbr_nm_id, 'work_ratio': rows.work_ratio} for rows in mbr_grp_list]
            df_mbr_grp = pd.DataFrame(data_list)
            mbr_grp_nm_list = BsMbrGrpNm.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=dept_selected)

            columns_2 = list(mbr_grp_nm_list.values_list('mbr_grp_nm', flat=True))  # 부서원 그룹 이름
            # print('columns', columns_2)

            df_mbr = pd.DataFrame()
            df_mbr['부서원'] = [mbr.mbr_nm for mbr in BsMbrArrange(prd_cd_selected, dept_selected)]

            for column in columns_2:
                df_mbr[column] = None

            for i in range(len(df_mbr_grp)): # df_mbr_grp의 행 수만큼 반복한다. df_mbr_grp는 BsMbrGrp 테이블을 dataframe으로 변환한 것이다.
                mbr_grp_nm = df_mbr_grp.iloc[i, 0] # mbr_grp_nm 데이터를 저장
                mbr_nm = df_mbr_grp.iloc[i, 1] # mbr_nm 데이터를 저장
                work_ratio = df_mbr_grp.iloc[i, 2] # work_ratio 데이터를 저장

                # df_mbr은 부서원 이름과 부서원 그룹을 컬럼으로 가지는 dataframe으로, json으로 변환하여 UI로 보내줄 데이터이다.
                # df_mbr의 bs_mbr_nm 컬럼 데이터가 df_mbr_grp의 mbr_nm이고, 열 이름이 df_mbr_grp의 mbr_grp_nm인 cell에 work_ratio를 넣어준다.
                df_mbr.loc[df_mbr['부서원'] == mbr_nm, mbr_grp_nm] = work_ratio

            # print(df_mbr)

            df_json = df_mbr.to_json(orient='records')

            context = {
                'title': '부서 기본정보',
                'prd_list': BsPrd.objects.all(),
                'prd_cd_selected': prd_cd_selected,
                'dept_selected': dept_selected,
                'tab': tab,
                'activate': 'yes',  # 버튼 컨트롤 on
                'status': 'tab_after',
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                'data' : df_json
            }

            if dept_login == "DD06":
                context.update({
                    'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
                })
            else:
                context.update({
                    'status': 'tab_after',
                    'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login),
                })

    return render(request, 'jobs/JB101.html', context)


def JB102_1(request): # 직무 기본정보 회기 선택화면 - 회기를 선택하면 로그인한 부서에 따라 다른 정보를 띄워줌. 직무유형 선택하게 함.

    if request.method == 'POST':

        #html에서 회기 선택
        prd_selected = request.POST["prd_selected"]
        key = request.POST["key_prd_select"]

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드

        try:
            dept_login_nm = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

            context = {

                'prd_list' : BsPrd.objects.all(),
                'title' : '직무 기본정보', # 제목
                'prd_selected' : prd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'job_type_selected' : "former" # 직무유형 선택 전

            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                context['dept_login_nm'] = dept_login_nm
                context['dept_selected'] = dept_login
            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_login_nm'] = dept_login_nm
                context['dept_selected'] = dept_login

        except ObjectDoesNotExist as e: # 새로운 부서가 로그인했을 때, 그 부서가 없었던 회기를 선택할 때만 나타난다.

            # 오류 메시지 띄워주고 라디오 버튼 선택 막는다.

            messages.error(request, '해당 회기에 로그인한 부서가 없습니다.')

            # last_prd_cd = BsPrd.objects.all().last().prd_cd # 가장 최근 회기. default로 띄워줌

            # dept_login_nm = BsDept.objects.get(prd_cd=last_prd_cd, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

            # context = {
            #     'title' : '직무 기본정보', # 제목
            #     'prd_list' : BsPrd.objects.all(),
            #     'prd_selected' : last_prd_cd,
            #     'job_type_selected' : "former", # 직무유형 선택 전
            # }

            # context['team_list'] = BsDept.objects.filter(prd_cd=last_prd_cd, dept_cd=dept_login)
            # context['dept_login_nm'] = dept_login_nm
            # context['dept_selected'] = dept_login

            context = {
                'title' : '직무 기본정보', # 제목
                'prd_list' : BsPrd.objects.all(),
                'prd_selected' : prd_selected,
                'job_type_selected' : "former", # 직무유형 선택 전
                'radio_activate': 'no', # 라디오 버튼 비활성화
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                context['dept_selected'] = dept_login
            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_selected'] = dept_login

    return render(request, 'jobs/JB102.html', context)


def JB102_2(request): # 직무 기본정보의 부서 선택을 받고 그 정보(dept_selected)를 html로 넘겨줌. 경영기획팀만 해당

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"]
        dept_selected = request.POST["dept_selected"]
        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드

        context = {
            'title' : '직무 기본정보', # 제목
            'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
            'team_list' : BsDept.objects.filter(prd_cd=prd_selected), # 선택한 회기의 부서 목록
            'dept_selected' : dept_selected,
            'prd_selected' : prd_selected,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'job_type_selected' : "former", # 직무 유형 선택 전 상태

        }

    return render(request, 'jobs/JB102.html', context)


def JB102_3(request): # 직무 기본정보의 직무 유형 선택할 수 있도록 띄워줌

    context = {
        'title': '직무 기본정보',
        'prd_list': BsPrd.objects.all().order_by,
        'team_list': None,
        'prd_selected': None,
        'dept_selected': None,
        'job_list': None,
        'job_by_list': None,
        'activate': "activate", # 직무 유형 라디오 버튼 작동하면 하단의 직무 기본정보 표시
        'job_type': None,
        'save': "no", # 저장 버튼 활성화
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
    }

    if request.method == 'POST':
        prd_selected = request.POST.get("prd_selected")
        dept_selected = request.POST.get('dept_selected')
        job_type = request.POST.get('job_type')

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        filtered_set = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected)
        filtered_value = list(filtered_set.values_list('job_cd', flat=True))

        # 직무 유형에 따라 저장 버튼을 나타낸다.
        job_type_filter = {}
        if job_type == 'all':
            pass  # 'all'일 경우 저장 버튼 필요없음
        elif job_type == 'common':
            job_type_filter['job_type'] = '공통'
            context['save'] = "yes"
        elif job_type == 'unique':
            job_type_filter['job_type'] = '고유'
            context['save'] = "yes"

        # 필터링된 결과 가져오기
        filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_cd__in=filtered_value, **job_type_filter)
        filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
        filtered_result_job_by = filtered_set.filter(job_cd__in=filtered_value_job_by)

        # context 업데이트
        context.update({
            'prd_selected': prd_selected,
            'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
            'dept_selected': dept_selected,
            'job_list': filtered_result,
            'job_by_list': filtered_result_job_by,
            'job_type': job_type,
            'job_type_selected': 'latter', # 직무 유형 선택 후 상태
        })

        if dept_login == "DD06":
            context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
            context['dept_login_nm'] = dept_login_nm

        else:
            context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
            context['dept_login_nm'] = dept_login_nm
            context['dept_selected'] = dept_login

    return render(request, 'jobs/JB102.html', context)


def JB102_4(request): # 직무 선택 후 직무 성과책임 조회 / 저장, 취소, 직무추가, 직무삭제 버튼 눌렀을 때

    if request.method == 'POST':
        prd_selected = request.POST["prd_selected"] # 회기
        dept_selected = request.POST['dept_selected'] # 부서
        job_type = request.POST['job_type'] # 직무 유형

        # 필터링 공통
        filtered_set = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected) # 해당 회기, 부서의 BsJobDept object들
        filtered_value = list(filtered_set.values_list('job_cd', flat=True)) # 위의 object들의 job_cd 리스트. 이를 이용해 BsJob 테이블에 접근.

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        if 'job_radio_102' in request.POST:

            radio_selected = request.POST['job_radio_102']

            # job_type의 값에 따라 다른 액션을 취함. job_type의 key도 필요하다.
            if job_type == 'all': # job_type이 all일 경우 해당 dept_cd에 해당하는 bs_job 데이터를 모두 가져옴

                # 전체 직무에 대한 필터링
                filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_cd__in=filtered_value)

                context = {
                    'title' : '직무 기본정보', # 제목
                    'prd_list': BsPrd.objects.all().order_by,
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_selected' : dept_selected,
                    'job_list' : filtered_result, # job_list는 html에 띄워줄 결과값으로, 전체를 다 가져옴
                    'job_by_list' : filtered_set, # job_by_list는 BsJobDept이고, job_by가 누군지 알 수 있다.
                    'job_type' : "all",
                    'activate' : "activate", # 직무유형 라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                    'save' : "no", #저장 버튼 deactivate
                    'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                    'radio_selected' : radio_selected,
                    'act_del' : "yes",
                    'job_type_selected' : 'latter',
                }

                if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                    context['dept_login_nm'] = dept_login_nm

                else:
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_selected'] = dept_login

            #job_type이 common일 경우 해당 dept_cd에서 job_type이 공통인 bs_job 데이터만 모두 가져옴
            elif job_type == 'common':

                # 공통 직무에 대한 필터링
                filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통", job_cd__in=filtered_value)
                filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

                context = {
                    'title' : '직무 기본정보', # 제목
                    'prd_list': BsPrd.objects.all().order_by,
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_selected' : dept_selected,
                    'job_list' : filtered_result, # job_list는 html에 띄워줄 결과값으로, 전체를 다 가져옴
                    'job_by_list' : filtered_result_job_by,
                    'activate' : "activate", #라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                    'job_type' : "common",
                    'save' : "yes", #저장 버튼 activate
                    'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                    'radio_selected' : radio_selected,
                    'act_del' : "yes", # 삭제 버튼 activate
                    'job_type_selected' : 'latter',
                }

                if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                    context['dept_login_nm'] = dept_login_nm

                else:
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_selected'] = dept_login

            #job_type이 spec일 경우 해당 dept_cd에서 job_type이 고유인 bs_job 데이터만 모두 가져옴
            elif job_type == 'unique':

                # 고유 직무에 대한 필터링
                filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유", job_cd__in=filtered_value)
                filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

                context = {
                    'title' : '직무 기본정보', # 제목
                    'prd_list': BsPrd.objects.all().order_by,
                    'prd_selected' : prd_selected,
                    'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                    'dept_selected' : dept_selected,
                    'job_list' : filtered_result, # job_list는 html에 띄워줄 결과값으로, 전체를 다 가져옴
                    'job_by_list' : filtered_result_job_by,
                    'activate' : "activate", #라디오 버튼이 작동하면 key값을 html로 넘겨주어 하단에 직무 기본사항 표시하는데 사용함.
                    'job_type' : 'unique',
                    'save' : "yes", #저장 버튼 activate
                    'job_resp_list' : BsJobResp.objects.filter(prd_cd_id=prd_selected, job_cd_id=radio_selected).order_by('job_resp_ordr'),
                    'radio_selected' : radio_selected,
                    'act_del' : "yes",
                    'job_type_selected' : 'latter',
                }

                if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                    context['dept_login_nm'] = dept_login_nm

                else:
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_selected'] = dept_login

        if 'action' in request.POST: # 저장, 삭제, 추가, 취소 버튼을 눌렀을 때

            #html에서 띄워진 결과값들을 수정한 결과를 가져옴.
            input_cd = request.POST.getlist('job_cd_hidden')
            input_job_by = request.POST.getlist('job_by_102')

            action = request.POST['action']

            # 저장
            if action == 'action1':

                if job_type == "unique":

                    for code, by in zip(input_cd, input_job_by):
                        BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=code).update(job_by=by)

                    # 고유 직무에 대한 필터링
                    filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유", job_cd__in=filtered_value)
                    filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                    filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

                elif job_type == "common":

                    for code, by in zip(input_cd, input_job_by):
                        BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=code).update(job_by=by)

                    # 공통 직무에 대한 필터링
                    filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통", job_cd__in=filtered_value)
                    filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                    filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

                context = {
                    'title' : '직무 기본정보', # 제목
                    'prd_list' : BsPrd.objects.all(), #회기 목록
                    'prd_selected' : prd_selected,
                    'job_list' : filtered_result,
                    'job_by_list' : filtered_result_job_by,
                    'dept_selected' : dept_selected,
                    'job_type' : job_type,
                    'activate' : "activate",
                    'save' : "yes", #저장 버튼 activate,
                    'job_type_selected' : 'latter',
                }

                if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                    context['dept_login_nm'] = dept_login_nm

                else:
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_selected'] = dept_login

            elif action == 'action2': #action이 삭제버튼 눌렀을 때(삭제) 라디오 버튼 값을 받는다(job_cd), job_cd와 해당 팀 정보를 이용해 삭제할 행을 삭제한다.

                # 삭제할 라디오 버튼 값을 받는다.
                radio_value = request.POST['job_radio_102']

                BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd=radio_value).delete()
                JobActivity.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=radio_value).delete()
                JobTask.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=radio_value).delete()

                filtered_set = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected) # 해당 회기, 부서의 BsJobDept object들
                filtered_value = list(filtered_set.values_list('job_cd', flat=True)) # 위의 object들의 job_cd 리스트. 이를 이용해 BsJob 테이블에 접근.

                if job_type == "unique":
                    filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유", job_cd__in=filtered_value)
                    filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                    filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)
                elif job_type == "common":
                    filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통", job_cd__in=filtered_value)
                    filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                    filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

                context = {
                    'title' : '직무 기본정보', # 제목
                    'prd_list' : BsPrd.objects.all(), #회기 목록
                    'prd_selected' : prd_selected,
                    'job_list' : filtered_result,
                    'job_by_list' : filtered_result_job_by,
                    'dept_selected' : dept_selected,
                    'job_type' : job_type,
                    'activate' : "activate",
                    'save' : "yes", #저장 버튼 activate
                    'job_type_selected' : 'latter',
                }

                if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                    context['dept_login_nm'] = dept_login_nm

                else:
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_selected'] = dept_login

            elif action == 'action3': #action이 추가버튼 눌렀을 때(추가) - 직무 선택란 보여주기

                if job_type == "unique": # 고유 직무 선택, 직무명 순으로 나열
                    job_new_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유").order_by('job_nm')
                elif job_type == "common":
                    job_new_list = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통").order_by('job_nm')

                context = {
                    'title' : '직무 기본정보', # 제목
                    'prd_list' : BsPrd.objects.all(), #회기 목록
                    'prd_selected' : prd_selected,
                    'dept_selected' : dept_selected,
                    'job_type' : job_type,
                    'activate' : "activate",
                    'save' : "yes", #저장 버튼 activate
                    'new_key' : "activate", # 새 칸을 만들어준다는 뜻으로 이걸로 화면 컨트롤 가능
                    'job_new_list' : job_new_list,
                }

                if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                    context['dept_login_nm'] = dept_login_nm

                else:
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_selected'] = dept_login

            elif action == 'action4': # 취소 버튼 눌렀을 때

                if job_type == "unique":
                    filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유", job_cd__in=filtered_value)
                    filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                    filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)
                elif job_type == "common":
                    filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통", job_cd__in=filtered_value)
                    filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                    filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

                context = {
                    'title' : '직무 기본정보', # 제목
                    'prd_list' : BsPrd.objects.all(), #회기 목록
                    'prd_selected' : prd_selected,
                    'job_list' : filtered_result,
                    'job_by_list' : filtered_result_job_by,
                    'dept_selected' : dept_selected,
                    'job_type' : job_type,
                    'activate' : "activate",
                    'save' : "yes", #저장 버튼 activate
                    'job_type_selected' : 'latter',
                }

                if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                    context['dept_login_nm'] = dept_login_nm

                else:
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                    context['dept_login_nm'] = dept_login_nm
                    context['dept_selected'] = dept_login

        context['dept_mgr_yn'] = get_dept_mgr_yn(request.user.username)

    return render(request, 'jobs/JB102.html', context)


def JB102_5(request): # 새로운 직무를 선택하고, 직무 수행자를 입력한 후 저장 버튼 누를 때 신규 직무 추가

    prd_selected = request.POST["prd_selected"] # 회기
    dept_selected = request.POST['dept_selected'] # 부서
    job_type = request.POST['job_type'] # 직무 유형

    dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
    dept_login_nm = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

    action = request.POST['action']

    if action == 'action1' : # 저장 버튼 누르면 BsJobDept object create

        try:
            new_job_cd = request.POST['new_job_cd'] # 그 해당 회기, 부서에 추가해줄 직무코드
            new_job_by = request.POST['new_job_by'] # 그 직무를 수행할 신규 직무 수행자

            BsJobDept.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=new_job_cd, job_by=new_job_by, create_dttm=now)

            # 해당 회기의 해당 부서의 직무 개수를 세고 새로운 직무를 JobTask테이블에 추가할 때 job_seq값에는 그 개수 + 1 값을 넣어준다.
            job_count = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected).count()

            # new_job_cd가 JC001이거나 JC002이거나 JC004이면,
            if new_job_cd == "JC001" or new_job_cd == "JC002" or new_job_cd == "JC004":
                JobTask.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=new_job_cd, duty_nm="책무1", task_nm="과업1",
                                       job_seq=job_count, duty_seq=1, task_seq=1)
                JobActivity.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=new_job_cd, duty_nm_id="책무1", task_nm_id="과업1",
                                           act_nm="활동1", act_prfrm_cnt=0, job_seq=job_count, duty_seq=1, task_seq=1, act_seq=1)
            else:
                JobTask.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=new_job_cd, duty_nm="책무1", task_nm="과업1",
                                       work_lv_imprt=1, work_lv_dfclt=1, work_lv_prfcn=1, work_lv_sum=3, work_grade_id="G5",
                                         prfrm_tm_ann=0, job_seq=job_count, duty_seq=1, task_seq=1)
                JobActivity.objects.create(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd_id=new_job_cd, duty_nm_id="책무1", task_nm_id="과업1",
                                           act_nm="활동1", act_prfrm_cnt=0, act_prfrm_cnt_ann=0, act_prfrm_tm_cs=0, act_prfrm_tm_ann=0,
                                             job_seq=job_count, duty_seq=1, task_seq=1, act_seq=1)

            filtered_set = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected) # 해당 회기, 부서의 BsJobDept object들
            filtered_value = list(filtered_set.values_list('job_cd', flat=True)) # 위의 object들의 job_cd 리스트. 이를 이용해 BsJob 테이블에 접근.

            if job_type == "unique":
                # 필터링된 결과 가져오기
                filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유", job_cd__in=filtered_value)
                filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

            elif job_type == "common":
                # 필터링된 결과 가져오기
                filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통", job_cd__in=filtered_value)
                filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

            context = {
                'title' : '직무 기본정보', # 제목
                'prd_list' : BsPrd.objects.all(), #회기 목록
                'prd_selected' : prd_selected,
                'job_list' : filtered_result,
                'job_by_list' : filtered_result_job_by,
                'dept_selected' : dept_selected,
                'job_type' : job_type,
                'activate' : "activate",
                'save' : "yes", #저장 버튼 activate
                'job_type_selected' : 'latter',
            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                context['dept_login_nm'] = dept_login_nm

            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_login_nm'] = dept_login_nm
                context['dept_selected'] = dept_login

        except IntegrityError as e:
            messages.error(request, f'해당 직무는 이미 존재합니다.')

            filtered_set = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected) # 해당 회기, 부서의 BsJobDept object들
            filtered_value = list(filtered_set.values_list('job_cd', flat=True)) # 위의 object들의 job_cd 리스트. 이를 이용해 BsJob 테이블에 접근.

            if job_type == "unique":
                # 필터링된 결과 가져오기
                filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유", job_cd__in=filtered_value)
                filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

            elif job_type == "common":
                # 필터링된 결과 가져오기
                filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통", job_cd__in=filtered_value)
                filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
                filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

            context = {
                'title' : '직무 기본정보', # 제목
                'prd_list' : BsPrd.objects.all(), #회기 목록
                'prd_selected' : prd_selected,
                'job_list' : filtered_result,
                'job_by_list' : filtered_result_job_by,
                'dept_selected' : dept_selected,
                'job_type' : job_type,
                'activate' : "activate",
                'save' : "yes", #저장 버튼 activate
                'job_type_selected' : 'latter',
            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                context['dept_login_nm'] = dept_login_nm

            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_login_nm'] = dept_login_nm
                context['dept_selected'] = dept_login


    elif action == 'action2' : # 취소 버튼 누르면 다시 해당 회기, 부서, 직무유형의 값들을 띄워준다.

        filtered_set = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected) # 해당 회기, 부서의 BsJobDept object들
        filtered_value = list(filtered_set.values_list('job_cd', flat=True)) # 위의 object들의 job_cd 리스트. 이를 이용해 BsJob 테이블에 접근.

        if job_type == "unique":
            # 필터링된 결과 가져오기
            filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="고유", job_cd__in=filtered_value)
            filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
            filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

        elif job_type == "common":
            # 필터링된 결과 가져오기
            filtered_result = BsJob.objects.filter(prd_cd_id=prd_selected, job_type="공통", job_cd__in=filtered_value)
            filtered_value_job_by = list(filtered_result.values_list('job_cd', flat=True))
            filtered_result_job_by = BsJobDept.objects.filter(prd_cd_id=prd_selected, dept_cd_id=dept_selected, job_cd__in=filtered_value_job_by)

        context = {
            'title' : '직무 기본정보', # 제목
            'prd_list' : BsPrd.objects.all(), #회기 목록
            'prd_selected' : prd_selected,
            'job_list' : filtered_result,
            'job_by_list' : filtered_result_job_by,
            'dept_selected' : dept_selected,
            'job_type' : job_type,
            'activate' : "activate",
            'save' : "yes", #저장 버튼 activate
            'job_type_selected' : 'latter',
        } #

        if dept_login == "DD06":
            context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
            context['dept_login_nm'] = dept_login_nm

        else:
            context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
            context['dept_login_nm'] = dept_login_nm
            context['dept_selected'] = dept_login

    context['dept_mgr_yn'] = get_dept_mgr_yn(request.user.username),

    return render(request, 'jobs/JB102.html', context)


def home(request):
    chat = 'Hello'
    name = 'inu'
    return render(request, 'jobs/BS102.html', {'user_chat':chat, 'user_name':name})


def test(request):

    context = {
            'today_date' : str(dt.datetime.today()).split()[0]
        }

    #BsPrd의 모든 object값을 original_rows라는 객체로 가져옴
    original_rows=BsPrd.objects.all()

    #datalist라는 리스트는 딕셔너리로 구성되어 있으며, 각 column name에 맞게 값들이 list화 되어있음.
    data_list = [{'prd_cd' : rows.prd_cd, 'year' : rows.year, 'turn': rows.turn, 'job_srv_str_dt' : rows.job_srv_str_dt,
               'job_srv_end_dt' : rows.job_srv_end_dt, 'job_srv_fix_dt' : rows.job_srv_fix_dt, 'prd_done_yn' : rows.prd_done_yn} for rows in original_rows]

    #datalist 리스트를 이용해 dataframe 생성
    df1 = pd.DataFrame(data_list)
    print(df1)

    if request.method == 'POST':

        #각 dynamic input을 리스트 형태로 받아온다.
        dynamic_inputs1 = request.POST.getlist('dynamic_input')
        dynamic_inputs2 = request.POST.getlist('dynamic_input2')
        dynamic_inputs3 = request.POST.getlist('dynamic_input3')
        dynamic_inputs4 = request.POST.getlist('dynamic_input4')
        dynamic_inputs5 = request.POST.getlist('dynamic_input5')
        dynamic_inputs6 = request.POST.getlist('dynamic_input6')
        dynamic_inputs7 = request.POST.getlist('dynamic_input7')

        #result는 리스트 형태로 되어 있는 각 dynamic input들을 합친 자료들이다.
        result = zip(dynamic_inputs1, dynamic_inputs2, dynamic_inputs3, dynamic_inputs4, dynamic_inputs5, dynamic_inputs6, dynamic_inputs7)

        #dynamic input들의 zip을 활용하여 new_row라는 리스트를 만들어준다.
        for i, j, k, l, m, n, o in result:
            new_row = [{'prd_cd':i, 'year':j, 'turn':k, 'job_srv_str_dt':l, 'job_srv_end_dt':m, 'job_srv_fix_dt':n, 'prd_done_yn':o}]

        # append를 못하니까, concat을 이용해서 기존 dataframe과 신규 dataframe을 합쳐준다.
        df2 = pd.concat([df1, pd.DataFrame(new_row)], ignore_index=True)
        print(df2)

        # for i, j, k, l, m, n, o in result:
        #     BsPrd.objects.create(prd_cd=i, year=j, turn=k, job_srv_str_dt=l, job_srv_end_dt=m, job_srv_fix_dt=n, prd_done_yn=o)

    return render(request, 'jobs/test.html', context)


def JB103_grid_1(request): # 회기 선택 후 Grid에 띄워주는 화면

    if request.method == 'POST':

        # 회기, 부서 데이터를 받아옴
        prd_cd_selected = request.POST['prd_cd']

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드

        try:
            dept_login_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

            # 회기, 부서 데이터에 해당하는 JobTask 값에 접근하여, dataframe 생성
            original_rows=JobTask.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함

            data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                        'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                        'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                        'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                        'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows]

            df1 = pd.DataFrame(data_list)

            # job_activity 접근
            original_rows_2=JobActivity.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login) # 나중에 prd_cd 바꿔줘야 함
            data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                        'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq,
                        'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                        'rpt_nm': rows.rpt_nm, 'job_seq':rows.job_seq, 'duty_seq':rows.duty_seq, 'task_seq':rows.task_seq, 'act_seq':rows.act_seq } for rows in original_rows_2]

            df2 = pd.DataFrame(data_list_2)


            df3 = pd.merge(df1, df2)

            # dataframe의 index를 열로 만들어줌
            df3.reset_index(inplace=True)

            # df3의 index 열을 복사하여, 새로운 열인 index_pos를 만들어줌. 이 값은 변하지 않는 값이며, grid에서 추가가 되면 999가 되는 값이다.
            df3['index_pos'] = df3['index']

            # job_nm 열을 추가
            df3['job_nm'] = df3['job_cd'].apply(lambda x: BsJob.objects.get(prd_cd=prd_cd_selected, job_cd=x).job_nm)

            # job_cd 열 삭제
            df3.drop('job_cd', axis=1, inplace=True)

            # job_seq, duty_seq, task_seq, act_seq 순으로 정렬
            df3 = df3.sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq'])

            # 데이터프레임을 JSON 형식으로 변환하여 전달
            df_json = df3.to_json(orient='records')

            context = {
                'data' : df_json,
                'prd_cd_selected' : prd_cd_selected,
                'prd' : BsPrd.objects.all(),
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_cd_selected)
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login
            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login)
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login

            return render(request, 'jobs/JB103_grid.html', context)

        except pd.errors.MergeError as e:

            messages.error(request, '해당 회기에 로그인한 부서의 정보가 없습니다.')

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                # 'error_message' : "해당 회기 및 부서에는 데이터가 없습니다.",
                'my_value' : "에러",
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_cd_selected) #마지막 회기의 부서 띄워주는게 좋을 듯
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login
            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login)
                context['dept_login_nm'] = dept_login_nm
                context['dept_cd_selected'] = dept_login

            return render(request, 'jobs/JB103_grid.html', context)

        except ObjectDoesNotExist as e:

            messages.error(request, '해당 회기에 로그인한 부서가 없습니다.')

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                'my_value' : "에러",
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            if dept_login == "DD06":
                    context['team_list'] = BsDept.objects.filter(prd_cd=prd_cd_selected) #마지막 회기의 부서 띄워주는게 좋을 듯
                    context['dept_cd_selected'] = dept_login
            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login)
                context['dept_cd_selected'] = dept_login

            return render(request, 'jobs/JB103_grid.html', context)


def JB103_grid_2(request): # 부서 선택 후 조회 화면(경영기획팀만 해당)

    if request.method == 'POST':

       # 회기, 부서 데이터를 받아옴
        prd_cd_selected = request.POST['prd_cd_selected']

        dept_cd_selected = request.POST['team_name']

        # 회기, 부서 데이터에 해당하는 JobTask 값에 접근하여, dataframe 생성
        original_rows=JobTask.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함

        data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm,
                    'task_nm': rows.task_nm, 'task_prsn_chrg': rows.task_prsn_chrg, 'work_lv_imprt': rows.work_lv_imprt,
                    'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                    'work_grade': rows.work_grade_id, 'work_attrbt': rows.work_attrbt,
                    'prfrm_tm_ann': rows.prfrm_tm_ann } for rows in original_rows]

        df1 = pd.DataFrame(data_list)

        # job_activity 접근
        original_rows_2=JobActivity.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_cd_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm_id,
                    'task_nm': rows.task_nm_id, 'act_nm': rows.act_nm, 'act_prsn_chrg': rows.act_prsn_chrg, 'act_prfrm_freq': rows.act_prfrm_freq,
                    'act_prfrm_cnt_ann': rows.act_prfrm_cnt_ann, 'act_prfrm_tm_cs': rows.act_prfrm_tm_cs, 'act_prfrm_tm_ann': rows.act_prfrm_tm_ann,
                    'rpt_nm': rows.rpt_nm, 'job_seq':rows.job_seq, 'duty_seq':rows.duty_seq, 'task_seq':rows.task_seq, 'act_seq':rows.act_seq } for rows in original_rows_2]

        df2 = pd.DataFrame(data_list_2)

        try:

            df3 = pd.merge(df1, df2)

            # dataframe의 index를 열로 만들어줌
            df3.reset_index(inplace=True)

            # df3의 index 열을 복사하여, 새로운 열인 index_pos를 만들어줌. 이 값은 변하지 않는 값이며, grid에서 추가가 되면 999가 되는 값이다.
            df3['index_pos'] = df3['index']

            # job_nm 열을 추가
            df3['job_nm'] = df3['job_cd'].apply(lambda x: BsJob.objects.get(prd_cd=prd_cd_selected, job_cd=x).job_nm)

            # job_cd 열 삭제
            df3.drop('job_cd', axis=1, inplace=True)

            # job_seq, duty_seq, task_seq, act_seq 순으로 정렬
            df3 = df3.sort_values(['job_seq', 'duty_seq', 'task_seq', 'act_seq'])

            # 데이터프레임을 JSON 형식으로 변환하여 전달
            df_json = df3.to_json(orient='records')

            context = {
                'data' : df_json,
                'prd_cd_selected' : prd_cd_selected,
                'prd' : BsPrd.objects.all(),
                'team_list' : BsDept.objects.filter(prd_cd=prd_cd_selected),
                'dept_cd_selected' : dept_cd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            return render(request, 'jobs/JB103_grid.html', context)

        except pd.errors.MergeError as e:

            messages.error(request, '해당 회기에 선택한 부서의 직무정보가 없습니다.')

            context = {
                'prd' : BsPrd.objects.all(),
                'prd_cd_selected' : prd_cd_selected,
                # 'error_message' : "해당 회기 및 부서에는 데이터가 없습니다.",
                'my_value' : "에러",
                'team_list' : BsDept.objects.filter(prd_cd=prd_cd_selected),
                'dept_cd_selected' : dept_cd_selected,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            return render(request, 'jobs/JB103_grid.html', context)


def JB108_1(request): # 직무현황 제출 - 회기 선택

    if request.method == 'POST':

        try:
            #html에서 회기 선택
            prd_selected = request.POST["prd_selected"]
            key = request.POST["key_prd_select"]
            # print(prd_selected)
            prd_done_yn = BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn
            dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드

            submit_yn = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_login).job_details_submit_yn

            if prd_done_yn == 'N':
                    if submit_yn == 'N':
                        confirm_text = "직무현황을 제출하지 않았습니다."
                    else:
                        confirm_text = "직무현황을 제출한 상태입니다."
            else:
                confirm_text = "마감된 회기입니다."

            context = {
                'title' : '직무 현황제출',
                'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                'key' : key, # 회기 바꿨다는 뜻,
                'prd_selected' : prd_selected,
                'prd_done_yn' : prd_done_yn,
                'modified' : "n",
                'confirm_text' : confirm_text,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                context['dept_selected'] = dept_login
            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_selected'] = dept_login

        except ObjectDoesNotExist as e:

            messages.error(request, '해당 회기에 로그인한 부서가 없습니다.')

            context = {
                'title' : '직무 현황제출',
                'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
                'prd_selected' : prd_selected,
                'modified' : "n",
                'confirm_text' : "해당 회기에 부서 정보가 없습니다.",
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            }

            if dept_login == "DD06":
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
                context['dept_selected'] = dept_login
            else:
                context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
                context['dept_selected'] = dept_login

    return render(request, 'jobs/JB108.html', context)


def JB108_2(request): # 직무현황 제출 부서 선택받고 그 값을 html로 넘겨줌. 경영기획팀만 해당.

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"]
        dept_selected = request.POST["dept_selected"]
        prd_done_yn = BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn
        submit_yn = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_selected).job_details_submit_yn

        if prd_done_yn == 'N':
            if submit_yn == 'N':
                confirm_text = "직무현황을 제출하지 않았습니다."
            else:
                confirm_text = "직무현황을 제출한 상태입니다."
        else:
            confirm_text = "마감된 회기입니다."

        context = {
            'title' : '직무 현황제출',
            'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
            'team_list' : BsDept.objects.filter(prd_cd=prd_selected), # 선택한 회기의 부서 목록
            'dept_selected' : dept_selected,
            'prd_selected' : prd_selected,
            'prd_done_yn' : prd_done_yn,
            'submit_yn' : submit_yn,
            'confirm_text' : confirm_text,
            'modified' : "n",
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

    return render(request, 'jobs/JB108.html', context)


def JB108_3(request): # 직무현황 제출/제출취소 버튼 누르고 난 후

    if request.method == 'POST':

        prd_selected = request.POST["prd_selected"]
        dept_selected = request.POST["dept_selected"]
        prd_done_yn = BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn
        submit_yn = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_selected).job_details_submit_yn

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드

        action = request.POST["action"]

        if action == 'action1':

            BsDept.objects.filter(prd_cd_id=prd_selected, dept_cd=dept_selected).update(job_details_submit_yn="Y")
            confirm_text = "직무현황을 제출하였습니다."
            submit_yn = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_selected).job_details_submit_yn

        elif action == 'action2':

            BsDept.objects.filter(prd_cd_id=prd_selected, dept_cd=dept_selected).update(job_details_submit_yn="N")
            confirm_text = "직무현황 제출을 취소하였습니다."
            submit_yn = BsDept.objects.get(prd_cd=prd_selected, dept_cd=dept_selected).job_details_submit_yn

        context = {
            'title' : '직무 현황제출',
            'prd_list' : BsPrd.objects.all().order_by, # 회기 리스트. 마지막 회기가 디폴트로 뜰 것임
            'prd_selected' : prd_selected,
            'prd_done_yn' : prd_done_yn,
            'submit_yn' : submit_yn,
            'confirm_text' : confirm_text,
            'modified' : "y",
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if dept_login == "DD06":
            context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected)
            context['dept_selected'] = dept_selected
        else:
            context['team_list'] = BsDept.objects.filter(prd_cd=prd_selected, dept_cd=dept_login)
            context['dept_selected'] = dept_login

    return render(request, 'jobs/JB108.html', context)


def JB109_1(request): # 업무량 분석화면 - 회기 선택 후 선택한 회기를 넘겨준다.

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd']

        context = {
            'prd_list' : BsPrd.objects.all(),
            'title' : '업무량 분석', # 제목
            'prd_cd_selected' : prd_cd_selected,
            'dept_selected_key' : 'former',
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

    return render(request, 'jobs/JB109.html', context)


def JB109_2(request): # 업무량 분석화면 - 탭 선택 후 선택한 탭을 넘겨준다.

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd_selected']
        span_name = request.POST.get('span_name', '') # 탭 선택에 따라 초기 데이터와 키 값을 보내준다.

        context = {
            'prd_list' : BsPrd.objects.all(),
            'title' : '업무량 분석', # 제목
            'prd_cd_selected' : prd_cd_selected,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if span_name == 'span1' : # 업무량 분석 탭 선택했을 때

            context.update({
                'tab' : 'tab1',
                'dept_selected_key' : 'former',
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            })

        if span_name == 'span1_1' : # 조정 업무량 분석 탭 선택했을 때

            context.update({
                'tab' : 'tab1_1',
                'dept_selected_key' : 'former',
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            })

        if span_name == 'span2': # 업무량 구성비 탭 선택했을 때

            context.update({
                'tab' : 'tab2',
                'dept_selected_key' : 'former',
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            })

        if span_name == 'span2_1': # 조정 업무량 구성비 탭 선택했을 때

            context.update({
                'tab' : 'tab2_1',
                'dept_selected_key' : 'former',
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            })

        if span_name == 'span3': # 적정인력 산정 탭 선택했을 때

            # 해당 회기의 dept_domain 리스트를 만들어준다. 해당 회기의 BsDeptGrpDomain 테이블에서 dept_domain값들을 가져오고 중복 제거한다.
            domain_list = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected).values_list('dept_domain', flat=True).distinct()

            context.update({
                'tab' : 'tab3',
                'domain_selected_key' : 'former',
                'domain_list' : domain_list,
            })

        if span_name == 'span3_1': # 적정인력 산정 탭 선택했을 때

            # 해당 회기의 dept_domain 리스트를 만들어준다. 해당 회기의 BsDeptGrpDomain 테이블에서 dept_domain값들을 가져오고 중복 제거한다.
            domain_list = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected).values_list('dept_domain', flat=True).distinct()

            context.update({
                'tab' : 'tab3_1',
                'domain_selected_key' : 'former',
                'domain_list' : domain_list,
            })

        if span_name == 'span4':

            # 해당 회기의 dept_domain 리스트를 만들어준다. 해당 회기의 BsDeptGrpDomain 테이블에서 dept_domain값들을 가져오고 중복 제거한다.
            domain_list = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected).values_list('dept_domain', flat=True).distinct()

            context.update({
                'tab' : 'tab4',
                'domain_selected_key' : 'former',
                'domain_list' : domain_list,
            })

    return render(request, 'jobs/JB109.html', context)


def JB109_3(request): # 업무량 분석화면 - 부서 선택한 후 - 업무량 분석 / 업무량 구성비

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd_selected']
        tab = request.POST['tab']
        dept_selected = request.POST['dept_selected']

        context = {

            'prd_list' : BsPrd.objects.all(),
            'title' : '업무량 분석', # 제목
            'prd_cd_selected' : prd_cd_selected,
            'tab' : tab,
            'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            'dept_selected' : dept_selected,
            'dept_selected_nm' : BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_selected).dept_nm,
            'dept_selected_key' : 'latter',
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        # if tab == 'tab1': # 업무량 분석 탭 선택했을 때
        # tab이 tab1이거나 a일때
        if tab == 'tab1' or tab == 'tab1_1':

            std_wrk_tm = float(BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_tm) # 표준근무시간

            work_grade_list = ['G1', 'G2', 'G3', 'G4', 'G5'] # 업무등급 리스트. 이 리스트는 변하지 않는다.

            mbr_of_dept = BsMbr.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected) # 해당 부서의 부서원 목록

            # mbr_of_dept에서 ttl_nm이 '팀장'인 object는 제외해준다. 왜냐하면 PO산정때 필요없기 때문이다. 어차피 팀장 업무는 다 업무수행시간 0시간임.
            mbr_of_dept = mbr_of_dept.exclude(ttl_nm='팀장')

            analysis = pd.DataFrame({'work_grade':work_grade_list}) # 분석 결과를 담을 데이터프레임

            if tab == 'tab1': # 업무량 분석 탭 선택했을 때

                task_target = JobTask.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected)
                data_list = [{'task_nm' : rows.task_nm, 'work_grade' : rows.work_grade_id, 'prfrm_tm_ann':rows.prfrm_tm_ann} for rows in task_target]
                df1 = pd.DataFrame(data_list)

            elif tab == 'tab1_1': # 조정 업무량 분석 탭 선택했을 때

                task_target = JobTaskAdj.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected)
                data_list = [{'task_nm' : rows.task_nm, 'work_grade' : rows.work_grade_id, 'prfrm_tm_ann':rows.prfrm_tm_ann} for rows in task_target]
                df1 = pd.DataFrame(data_list)

            task_count = [] # 과업수
            prfrm_tm_ann = [] # 연간 수행 시간
            m_result = [] # 인력 산정 결과
            po = [] #po
            workload = BsWorkGrade.objects.filter(prd_cd=prd_cd_selected) # 업무량 가중치 object
            workload_wt = [float(n.workload_wt) for n in workload ] # 업무량 가중치 list(float)

            try:

                for grade in work_grade_list: # 업무등급 각각에 대하여 작업.

                    cnt = len(df1.loc[df1['work_grade'] == grade]) # 그 grade에 해당하는 task의 수를 센다.
                    task_count.append(cnt) # task_count라는 리스트에 cnt를 추가. 즉 그 grade에 해당하는 task의 수이다.

                    df2 = df1[df1['work_grade'] == grade] # JobTask 혹은 JobTaskAdj을 이용해 만든 df1에서 그 grade에 해당하는 task만 추출하여 df2 만들기

                    prfrm_tm_ann.append(float(df2['prfrm_tm_ann'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann의 합을 구한다.

                    # 인력 산정
                    m_result.append(round(float(df2['prfrm_tm_ann'].sum())/std_wrk_tm, 1))

                    # 그 grade에 해당하는 pos_nm 리스트를 만든다.
                    pos_nm_of_grade = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, work_grade_id=grade).values_list('pos_nm', flat=True)

                    cnt_grade = 0 # 그 grade에 해당하는 po

                    # 위에서 만들어준 pos_nm 리스트를 갖고 Mbr에 접근
                    for row in mbr_of_dept:

                        if row.pos_nm in pos_nm_of_grade: # 직위 이름이 pos_nm_gr_grade(그 grade에 해당하는 pos_nm 리스트)에 있으면
                            cnt_grade = cnt_grade+1 # PO를 1 늘려준다.

                    po.append(cnt_grade)

                analysis['task_count'] = task_count # 과업수
                analysis['prfrm_tm_ann'] = prfrm_tm_ann # 연간 업무량
                analysis['po_result'] = m_result # 인력 산정
                analysis['po'] = po # 업무등급별 PO
                analysis['workload_wt'] = workload_wt # 업무량 가중치
                analysis['po_cal'] = round(analysis['po'] * analysis['workload_wt'], 2) # 환산 PO
                analysis['prfrm_tm_ann_cal'] = analysis['prfrm_tm_ann'] * analysis['workload_wt']
                analysis['po_right'] = round(analysis['prfrm_tm_ann_cal']/std_wrk_tm, 1) # 적정 인력 산정
                analysis['overless'] = round(analysis['po_cal']-analysis['po_right'], 1)

                sum_1 = analysis['task_count'].sum()
                sum_2 = round(analysis['prfrm_tm_ann'].sum(), 1)
                sum_3 = round(analysis['po_result'].sum(), 1)
                sum_4 = round(analysis['po'].sum(), 1)
                sum_5 = round(analysis['po_cal'].sum(), 2)
                # sum_6 = round(round(analysis['prfrm_tm_ann_cal'].sum(), 2), 1)
                sum_6 = round(analysis['prfrm_tm_ann_cal'].sum(), 1)
                # sum_6 = analysis['prfrm_tm_ann_cal'].sum()
                sum_7 = round(analysis['po_right'].sum(), 1)
                sum_8 = round(analysis['overless'].sum(), 1)

                # anaylsis['prfrm_tm_ann_cal']을 반올림
                analysis['prfrm_tm_ann_cal'] = analysis['prfrm_tm_ann_cal'].apply(lambda x: round(x, 1))

                if sum_8 > BsWlOvSht.objects.get(prd_cd=prd_cd_selected).ov_sht_max:
                    overless = "여유"
                elif sum_8 <= BsWlOvSht.objects.get(prd_cd=prd_cd_selected).ov_sht_max and sum_8 >= BsWlOvSht.objects.get(prd_cd=prd_cd_selected).ov_sht_min:
                    overless = "적정"
                elif sum_8:
                    overless = "부족"

                context.update({
                        'analysis' : analysis,
                        'sum_1' : sum_1,
                        'sum_2' : sum_2,
                        'sum_3' : sum_3,
                        'sum_4' : sum_4,
                        'sum_5' : sum_5,
                        'sum_6' : sum_6,
                        'sum_7' : sum_7,
                        'sum_8' : sum_8,
                        'overless' : overless,
                        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                    })

            except KeyError as e:

                messages.error(request, '해당 회기에 선택한 부서의 조정 업무량 분석 정보가 없습니다.')

        if tab == 'tab2' or tab == 'tab2_1': # 업무량 구성비 탭 선택했을 때

            # JobTask 테이블을 이용해 해당 회기의 해당 부서에 대한 dataframe 생성
            if tab == 'tab2':
                job_task = JobTask.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected)
            elif tab == 'tab2_1':
                job_task = JobTaskAdj.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected)

            data_list = [{'job_cd' : rows.job_cd_id, 'duty_nm' : rows.duty_nm, 'task_nm' : rows.task_nm,
                          'work_lv_imprt' : rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn,
                           'work_lv_sum' : rows.work_lv_sum, 'work_grade' : rows.work_grade_id, 'prfrm_tm_ann':rows.prfrm_tm_ann,
                             'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq } for rows in job_task]

            df1 = pd.DataFrame(data_list) # JobTask 테이블에서 필요한 rough data를 가져왔다.

            try:
                # job_cd, duty_nm, job_seq, duty_seq열을 가져와서 중복을 제거한 후, job_seq, duty_seq 순으로 정렬한다. 이후 인덱스 리셋한다.
                df2 = df1[['job_cd', 'duty_nm', 'job_seq', 'duty_seq']].drop_duplicates().sort_values(['job_seq', 'duty_seq']).reset_index(drop=True)

                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행 수를 센다. 그러면 task의 개수이다. 이를 task_cnt 열로 추가한다.
                # df2에 대해서 for문을 활용하여 task_cnt열 데이터를 추가해준다. 먼저 task_cnt열부터 만들어놓고 시작해야 한다.
                df2['task_cnt'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'task_cnt'] = len(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])])

                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구해서 그 값을 job_prfrm_tm_ann 열로 추가한다.
                # 우선 df1의 prfrm_tm_ann열 중에서 Null값을 0으로 치환한다.
                df1['prfrm_tm_ann'] = df1['prfrm_tm_ann'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_tm_ann열 데이터를 추가해준다. 먼저 duty_tm_ann열 만들어놓고 시작해야 한다.
                df2['duty_tm_ann'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'duty_tm_ann'] = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum()

                # df2의 duty_tm_ann열의 합을 구한 후, 그 값을 sum_duty_tm_ann이라는 변수에 저장한다.
                sum_duty_tm_ann = df2['duty_tm_ann'].sum()

                # df2에 각 duty의 ratio열인 duty_ratio을 추가한다. for문을 활용한다.
                df2['duty_ratio'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'duty_ratio'] = round((df2.loc[i, 'duty_tm_ann']/sum_duty_tm_ann)*100, 1)

                # 리스트를 만드는데, 앞에서는 duty_ratio를 반올림해서 만들었다. 이번에는 반올림하지 않고 만든다.
                duty_ratio_list = [(x/sum_duty_tm_ann)*100 for x in df2['duty_tm_ann']]

                # 이 리스트의 합을 구한다. 나중에 sum값으로 사용할 값이다.
                sum_duty_ratio = round(sum(duty_ratio_list), 1)

                ### 바꿔야 하는 부분 ###
                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_imprt를 해당 duty의 해당 duty_nm의 전체 prfrm_tm_ann 합에서 그 행의 비율을 곱한 값의 합을 구해서 duty_imprt열로 추가한다.
                # 우선 df1의 work_lv_imprt열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_imprt'] = df1['work_lv_imprt'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_imprt열 데이터를 추가해준다. 먼저 duty_imprt열 만들어놓고 시작해야 한다.
                df2['duty_imprt'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_imprt'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].mean(), 1)


                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_imprt를 곱한 값을 구하여 duty_imprt열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_imprt열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_imprt'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_imprt열을 구한다.
                        # df2.loc[i, 'duty_imprt'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_imprt 리스트 생성
                        work_lv_imprt_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].tolist()
                        # work_lv_imprt_list의 자료형을 float으로 변경
                        work_lv_imprt_list = [float(x) for x in work_lv_imprt_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나눈 값으로 변경
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_imprt열에 추가한다.
                        df2.loc[i, 'duty_imprt'] = round(sum([a*b for a, b in zip(work_lv_imprt_list, prfrm_tm_ann_ratio_list)]), 1)


                        # prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']/sum_prfrm_tm_ann
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        # prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_imprt열에 추가한다.
                        # df2.loc[i, 'duty_imprt'] = round(sum([a*b for a, b in zip(work_lv_imprt_list, prfrm_tm_ann_ratio_list)])*100, 1)


                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_dfclt 평균의 소수 첫째자리까지 반올림을 구해서 duty_dfclt열로 추가한다.
                # 우선 df1의 work_lv_dfclt열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_dfclt'] = df1['work_lv_dfclt'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_dfclt열 데이터를 추가해준다. 먼저 duty_dfclt열 만들어놓고 시작해야 한다.
                df2['duty_dfclt'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_dfclt'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].mean(), 1)

                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_dfclt를 곱한 값을 구하여 duty_dfclt열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_dfclt열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_dfclt'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_dfclt열을 구한다.
                        # df2.loc[i, 'duty_dfclt'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_dfclt 리스트 생성
                        work_lv_dfclt_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].tolist()
                        # work_lv_dfclt_list의 자료형을 float으로 변경
                        work_lv_dfclt_list = [float(x) for x in work_lv_dfclt_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_dfclt열에 추가한다.
                        df2.loc[i, 'duty_dfclt'] = round(sum([a*b for a, b in zip(work_lv_dfclt_list, prfrm_tm_ann_ratio_list)]), 1)



                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_prfcn 평균의 소수 첫째자리까지 반올림을 구해서 duty_prfcn열로 추가한다.
                # 우선 df1의 work_lv_prfcn열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_prfcn'] = df1['work_lv_prfcn'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_prfcn열 데이터를 추가해준다. 먼저 duty_prfcn열 만들어놓고 시작해야 한다.
                df2['duty_prfcn'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_prfcn'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].mean(), 1)

                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_prfcn을 곱한 값을 구하여 duty_prfcn열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_prfcn열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_prfcn'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_prfcn열을 구한다.
                        # df2.loc[i, 'duty_prfcn'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_prfcn 리스트 생성
                        work_lv_prfcn_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].tolist()
                        # work_lv_prfcn_list의 자료형을 float으로 변경
                        work_lv_prfcn_list = [float(x) for x in work_lv_prfcn_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나눈 값으로 변경
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_prfcn열에 추가한다.
                        df2.loc[i, 'duty_prfcn'] = round(sum([a*b for a, b in zip(work_lv_prfcn_list, prfrm_tm_ann_ratio_list)]), 1)



                # work_lv_mean열 추가. duty_imprt, duty_dfclt, duty_prfcn의 평균을 구해서 추가한다.
                df2['work_lv_mean'] = round((df2['duty_imprt'] + df2['duty_dfclt'] + df2['duty_prfcn']), 1)

                # # 바꿔야 되는 부분
                g1_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G1').work_lv_min
                g1_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G1').work_lv_max
                g2_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G2').work_lv_min
                g2_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G2').work_lv_max
                g3_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G3').work_lv_min
                g3_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G3').work_lv_max
                g4_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G4').work_lv_min
                g4_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G4').work_lv_max
                g5_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G5').work_lv_min
                g5_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G5').work_lv_max

                # df2에서 각 행에 대해 work_lv_mean열에 대해 조건문을 적용하여 G1, G2, G3, G4, G5를 부여한다.
                # for문을 이용해 각 row에 접근해서 바꿔준다.
                for i in range(len(df2)):

                    # 그 행의 job_cd가 JC001이면 work_lv_mean열에 빈칸을 부여한다.
                    if df2.loc[i, 'job_cd'] == 'JC001':
                        df2.loc[i, 'work_lv_mean'] = ''

                    # 그 행의 job_cd가 JC001이 아니면
                    else:

                        if round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) > g1_max:
                            df2.loc[i, 'work_lv_mean'] = 'G1'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g1_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g1_min:
                            df2.loc[i, 'work_lv_mean'] = 'G1'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g2_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g2_min:
                            df2.loc[i, 'work_lv_mean'] = 'G2'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g3_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g3_min:
                            df2.loc[i, 'work_lv_mean'] = 'G3'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g4_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g4_min:
                            df2.loc[i, 'work_lv_mean'] = 'G4'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g5_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g5_min:
                            df2.loc[i, 'work_lv_mean'] = 'G5'

                # BsJob 테이블을 이용해 해당 job_cd의 job_nm 열을 추가할 것이다.
                job_nm_list = [BsJob.objects.get(prd_cd=prd_cd_selected, job_cd=x).job_nm for x in df2['job_cd']]
                df2['job_nm'] = job_nm_list

                # job_nm열을 job_cd열 뒤로 이동
                cols = df2.columns.tolist()
                cols = cols[:1] + cols[-1:] + cols[1:-1]
                df2 = df2[cols]

                context.update({
                    'analysis' : df2,
                    'sum_duty_tm_ann' : sum_duty_tm_ann,
                    'sum_duty_ratio' : sum_duty_ratio,
                })

            except KeyError as e:
                messages.error(request, '해당 회기에 선택한 부서의 조정 업무량 구성비 정보가 없습니다.')

    return render(request, 'jobs/JB109.html', context)


def JB109_4(request): # 업무량 분석 - 조직그룹 선택한 후 - 적정인력 산정 / 인력산정 결과

    if request.method == 'POST':

        prd_cd_selected = request.POST['prd_cd_selected']
        tab = request.POST['tab']
        domain_selected = request.POST['domain_selected']

        # 해당 회기의 dept_domain 리스트를 만들어준다. 해당 회기의 BsDeptGrpDomain 테이블에서 dept_domain값들을 가져오고 중복 제거한다.
        domain_list = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected).values_list('dept_domain', flat=True).distinct()

        # BsDeptGrpDomain 테이블로부터 dept_domain, dept_grp_nm, domain_seq, grp_seq를 가져온다.
        # BsDeptGrp 테이블로부터 dept_domain, dept_grp_nm, dept_cd, dept_seq를 가져온다.
        # 가져와서 dataframe 각각 만들어주고 merge한다.
        target = BsDeptGrpDomain.objects.filter(prd_cd_id=prd_cd_selected, dept_domain=domain_selected)
        data_list = [{'dept_domain' : rows.dept_domain, 'dept_grp_nm' : rows.dept_grp_nm, 'domain_seq': rows.domain_seq,
                    'grp_seq': rows.grp_seq} for rows in target]
        df1 = pd.DataFrame(data_list)
        target2 = BsDeptGrp.objects.filter(prd_cd_id=prd_cd_selected, dept_domain_id=domain_selected)
        data_list2 = [{'dept_domain' : rows.dept_domain_id, 'dept_grp_nm' : rows.dept_grp_nm_id, 'dept_cd': rows.dept_cd_id,
                    'dept_seq': rows.dept_seq} for rows in target2]
        df2 = pd.DataFrame(data_list2)
        df3 = pd.merge(df1, df2)

        if tab == 'tab3' or tab == 'tab3_1' : # 적정 인력 산정 탭 선택했을 때

            # # df3에 dept_po 열을 추가한다. dept_po 데이터는 BsDept 테이블에서 해당 부서코드의 dept_po 값을 가져온다.
            # df3['dept_po'] = df3['dept_cd'].apply(lambda x: BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=x).dept_po)
            # df3에 dept_po열을 추가한다. dept_po 데이터는 BsMbr에서 그 부서에 해당하는 object 중 ttl_nm이 '팀장'인 것을 제외한 object의 수를 가져온다.
            df3['dept_po'] = df3['dept_cd'].apply(lambda x: len(BsMbr.objects.filter(prd_cd=prd_cd_selected, dept_cd=x).exclude(ttl_nm='팀장')))

            work_grade_list = ['G1', 'G2', 'G3', 'G4', 'G5'] # 업무등급 리스트

            # df3에 prfrm_tm_ann 열을 추가한다. df3의 각 행에 대하여(for문 활용), 그 행으 dept_cd에 해당하는 JobTask 테이블의 prfrm_tm_ann 값을 가져온다.
            # JobTask object들의 prfrm_tm_ann 값을 더해주면 그 행의 prfrm_tm_ann이다.
            for i in range(len(df3)):

                sum_prfrm_tm_ann = 0

                if tab == 'tab3':

                    sum_target = JobTask.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=df3.iloc[i]['dept_cd'])

                elif tab == 'tab3_1':

                    sum_target = JobTaskAdj.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=df3.iloc[i]['dept_cd'])

                for rows in sum_target:
                    if rows.prfrm_tm_ann == None:
                        rows.prfrm_tm_ann = 0
                    sum_prfrm_tm_ann = sum_prfrm_tm_ann + rows.prfrm_tm_ann

                df3.loc[i, 'prfrm_tm_ann'] = float(sum_prfrm_tm_ann)

            #df3에 po_nec 열을 추가한다. df3의 prfrm_tm_ann열을 BsStdWrkTm 테이블의 std_wrk_tm으로 나눈 값을 넣어준다.
            std_wrk_tm = float(BsStdWrkTm.objects.get(prd_cd=prd_cd_selected).std_wrk_tm)

            df3['po_nec'] = round(df3['prfrm_tm_ann']/std_wrk_tm, 1)

            # df3에 환산 업무량, 환산 PO, 적정인력 산정, 과부족 열을 추가한다.
            for i in range(len(df3)): # 각 부서에 대해서

                if tab == 'tab3':
                    task_list = JobTask.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=df3.iloc[i]['dept_cd']) # 그 부서의 task_list를 가져온다.
                elif tab == 'tab3_1':
                    task_list = JobTaskAdj.objects.filter(prd_cd_id=prd_cd_selected, dept_cd_id=df3.iloc[i]['dept_cd']) # 그 부서의 task_list를 가져온다.

                sum_tm_cal = 0 # 그 부서의 환산 업무량 초기값을 0으로 설정한다.
                sum_po_cal = 0 # 그 부서의 환산 PO 초기값을 0으로 설정한다.

                mbr_of_dept = BsMbr.objects.filter(prd_cd=prd_cd_selected, dept_cd=df3.iloc[i]['dept_cd']) # 해당 부서의 부서원 목록

                # mbr_of_dept에서 ttl_nm이 '팀장'인 object는 제외해준다. 왜냐하면 PO산정때 필요없기 때문이다. 어차피 팀장 업무는 다 업무수행시간 0시간임.
                mbr_of_dept = mbr_of_dept.exclude(ttl_nm='팀장')

                # task_list를 이용해서 dataframe을 만든다.
                data_list = [{'task_nm' : rows.task_nm, 'work_grade' : rows.work_grade_id, 'prfrm_tm_ann':rows.prfrm_tm_ann} for rows in task_list]
                df3_1 = pd.DataFrame(data_list)

                df4 = pd.DataFrame({'work_grade':work_grade_list}) # 분석 결과를 담을 데이터프레임

                for grade in work_grade_list: # 업무등급 각각에 대하여 작업.

                    # # df3_1에서 그 grade에 해당하는 task의 prfrm_tm_ann을 더해서, 거기에 업무량 가중치를 곱한다.
                    # sum_tm_cal = sum_tm_cal + float(df3_1.loc[df3_1['work_grade'] == grade, 'prfrm_tm_ann'].sum()) * float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade=grade).workload_wt)

                    # df3_1에서 그 grade에 해당하는 task의 prfrm_tm_ann을 더해서 df4에 추가한다.
                    sum_prfrm_tm_ann = float(df3_1.loc[df3_1['work_grade'] == grade, 'prfrm_tm_ann'].sum())
                    df4.loc[df4['work_grade'] == grade, 'prfrm_tm_ann'] = sum_prfrm_tm_ann

                    # 그 gradede에 해당하는 가중치를 구하여 df4에 추가한다.
                    df4.loc[df4['work_grade'] == grade, 'workload_wt'] = float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade=grade).workload_wt)

                    # 그 grade에 해당하는 pos_nm 리스트를 만든다.
                    pos_nm_of_grade = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, work_grade_id=grade).values_list('pos_nm', flat=True)

                    cnt_grade = 0 # 그 grade에 해당하는 po

                    # 위에서 만들어준 pos_nm 리스트를 갖고 Mbr에 접근
                    for row in mbr_of_dept:

                        if row.pos_nm in pos_nm_of_grade: # 직위 이름이 pos_nm_gr_grade(그 grade에 해당하는 pos_nm 리스트)에 있으면
                            cnt_grade = cnt_grade + 1 # PO를 1 늘려준다.

                    # sum_po_cal = sum_po_cal + cnt_grade * float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade=grade).workload_wt)

                    # 그 grade에 해당하는 po를 df4에 추가한다.
                    df4.loc[df4['work_grade'] == grade, 'po'] = cnt_grade

                    # 그 grade의 환산 업무량을 구하여 df4에 추가한다. 환산 업무량은 prfrm_tm_ann열과 workload_wt열을 곱한 것이다.
                    df4.loc[df4['work_grade'] == grade, 'prfrm_tm_ann_cal'] = df4.loc[df4['work_grade'] == grade, 'prfrm_tm_ann'] * df4.loc[df4['work_grade'] == grade, 'workload_wt']

                    # 적정인력 산정 열 추가. 적정 인력 산정은 각 grade에 대해서 환산 업무량을 표준 근무시간으로 나눈 값이다. 반올림해준다.
                    df4.loc[df4['work_grade'] == grade, 'po_appr'] = round(df4.loc[df4['work_grade'] == grade, 'prfrm_tm_ann_cal']/std_wrk_tm, 1)

                # sum_tm_cal = round(sum_tm_cal, 1) # 환산 업무량 열 추가

                # df3.loc[i, 'tm_cal'] = sum_tm_cal # 환산 업무량 열 추가

                # 환산 업무량 열 추가, df4를 이용한다. df4의 prfrm_tm_ann과 workload_wt를 곱해서 그 값을 더한다.
                df3.loc[i, 'tm_cal'] = round((df4['prfrm_tm_ann'] * df4['workload_wt']).sum(), 1)

                # 환산 PO열 추가. 환산 PO = PO * 업무량 가중치
                # df3.loc[i, 'po_cal'] = round(sum_po_cal, 2) # 환산 PO열 추가
                df3.loc[i, 'po_cal'] = round(round((df4['po'] * df4['workload_wt']), 2).sum(), 2)

                # 적정인력 산정 열 추가 -
                # df3.loc[i, 'po_appr'] = round(df3.loc[i, 'tm_cal']/std_wrk_tm, 1) # 적정인력 산정 열 추가
                df3.loc[i, 'po_appr'] = round(df4['po_appr'].sum(), 1)

                # 과부족 열 추가
                df3.loc[i, 'overless'] = round(df3.loc[i, 'po_cal'] - df3.loc[i, 'po_appr'], 1)

                # 부서명 열 추가
                df3['dept_nm'] = df3['dept_cd'].apply(lambda x: BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=x).dept_nm)

                ########################################### 변경전 ###########################################
                # # 각 task_list의 행에 대하여, task_list의 work_grade를 보고, work_grade 에 따른 업무량 가중치를 prfrm_tm_ann에 곱한다. 그리고 그 값을 sum_tm_cal에 더해준다.
                # # task_list는 그 부서의 task들을 가져온 것이다.
                # for rows in task_list:

                #     if rows.work_grade_id == None:
                #         weight = 0.0

                #     elif rows.work_grade_id == 'G1':
                #         weight = BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G1').workload_wt

                #     elif rows.work_grade_id == 'G2':
                #         weight = BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G2').workload_wt

                #     elif rows.work_grade_id == 'G3':
                #         weight = BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G3').workload_wt

                #     elif rows.work_grade_id == 'G4':
                #         weight = BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G4').workload_wt

                #     elif rows.work_grade_id == 'G5':
                #         weight = BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G5').workload_wt

                #     if rows.prfrm_tm_ann == None:
                #         rows.prfrm_tm_ann = 0.0
                #         sum_tm_cal = sum_tm_cal + rows.prfrm_tm_ann * float(weight)

                #     else:
                #         sum_tm_cal = sum_tm_cal + float(rows.prfrm_tm_ann) * float(weight)

                # # 환산 업무량 열 추가
                # df3.loc[i, 'tm_cal'] = round(sum_tm_cal, 1)

                # # 환산 PO열 추가. 환산 PO = PO * 업무량 가중치
                # # 그 팀 내의 G1, G2, G3, G4, G5에 해당하는 사람 수를 가져옴
                # for grade in work_grade_list: # 업무등급 각각에 대하여 작업.
                #     # 그 grade에 해당하는 pos_nm 리스트를 만든다.
                #     pos_nm_of_grade = BsPosGrade.objects.filter(prd_cd=prd_cd_selected, work_grade_id=grade).values_list('pos_nm', flat=True)

                #     cnt_grade = 0 # 그 grade에 해당하는 po

                #     # 위에서 만들어준 pos_nm 리스트를 갖고 Mbr에 접근
                #     for row in mbr_of_dept:

                #         if row.pos_nm in pos_nm_of_grade: # 직위 이름이 pos_nm_gr_grade(그 grade에 해당하는 pos_nm 리스트)에 있으면
                #             cnt_grade = cnt_grade + 1 # PO를 1 늘려준다.

                #     if grade == 'G1':
                #         sum_po_cal = sum_po_cal + cnt_grade * float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G1').workload_wt)
                #     elif grade == 'G2':
                #         sum_po_cal = sum_po_cal + cnt_grade * float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G2').workload_wt)
                #     elif grade == 'G3':
                #         sum_po_cal = sum_po_cal + cnt_grade * float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G3').workload_wt)
                #     elif grade == 'G4':
                #         sum_po_cal = sum_po_cal + cnt_grade * float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G4').workload_wt)
                #     elif grade == 'G5':
                #         sum_po_cal = sum_po_cal + cnt_grade * float(BsWorkGrade.objects.get(prd_cd_id=prd_cd_selected, work_grade='G5').workload_wt)

                # # 환산 PO열 추가
                # df3.loc[i, 'po_cal'] = round(sum_po_cal, 2)

                # # 적정인력 산정 열 추가
                # df3.loc[i, 'po_appr'] = round(df3.loc[i, 'tm_cal']/std_wrk_tm, 1)

                # # 과부족 열 추가
                # df3.loc[i, 'overless'] = round(df3.loc[i, 'po_cal'] - df3.loc[i, 'po_appr'], 1)

                # # 부서명 열 추가
                # df3['dept_nm'] = df3['dept_cd'].apply(lambda x: BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=x).dept_nm)

            # df3를 domain_seq, grq_seq, dept_seq 순으로 정렬한다.
            df3 = df3.sort_values(['domain_seq', 'grp_seq', 'dept_seq']).reset_index(drop=True)

            context = {
            'prd_list' : BsPrd.objects.all(),
            'title' : '업무량 분석', # 제목
            'prd_cd_selected' : prd_cd_selected,
            'tab' : tab,
            'domain_selected' : domain_selected,
            'domain_selected_key' : 'latter',
            'domain_list' : domain_list,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
            'analysis' : df3,
            }

        if tab == 'tab4': # 인력 산정 결과 탭일 때

            # BsTtlList 테이블로부터 해당 회기의 직책 리스트를 가져오고 그 가져온 리스트 수 만큼 df3의 열을 만들어준다.
            ttl_list = BsTtlList.objects.filter(prd_cd=prd_cd_selected).order_by('ttl_ordr')

            for rows in ttl_list:
                df3[rows.ttl_nm] = 0

            df3['TO합계'] = 0

            # df3의 각 행(부서)에 대하여, BsTtlCnt 테이블로부터 해당 부서의 직책별 인원수를 가져온다.
            # 이 때, 가져온 인원수를 df3의 해당 직책 열에 넣어준다.
            for i in range(len(df3)): # 각 부서에 대하여

                to_sum = 0

                for rows in ttl_list:
                    df3.loc[i, rows.ttl_nm] = BsTtlCnt.objects.get(prd_cd=prd_cd_selected, dept_cd=df3.iloc[i]['dept_cd'], ttl_nm=rows.ttl_nm).ttl_cnt
                    to_sum = to_sum + df3.loc[i, rows.ttl_nm]

                # TO합계 열 업데이트
                df3.loc[i, 'TO합계'] = to_sum

            for rows in ttl_list:
                df3[rows.ttl_nm+' PO'] = 0

            df3['PO합계'] = 0

            # df3의 각 행(부서)에 대하여, BsMbr 테이블을 이용해 각 직책별 현재 인원수를 가져온다.
            # 이 때, 가져온 인원수를 df3의 해당 직책 열에 넣어준다.
            for i in range(len(df3)): # 각 부서에 대하여

                po_sum = 0

                for rows in ttl_list:
                    df3.loc[i, rows.ttl_nm+' PO'] = BsMbr.objects.filter(prd_cd=prd_cd_selected, dept_cd=df3.iloc[i]['dept_cd'], ttl_nm=rows.ttl_nm).count()
                    po_sum = po_sum + df3.loc[i, rows.ttl_nm+' PO']

                # PO합계 열 업데이트
                df3.loc[i, 'PO합계'] = po_sum

            # 부서명 열 추가
            df3['부서명'] = df3['dept_cd'].apply(lambda x: BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=x).dept_nm)

            # 부서명 열을 dept_grp_nm 뒤로 이동
            cols = df3.columns.tolist()
            cols = cols[:2] + cols[-1:] + cols[2:-1]
            df3 = df3[cols]

            # df3를 domain_seq, grq_seq, dept_seq 순으로 정렬한다.
            df3 = df3.sort_values(['domain_seq', 'grp_seq', 'dept_seq']).reset_index(drop=True)

            # df3에서 dept_domain, domain_seq, grp_seq, dept_cd, dept_seq 열 없앤다.
            df3 = df3.drop(['dept_domain', 'domain_seq', 'grp_seq', 'dept_cd', 'dept_seq'], axis=1)

            # df3에 TO대비 열 추가해준다. TO대비 = TO합계 - PO합계
            df3['TO 대비'] = df3['TO합계'] - df3['PO합계']

            # df3의 dept_grp_nm 열 이름을 그룹명으로 변경
            df3 = df3.rename(columns={'dept_grp_nm':'그룹명'})

            # df3를 json으로 변환하여 man_result라는 데이터로 저장
            man_result = df3.to_json(orient='records')

            context = {
                'prd_list' : BsPrd.objects.all(),
                'title' : '업무량 분석', # 제목
                'prd_cd_selected' : prd_cd_selected,
                'tab' : tab,
                'domain_selected' : domain_selected,
                'domain_selected_key' : 'latter',
                'domain_list' : domain_list,
                'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
                'man_result' : man_result,
            }

    return render(request, 'jobs/JB109.html', context)


def JB110_1(request): # 부서 업무량 분석 - 탭 선택 후, 로그인한 부서의 부서 업무량 정보를 탭에 따라 띄워준다.

    context = {}

    if request.method == 'POST':
        # 선택한 회기를 input으로 받아옴.
        prd_cd_selected = request.POST['prd_cd_selected']
        # span을 탭을 어느 것을 선택하느냐에 따라 다르게 연산을 수행할 것임. 일단 그 span이 뭔지 알아낼 것임.
        span_name = request.POST.get('span_name', '')

        dept_login = get_dept_code(request.user.username) # 로그인한 부서의 부서코드
        dept_login_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_login).dept_nm # 로그인한 부서의 부서명

        # 공통으로 사용하는 context 설정
        context = {
            'title': '부서 업무량 분석',  # 제목
            'prd_list': BsPrd.objects.all(),
            'prd_cd_selected': prd_cd_selected,
            'status': 'tab_after',
            'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
            'dept_login_nm' : dept_login_nm,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if dept_login == 'DD06': # 경영기획팀이 로그인했을 경우 다른 팀들도 선택할 수 있도록 함.
            context.update({
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            })
        else:
            context.update({
                'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login),
            })

        # 선택된 탭에 따라 tab 값 설정
        if span_name == 'span1': # 직무별 업무량 분석 탭일 경우

            try:
                context['tab'] = "tab1"

                # v_jb110f 뷰에서 해당 부서의 업무량 분석 정보를 가져와서 dataframe 생성
                analysis_target = VJb110F.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login)
                data_list = [{'job_nm' : rows.job_nm, 'cnt_task' : rows.cnt_task, 'wrk_tm' : rows.wrk_tm, 'wrk_ratio1' : rows.wrk_ratio1,
                            'imprt' : rows.imprt, 'dfclt' : rows.dfclt, 'prfcn' : rows.prfcn, 'wrk_lv_sum' : rows.wrk_lv_sum,
                                'work_grade' : rows.work_grade  } for rows in analysis_target]

                df1 = pd.DataFrame(data_list)

                # job_cd열 추가. BsJob열 참조
                job_cd_list = [BsJob.objects.get(prd_cd=prd_cd_selected, job_nm=x).job_cd for x in df1['job_nm']]
                df1['job_cd'] = job_cd_list

                # df1을 job_cd 순으로 정렬
                df1 = df1.sort_values(by='job_cd')

                df1 = df1.fillna('') # NaN값을 ''로 채워줌

                sum_1 = df1['cnt_task'].sum() # 과업수 합계
                # ''을 제외하고 wrK_tm의 합계를 구한다.
                sum_2 = df1.loc[df1['wrk_tm'] != '', 'wrk_tm'].sum() # 업무량 합계

                # 연간 수행시간의 합에서 각 연간 수행시간을 나누어 구성비를 구한다. 그 구성비를 리스트로 만든다. 단, NaN값은 0으로 처리한다.
                ratio_list = [(x/sum_2)*100 if x != '' else 0 for x in df1['wrk_tm']]
                # 이 리스트의 합을 구한다.
                sum_3 = round(sum(ratio_list), 1)

                context.update({
                    'dept_selected': dept_login,
                    'dept_selected_nm' : dept_login_nm,
                    'analysis' : df1,
                    'sum_1' : sum_1,
                    'sum_2' : sum_2,
                    'sum_3' : sum_3,
                })

            except KeyError as e:

                messages.error(request, '해당 회기에 로그인한 부서의 정보가 없습니다.')

                context.update({
                    'dept_selected': dept_login,
                    'dept_selected_nm' : dept_login_nm,
                })

        elif span_name == 'span2': # 책무별 업무량 분석 탭일 경우
            context['tab'] = "tab2"

            job_task = JobTask.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login)

            data_list = [{'job_cd' : rows.job_cd_id, 'duty_nm' : rows.duty_nm, 'task_nm' : rows.task_nm,
                'work_lv_imprt' : rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn,
                'work_lv_sum' : rows.work_lv_sum, 'work_grade' : rows.work_grade_id, 'prfrm_tm_ann':rows.prfrm_tm_ann,
                    'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq } for rows in job_task]

            df1 = pd.DataFrame(data_list) # JobTask 테이블에서 필요한 rough data를 가져왔다.

            try:
                # job_cd, duty_nm, job_seq, duty_seq열을 가져와서 중복을 제거한 후, job_seq, duty_seq 순으로 정렬한다. 이후 인덱스 리셋한다.
                df2 = df1[['job_cd', 'duty_nm', 'job_seq', 'duty_seq']].drop_duplicates().sort_values(['job_seq', 'duty_seq']).reset_index(drop=True)

                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행 수를 센다. 그러면 task의 개수이다. 이를 task_cnt 열로 추가한다.
                # df2에 대해서 for문을 활용하여 task_cnt열 데이터를 추가해준다. 먼저 task_cnt열부터 만들어놓고 시작해야 한다.
                df2['task_cnt'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'task_cnt'] = len(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])])

                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구해서 그 값을 job_prfrm_tm_ann 열로 추가한다.
                # 우선 df1의 prfrm_tm_ann열 중에서 Null값을 0으로 치환한다.
                df1['prfrm_tm_ann'] = df1['prfrm_tm_ann'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_tm_ann열 데이터를 추가해준다. 먼저 duty_tm_ann열 만들어놓고 시작해야 한다.
                df2['duty_tm_ann'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'duty_tm_ann'] = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum()

                # df2의 duty_tm_ann열의 합을 구한 후, 그 값을 sum_duty_tm_ann이라는 변수에 저장한다.
                sum_duty_tm_ann = df2['duty_tm_ann'].sum()

                # df2에 각 duty의 ratio열인 duty_ratio을 추가한다. for문을 활용한다.
                df2['duty_ratio'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'duty_ratio'] = round((df2.loc[i, 'duty_tm_ann']/sum_duty_tm_ann)*100, 1)

                # 리스트를 만드는데, 앞에서는 duty_ratio를 반올림해서 만들었다. 이번에는 반올림하지 않고 만든다.
                duty_ratio_list = [(x/sum_duty_tm_ann)*100 for x in df2['duty_tm_ann']]

                # 이 리스트의 합을 구한다. 나중에 sum값으로 사용할 값이다.
                sum_duty_ratio = round(sum(duty_ratio_list), 1)

                ### 바꿔야 하는 부분 ###
                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_imprt를 해당 duty의 해당 duty_nm의 전체 prfrm_tm_ann 합에서 그 행의 비율을 곱한 값의 합을 구해서 duty_imprt열로 추가한다.
                # 우선 df1의 work_lv_imprt열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_imprt'] = df1['work_lv_imprt'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_imprt열 데이터를 추가해준다. 먼저 duty_imprt열 만들어놓고 시작해야 한다.
                df2['duty_imprt'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_imprt'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].mean(), 1)


                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_imprt를 곱한 값을 구하여 duty_imprt열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_imprt열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_imprt'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_imprt열을 구한다.
                        # df2.loc[i, 'duty_imprt'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_imprt 리스트 생성
                        work_lv_imprt_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].tolist()
                        # work_lv_imprt_list의 자료형을 float으로 변경
                        work_lv_imprt_list = [float(x) for x in work_lv_imprt_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나눈 값으로 변경
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_imprt열에 추가한다.
                        df2.loc[i, 'duty_imprt'] = round(sum([a*b for a, b in zip(work_lv_imprt_list, prfrm_tm_ann_ratio_list)]), 1)


                        # prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']/sum_prfrm_tm_ann
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        # prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_imprt열에 추가한다.
                        # df2.loc[i, 'duty_imprt'] = round(sum([a*b for a, b in zip(work_lv_imprt_list, prfrm_tm_ann_ratio_list)])*100, 1)


                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_dfclt 평균의 소수 첫째자리까지 반올림을 구해서 duty_dfclt열로 추가한다.
                # 우선 df1의 work_lv_dfclt열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_dfclt'] = df1['work_lv_dfclt'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_dfclt열 데이터를 추가해준다. 먼저 duty_dfclt열 만들어놓고 시작해야 한다.
                df2['duty_dfclt'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_dfclt'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].mean(), 1)

                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_dfclt를 곱한 값을 구하여 duty_dfclt열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_dfclt열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_dfclt'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_dfclt열을 구한다.
                        # df2.loc[i, 'duty_dfclt'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_dfclt 리스트 생성
                        work_lv_dfclt_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].tolist()
                        # work_lv_dfclt_list의 자료형을 float으로 변경
                        work_lv_dfclt_list = [float(x) for x in work_lv_dfclt_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_dfclt열에 추가한다.
                        df2.loc[i, 'duty_dfclt'] = round(sum([a*b for a, b in zip(work_lv_dfclt_list, prfrm_tm_ann_ratio_list)]), 1)



                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_prfcn 평균의 소수 첫째자리까지 반올림을 구해서 duty_prfcn열로 추가한다.
                # 우선 df1의 work_lv_prfcn열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_prfcn'] = df1['work_lv_prfcn'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_prfcn열 데이터를 추가해준다. 먼저 duty_prfcn열 만들어놓고 시작해야 한다.
                df2['duty_prfcn'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_prfcn'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].mean(), 1)

                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_prfcn을 곱한 값을 구하여 duty_prfcn열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_prfcn열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_prfcn'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_prfcn열을 구한다.
                        # df2.loc[i, 'duty_prfcn'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_prfcn 리스트 생성
                        work_lv_prfcn_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].tolist()
                        # work_lv_prfcn_list의 자료형을 float으로 변경
                        work_lv_prfcn_list = [float(x) for x in work_lv_prfcn_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나눈 값으로 변경
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_prfcn열에 추가한다.
                        df2.loc[i, 'duty_prfcn'] = round(sum([a*b for a, b in zip(work_lv_prfcn_list, prfrm_tm_ann_ratio_list)]), 1)

                # work_lv_mean열 추가. duty_imprt, duty_dfclt, duty_prfcn의 평균을 구해서 추가한다.
                df2['work_lv_mean'] = round((df2['duty_imprt'] + df2['duty_dfclt'] + df2['duty_prfcn']), 1)

                # # 바꿔야 되는 부분
                g1_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G1').work_lv_min
                g1_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G1').work_lv_max
                g2_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G2').work_lv_min
                g2_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G2').work_lv_max
                g3_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G3').work_lv_min
                g3_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G3').work_lv_max
                g4_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G4').work_lv_min
                g4_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G4').work_lv_max
                g5_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G5').work_lv_min
                g5_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G5').work_lv_max

                # df2에서 각 행에 대해 work_lv_mean열에 대해 조건문을 적용하여 G1, G2, G3, G4, G5를 부여한다.
                # for문을 이용해 각 row에 접근해서 바꿔준다.
                for i in range(len(df2)):

                    # 그 행의 job_cd가 JC001이면 work_lv_mean열에 빈칸을 부여한다.
                    if df2.loc[i, 'job_cd'] == 'JC001':
                        df2.loc[i, 'work_lv_mean'] = ''

                    # 그 행의 job_cd가 JC001이 아니면
                    else:

                        if round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) > g1_max:
                            df2.loc[i, 'work_lv_mean'] = 'G1'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g1_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g1_min:
                            df2.loc[i, 'work_lv_mean'] = 'G1'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g2_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g2_min:
                            df2.loc[i, 'work_lv_mean'] = 'G2'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g3_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g3_min:
                            df2.loc[i, 'work_lv_mean'] = 'G3'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g4_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g4_min:
                            df2.loc[i, 'work_lv_mean'] = 'G4'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g5_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g5_min:
                            df2.loc[i, 'work_lv_mean'] = 'G5'

                # BsJob 테이블을 이용해 해당 job_cd의 job_nm 열을 추가할 것이다.
                job_nm_list = [BsJob.objects.get(prd_cd=prd_cd_selected, job_cd=x).job_nm for x in df2['job_cd']]
                df2['job_nm'] = job_nm_list

                # job_nm열을 job_cd열 뒤로 이동
                cols = df2.columns.tolist()
                cols = cols[:1] + cols[-1:] + cols[1:-1]
                df2 = df2[cols]

                context.update({
                    'dept_selected': dept_login,
                    'dept_selected_nm' : dept_login_nm,
                    'analysis' : df2,
                    'sum_duty_tm_ann' : sum_duty_tm_ann,
                    'sum_duty_ratio' : sum_duty_ratio,
                })

            except KeyError as e:
                messages.error(request, '해당 회기에 선택한 부서의 조정 업무량 구성비 정보가 없습니다.')

        elif span_name == 'span3': # 담당자별 업무량 분석 탭일 경우
            context['tab'] = "tab3"

            try:
                # v_jb111 뷰에서 해당 부서의 담당자별 업무량 분석 정보를 가져와서 데이터프레임 생성
                analysis_target = VJb111.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_login)
                data_list = [{'task_prsn_chrg' : rows.task_prsn_chrg, 'job_nm' : rows.job_nm, 'wrk_tm_std' : rows.wrk_tm_std, 'total_prfrm_tm_ann' : rows.total_prfrm_tm_ann,
                            'wrk_tm_std2' : rows.wrk_tm_std2, 'work_ratio' : rows.work_ratio  } for rows in analysis_target]

                df1 = pd.DataFrame(data_list)
                sum_1 = df1['wrk_tm_std'].sum() # 표준 업무량 합계
                sum_2 = df1['total_prfrm_tm_ann'].sum() # 연간 수행시간 합계
                sum_3 = df1['wrk_tm_std2'].sum() # 표준 업무량 합계2
                sum_4 = df1['work_ratio'].sum() # 업무량 비율 합계

                context.update({
                    'dept_selected': dept_login,
                    'dept_selected_nm' : dept_login_nm,
                    'analysis' : df1,
                    'sum_1' : sum_1,
                    'sum_2' : sum_2,
                    'sum_3' : sum_3,
                    'sum_4' : sum_4,
                })

            except KeyError as e:

                messages.error(request, '해당 회기에 로그인한 부서의 정보가 없습니다.')

            context.update({
                'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
                'dept_selected' : dept_login
            })

    return render(request, 'jobs/JB110.html', context)


def JB110_2(request): # 탭이 선택된 상태에서 부서를 선택했을 때. 경영기획팀만 해당

    if request.method == 'POST':
        prd_cd_selected = request.POST['prd_cd_selected']
        dept_selected = request.POST['dept_selected']

        tab = request.POST['tab']  # 탭 정보

        # 공통 context 설정
        context = {
            'title': '부서 업무량 분석',
            'prd_list': BsPrd.objects.all(),
            'prd_cd_selected': prd_cd_selected,
            'dept_list': BsDept.objects.filter(prd_cd=prd_cd_selected),
            'dept_selected': dept_selected,
            'dept_selected_nm' : BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_selected).dept_nm,
            'tab': tab,
            # 'activate': 'yes', # 버튼 컨트롤 on
            'status': 'tab_after',
            'prd_done' : BsPrd.objects.get(prd_cd=prd_cd_selected).prd_done_yn,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        if tab == "tab1": # 직무별 업무량 분석 탭일 경우

            try:

                # v_jb110f 뷰에서 해당 부서의 업무량 분석 정보를 가져와서 dataframe 생성
                analysis_target = VJb110F.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected)
                data_list = [{'job_nm' : rows.job_nm, 'cnt_task' : rows.cnt_task, 'wrk_tm' : rows.wrk_tm, 'wrk_ratio1' : rows.wrk_ratio1,
                            'imprt' : rows.imprt, 'dfclt' : rows.dfclt, 'prfcn' : rows.prfcn, 'wrk_lv_sum' : rows.wrk_lv_sum,
                                'work_grade' : rows.work_grade  } for rows in analysis_target]
                df1 = pd.DataFrame(data_list)

                # job_cd열 추가. BsJob열 참조
                job_cd_list = [BsJob.objects.get(prd_cd=prd_cd_selected, job_nm=x).job_cd for x in df1['job_nm']]
                df1['job_cd'] = job_cd_list

                # df1을 job_cd 순으로 정렬
                df1 = df1.sort_values(by='job_cd')

                df1 = df1.fillna('') # NaN값을 ''로 채워줌

                sum_1 = df1['cnt_task'].sum() # 과업수 합계
                # ''을 제외하고 wrK_tm의 합계를 구한다.
                sum_2 = df1.loc[df1['wrk_tm'] != '', 'wrk_tm'].sum() # 업무량 합계
                # 연간 수행시간의 합에서 각 연간 수행시간을 나누어 구성비를 구한다. 그 구성비를 리스트로 만든다. 단, NaN값은 0으로 처리한다.

                ratio_list = [(x/sum_2)*100 if x != '' else 0 for x in df1['wrk_tm']]
                # 이 리스트의 합을 구한다.
                sum_3 = round(sum(ratio_list), 1)

                context.update({
                    'analysis' : df1,
                    'sum_1' : sum_1,
                    'sum_2' : sum_2,
                    'sum_3' : sum_3,
                })

            except KeyError as e:

                messages.error(request, '해당 회기에 선택한 부서의 정보가 없습니다.')


        elif tab == "tab2": # 책무별 업무량 분석 탭일 경우

            context['tab'] = "tab2"

            job_task = JobTask.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected)

            data_list = [{'job_cd' : rows.job_cd_id, 'duty_nm' : rows.duty_nm, 'task_nm' : rows.task_nm,
                'work_lv_imprt' : rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn,
                'work_lv_sum' : rows.work_lv_sum, 'work_grade' : rows.work_grade_id, 'prfrm_tm_ann':rows.prfrm_tm_ann,
                    'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq } for rows in job_task]

            df1 = pd.DataFrame(data_list) # JobTask 테이블에서 필요한 rough data를 가져왔다.

            try:
                # job_cd, duty_nm, job_seq, duty_seq열을 가져와서 중복을 제거한 후, job_seq, duty_seq 순으로 정렬한다. 이후 인덱스 리셋한다.
                df2 = df1[['job_cd', 'duty_nm', 'job_seq', 'duty_seq']].drop_duplicates().sort_values(['job_seq', 'duty_seq']).reset_index(drop=True)

                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행 수를 센다. 그러면 task의 개수이다. 이를 task_cnt 열로 추가한다.
                # df2에 대해서 for문을 활용하여 task_cnt열 데이터를 추가해준다. 먼저 task_cnt열부터 만들어놓고 시작해야 한다.
                df2['task_cnt'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'task_cnt'] = len(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])])

                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구해서 그 값을 job_prfrm_tm_ann 열로 추가한다.
                # 우선 df1의 prfrm_tm_ann열 중에서 Null값을 0으로 치환한다.
                df1['prfrm_tm_ann'] = df1['prfrm_tm_ann'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_tm_ann열 데이터를 추가해준다. 먼저 duty_tm_ann열 만들어놓고 시작해야 한다.
                df2['duty_tm_ann'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'duty_tm_ann'] = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum()

                # df2의 duty_tm_ann열의 합을 구한 후, 그 값을 sum_duty_tm_ann이라는 변수에 저장한다.
                sum_duty_tm_ann = df2['duty_tm_ann'].sum()

                # df2에 각 duty의 ratio열인 duty_ratio을 추가한다. for문을 활용한다.
                df2['duty_ratio'] = 0
                for i in range(len(df2)):
                    df2.loc[i, 'duty_ratio'] = round((df2.loc[i, 'duty_tm_ann']/sum_duty_tm_ann)*100, 1)

                # 리스트를 만드는데, 앞에서는 duty_ratio를 반올림해서 만들었다. 이번에는 반올림하지 않고 만든다.
                duty_ratio_list = [(x/sum_duty_tm_ann)*100 for x in df2['duty_tm_ann']]

                # 이 리스트의 합을 구한다. 나중에 sum값으로 사용할 값이다.
                sum_duty_ratio = round(sum(duty_ratio_list), 1)

                ### 바꿔야 하는 부분 ###
                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_imprt를 해당 duty의 해당 duty_nm의 전체 prfrm_tm_ann 합에서 그 행의 비율을 곱한 값의 합을 구해서 duty_imprt열로 추가한다.
                # 우선 df1의 work_lv_imprt열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_imprt'] = df1['work_lv_imprt'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_imprt열 데이터를 추가해준다. 먼저 duty_imprt열 만들어놓고 시작해야 한다.
                df2['duty_imprt'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_imprt'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].mean(), 1)


                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_imprt를 곱한 값을 구하여 duty_imprt열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_imprt열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_imprt'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_imprt열을 구한다.
                        # df2.loc[i, 'duty_imprt'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_imprt 리스트 생성
                        work_lv_imprt_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_imprt'].tolist()
                        # work_lv_imprt_list의 자료형을 float으로 변경
                        work_lv_imprt_list = [float(x) for x in work_lv_imprt_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나눈 값으로 변경
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_imprt열에 추가한다.
                        df2.loc[i, 'duty_imprt'] = round(sum([a*b for a, b in zip(work_lv_imprt_list, prfrm_tm_ann_ratio_list)]), 1)


                        # prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']/sum_prfrm_tm_ann
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        # prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_imprt열에 추가한다.
                        # df2.loc[i, 'duty_imprt'] = round(sum([a*b for a, b in zip(work_lv_imprt_list, prfrm_tm_ann_ratio_list)])*100, 1)


                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_dfclt 평균의 소수 첫째자리까지 반올림을 구해서 duty_dfclt열로 추가한다.
                # 우선 df1의 work_lv_dfclt열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_dfclt'] = df1['work_lv_dfclt'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_dfclt열 데이터를 추가해준다. 먼저 duty_dfclt열 만들어놓고 시작해야 한다.
                df2['duty_dfclt'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_dfclt'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].mean(), 1)

                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_dfclt를 곱한 값을 구하여 duty_dfclt열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_dfclt열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_dfclt'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_dfclt열을 구한다.
                        # df2.loc[i, 'duty_dfclt'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_dfclt 리스트 생성
                        work_lv_dfclt_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_dfclt'].tolist()
                        # work_lv_dfclt_list의 자료형을 float으로 변경
                        work_lv_dfclt_list = [float(x) for x in work_lv_dfclt_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_dfclt열에 추가한다.
                        df2.loc[i, 'duty_dfclt'] = round(sum([a*b for a, b in zip(work_lv_dfclt_list, prfrm_tm_ann_ratio_list)]), 1)



                # df1을 참고하여, 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_prfcn 평균의 소수 첫째자리까지 반올림을 구해서 duty_prfcn열로 추가한다.
                # 우선 df1의 work_lv_prfcn열 중에서 Null값을 0으로 치환한다.
                df1['work_lv_prfcn'] = df1['work_lv_prfcn'].fillna(0)
                # 그 후 df2에 대해서 for문을 활용하여 duty_prfcn열 데이터를 추가해준다. 먼저 duty_prfcn열 만들어놓고 시작해야 한다.
                df2['duty_prfcn'] = 0
                # for i in range(len(df2)):
                #     df2.loc[i, 'duty_prfcn'] = round(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].mean(), 1)

                for i in range(len(df2)): # df2의 행의 개수만큼 반복한다.

                    # 그 행이 해당 job_cd의 해당 duty_nm을 가지고 있는 행들 중에서 prfrm_tm_ann의 비율이 얼마나 되는지 구하고 그 값에 work_lv_prfcn을 곱한 값을 구하여 duty_prfcn열에 추가한다.
                    # 이를 위해서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann의 합을 구한다.
                    sum_prfrm_tm_ann = float(df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann'].sum())

                    # sum_prfrm_tm_ann이 0인 경우는 duty_prfcn열을 0으로 처리한다.
                    if sum_prfrm_tm_ann == 0:
                        df2.loc[i, 'duty_prfcn'] = 0
                    else:
                        # sum_prfrm_tm_ann이 0이 아닌 경우는 duty_prfcn열을 구한다.
                        # df2.loc[i, 'duty_prfcn'] = round(float((df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].sum())/sum_prfrm_tm_ann)*100, 1)
                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 work_lv_prfcn 리스트 생성
                        work_lv_prfcn_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['work_lv_prfcn'].tolist()
                        # work_lv_prfcn_list의 자료형을 float으로 변경
                        work_lv_prfcn_list = [float(x) for x in work_lv_prfcn_list]

                        # df1에서 해당 job_cd의 해당 duty_nm을 가지고 있는 행들의 prfrm_tm_ann을 float으로 변경한 것을 sum_prfrm_tm_ann으로 나눈 값의 리스트 생성
                        prfrm_tm_ann_ratio_list = df1[(df1['job_cd'] == df2.loc[i, 'job_cd']) & (df1['duty_nm'] == df2.loc[i, 'duty_nm'])]['prfrm_tm_ann']
                        # prfrm_tm_ann_ratio_list의 자료형을 float으로 변경
                        prfrm_tm_ann_ratio_list = [float(x) for x in prfrm_tm_ann_ratio_list]
                        # prfrm_tm_ann_ratio_list의 모든 값을 sum_prfrm_tm_ann으로 나눈 값으로 변경
                        prfrm_tm_ann_ratio_list = [x/sum_prfrm_tm_ann for x in prfrm_tm_ann_ratio_list]

                        # 두 리스트를 곱한 값을 구하여 duty_prfcn열에 추가한다.
                        df2.loc[i, 'duty_prfcn'] = round(sum([a*b for a, b in zip(work_lv_prfcn_list, prfrm_tm_ann_ratio_list)]), 1)

                # work_lv_mean열 추가. duty_imprt, duty_dfclt, duty_prfcn의 평균을 구해서 추가한다.
                df2['work_lv_mean'] = round((df2['duty_imprt'] + df2['duty_dfclt'] + df2['duty_prfcn']), 1)

                # # 바꿔야 되는 부분
                g1_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G1').work_lv_min
                g1_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G1').work_lv_max
                g2_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G2').work_lv_min
                g2_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G2').work_lv_max
                g3_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G3').work_lv_min
                g3_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G3').work_lv_max
                g4_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G4').work_lv_min
                g4_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G4').work_lv_max
                g5_min = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G5').work_lv_min
                g5_max = BsWorkGrade.objects.get(prd_cd_id = prd_cd_selected, work_grade='G5').work_lv_max

                # df2에서 각 행에 대해 work_lv_mean열에 대해 조건문을 적용하여 G1, G2, G3, G4, G5를 부여한다.
                # for문을 이용해 각 row에 접근해서 바꿔준다.
                for i in range(len(df2)):

                    # 그 행의 job_cd가 JC001이면 work_lv_mean열에 빈칸을 부여한다.
                    if df2.loc[i, 'job_cd'] == 'JC001':
                        df2.loc[i, 'work_lv_mean'] = ''

                    # 그 행의 job_cd가 JC001이 아니면
                    else:

                        if round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) > g1_max:
                            df2.loc[i, 'work_lv_mean'] = 'G1'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g1_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g1_min:
                            df2.loc[i, 'work_lv_mean'] = 'G1'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g2_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g2_min:
                            df2.loc[i, 'work_lv_mean'] = 'G2'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g3_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g3_min:
                            df2.loc[i, 'work_lv_mean'] = 'G3'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g4_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g4_min:
                            df2.loc[i, 'work_lv_mean'] = 'G4'
                        elif round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) < g5_max+1 and round((df2.loc[i, 'duty_imprt'] + df2.loc[i, 'duty_dfclt'] + df2.loc[i, 'duty_prfcn']), 1) >= g5_min:
                            df2.loc[i, 'work_lv_mean'] = 'G5'

                # BsJob 테이블을 이용해 해당 job_cd의 job_nm 열을 추가할 것이다.
                job_nm_list = [BsJob.objects.get(prd_cd=prd_cd_selected, job_cd=x).job_nm for x in df2['job_cd']]
                df2['job_nm'] = job_nm_list

                # job_nm열을 job_cd열 뒤로 이동
                cols = df2.columns.tolist()
                cols = cols[:1] + cols[-1:] + cols[1:-1]
                df2 = df2[cols]

                context.update({
                    'dept_selected': dept_selected,
                    'dept_selected_nm' : BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_selected).dept_nm,
                    'analysis' : df2,
                    'sum_duty_tm_ann' : sum_duty_tm_ann,
                    'sum_duty_ratio' : sum_duty_ratio,
                })

            except KeyError as e:
                messages.error(request, '해당 회기에 선택한 부서의 조정 업무량 구성비 정보가 없습니다.')

        elif tab == "tab3": # 담당자별 업무량 분석 탭일 경우

            try:

                # v_jb111 뷰에서 해당 부서의 담당자별 업무량 분석 정보를 가져와서 데이터프레임 생성
                analysis_target = VJb111.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_selected)
                data_list = [{'task_prsn_chrg' : rows.task_prsn_chrg, 'job_nm' : rows.job_nm, 'wrk_tm_std' : rows.wrk_tm_std, 'total_prfrm_tm_ann' : rows.total_prfrm_tm_ann,
                            'wrk_tm_std2' : rows.wrk_tm_std2, 'work_ratio' : rows.work_ratio  } for rows in analysis_target]

                df1 = pd.DataFrame(data_list)

                sum_1 = df1['wrk_tm_std'].sum() # 표준 업무량 합계
                sum_2 = df1['total_prfrm_tm_ann'].sum() # 연간 수행시간 합계
                sum_3 = df1['wrk_tm_std2'].sum() # 표준 업무량 합계2
                sum_4 = df1['work_ratio'].sum() # 업무량 비율 합계

                context.update({
                    'analysis' : df1,
                    'sum_1' : sum_1,
                    'sum_2' : sum_2,
                    'sum_3' : sum_3,
                    'sum_4' : sum_4,
                })

            except KeyError as e:

                messages.error(request, '해당 회기에 로그인한 부서의 정보가 없습니다.')

    return render(request, 'jobs/JB110.html', context)


def JB200_1(request): # 분석 기초 자료 화면 - 부서 변경 시

    if request.method == 'POST':
        prd_selected = request.POST['prd_selected']
        dept_selected = request.POST['dept_cd_selected']
        dept_list = BsDept.objects.filter(prd_cd=prd_selected)
        # std_wrk_tm = BsStdWrkTm.objects.get(prd_cd=prd_selected).std_wrk_tm
        std_wrk_tm = float(BsStdWrkTm.objects.get(prd_cd=prd_selected).std_wrk_tm)

        # job_task 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
        original_rows=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                    'work_lv_imprt': rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                        'work_grade': rows.work_grade_id, 'prfrm_tm_ann': rows.prfrm_tm_ann,
                          'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq } for rows in original_rows]
        df1 = pd.DataFrame(data_list)

        # df1에 prfrm_tm_ann_cal 열을 추가해준다. 초기값은 null이다.
        df1['prfrm_tm_ann_cal'] = None

        # df1에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
        # work_grade가 null일 경우 prfrm_tm_ann_cal열도 null값으로 한다.
        # df1의 각 행에 대하여 수행해준다. for문을 활용한다.
        for i in range(len(df1)):
            if df1['work_grade'][i] == 'G1':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G1').workload_wt)
            elif df1['work_grade'][i] == 'G2':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G2').workload_wt)
            elif df1['work_grade'][i] == 'G3':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G3').workload_wt)
            elif df1['work_grade'][i] == 'G4':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G4').workload_wt)
            elif df1['work_grade'][i] == 'G5':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G5').workload_wt)
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None

        # df1['prfrm_tm_ann_cal']의 자료형을 float으로 변경
        df1['prfrm_tm_ann_cal'] = df1['prfrm_tm_ann_cal'].astype(float)

        # job_task_adj 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
        original_rows_2=JobTaskAdj.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                        'work_lv_imprt_adj': rows.work_lv_imprt, 'work_lv_dfclt_adj': rows.work_lv_dfclt, 'work_lv_prfcn_adj': rows.work_lv_prfcn, 'work_lv_sum_adj': rows.work_lv_sum,
                        'work_grade_adj': rows.work_grade_id, 'prfrm_tm_ann_adj': rows.prfrm_tm_ann, 'adj_yn': rows.adj_yn } for rows in original_rows_2]
        df2 = pd.DataFrame(data_list_2)

        # df2에 prfrm_tm_ann_cal_adj열을 추가해준다. 초기값은 null이다.
        df2['prfrm_tm_ann_cal_adj'] = None

        # df2에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal_adj)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade_adj가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
        # work_grade_adj가 null일 경우 prfrm_tm_ann_cal_adj열도 null값으로 한다.
        # df2의 각 행에 대하여 수행해준다. for문을 활용한다.
        for i in range(len(df2)):
            if df2['work_grade_adj'][i] == 'G1':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G1').workload_wt)
            elif df2['work_grade_adj'][i] == 'G2':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G2').workload_wt)
            elif df2['work_grade_adj'][i] == 'G3':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G3').workload_wt)
            elif df2['work_grade_adj'][i] == 'G4':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G4').workload_wt)
            elif df2['work_grade_adj'][i] == 'G5':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G5').workload_wt)
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None

        # df2['prfrm_tm_ann_cal_adj']의 자료형을 float으로 변경
        df2['prfrm_tm_ann_cal_adj'] = df2['prfrm_tm_ann_cal_adj'].astype(float)

        try:
            df3 = pd.merge(df1, df2).sort_values(['job_seq', 'duty_seq', 'task_seq']) # job_task와 job_activity merge, 순서는 job_seq, duty_seq, task_seq, act_seq 순

            # job_nm 찾기
            original_rows_3 = BsJob.objects.filter(prd_cd=prd_selected)
            data_list_3 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_3]
            df4 = pd.DataFrame(data_list_3)

            df3 = pd.merge(df3, df4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

            # df3.to_excel('df3.xlsx')
            df_json = df3.to_json(orient='records')

             # BsWorkGrade에서 데이터를 가져와서, json으로 만들어서 넘겨준다.
            original_rows_4 = BsWorkGrade.objects.filter(prd_cd=prd_selected)
            data_list_4 = [{'work_grade': rows.work_grade, 'work_lv_min': rows.work_lv_min, 'work_lv_max': rows.work_lv_max, 'workload_wt': rows.workload_wt} for rows in original_rows_4]
            df5 = pd.DataFrame(data_list_4)
            df_json_2 = df5.to_json(orient='records')

            # 환산 업무량 합 구하기 #####################
            work_grade_list = ['G1', 'G2', 'G3', 'G4', 'G5'] # 업무등급 리스트. 이 리스트는 변하지 않는다.
            mbr_of_dept = BsMbr.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 해당 부서의 부서원 목록
            mbr_of_dept = mbr_of_dept.exclude(ttl_nm='팀장')

            analysis = pd.DataFrame({'work_grade':work_grade_list}) # 분석 결과를 담을 데이터프레임
            prfrm_tm_ann = [] # 각 업무등급별 총 업무량을 담을 리스트
            prfrm_tm_ann_adj = [] # 각 업무등급별 총 조정업무량을 담을 리스트

            for grade in work_grade_list: # 각 업무 등급에 대해서

                df6 = df3[df3['work_grade'] == grade] # JobTask 혹은 JobTaskAdj을 이용해 만든 df1에서 그 grade에 해당하는 task만 추출하여 df2 만들기
                df7 = df3[df3['work_grade_adj'] == grade]

                prfrm_tm_ann.append(float(df6['prfrm_tm_ann'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann의 합을 구한다.
                prfrm_tm_ann_adj.append(float(df7['prfrm_tm_ann_adj'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann_adj의 합을 구한다.

            analysis['prfrm_tm_ann'] = prfrm_tm_ann # 분석 결과 데이터프레임에 prfrm_tm_ann 열 추가
            analysis['prfrm_tm_ann_adj'] = prfrm_tm_ann_adj # 분석 결과 데이터프레임에 prfrm_tm_ann_adj 열 추가

            workload = BsWorkGrade.objects.filter(prd_cd=prd_selected) # 업무량 가중치 object
            workload_wt = [float(n.workload_wt) for n in workload ] # 업무량 가중치 list(float)

            analysis['workload_wt'] = workload_wt # 업무량 가중치

            analysis['prfrm_tm_ann_cal'] = analysis['prfrm_tm_ann'] * analysis['workload_wt']
            analysis['prfrm_tm_ann_cal_adj'] = analysis['prfrm_tm_ann_adj'] * analysis['workload_wt']

            sum_prfrm_tm_ann_cal = round(analysis['prfrm_tm_ann_cal'].sum(), 1)
            sum_prfrm_tm_ann_cal_adj = round(analysis['prfrm_tm_ann_cal_adj'].sum(), 1)

            analysis['po_right'] = round(analysis['prfrm_tm_ann_cal']/std_wrk_tm, 1) # 적정 인력 산정
            analysis['po_right_adj'] = round(analysis['prfrm_tm_ann_cal_adj']/std_wrk_tm, 1) # 조정 적정 인력 산정

            po_right = round(analysis['po_right'].sum(), 1)
            po_right_adj = round(analysis['po_right_adj'].sum(), 1)

            context = {
                'prd_list' : BsPrd.objects.all(),
                'title' : '분석 기초자료', # 제목
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'dept_list' : dept_list,
                'dept_cd_selected' : dept_selected,
                'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
                'data' : df_json,
                'standard' : df_json_2,
                'std_wrk_tm' : std_wrk_tm,
                'sum_prfrm_tm_ann_cal' : sum_prfrm_tm_ann_cal,
                'po_right' : po_right,
                'sum_prfrm_tm_ann_cal_adj' : sum_prfrm_tm_ann_cal_adj,
                'po_right_adj' : po_right_adj,
            }
        except pd.errors.MergeError as e:

            messages.error(request, '해당 회기에 분석 정보 기초자료 데이터가 없습니다.')

            context = {
                'prd_list' : BsPrd.objects.all(),
                'title' : '분석 기초자료', # 제목
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'dept_list' : dept_list,
                'dept_cd_selected' : dept_selected,
                'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
                'data' : 'null',
            }


    return render(request, 'jobs/JB200.html', context)


def JB200_2(request): # 업무량 분석 기초 자료 화면- 저장/취소 버튼 누를 시

    if request.method == 'POST':

        prd_selected = request.POST['prd_selected']
        dept_selected = request.POST['dept_cd_selected']
        dept_list = BsDept.objects.filter(prd_cd=prd_selected)

        action = request.POST['action'] # 저장/취소 버튼 중 어떤 것을 눌렀는지

        context = {
                'prd_list' : BsPrd.objects.all(),
                'title' : '분석 기초자료', # 제목
                'prd_selected' : prd_selected,
                'prd_done' : BsPrd.objects.get(prd_cd=prd_selected).prd_done_yn,
                'dept_list' : dept_list,
                'dept_cd_selected' : dept_selected,
                'dept_mgr_yn': get_dept_mgr_yn(request.user.username),
            }

        if action == 'action1':

            json_data = request.POST.get('jsonData')
            data = json.loads(json_data) # JSON 문자열을 Python 객체로 변환
            df_adj = pd.DataFrame(data) # Pandas DataFrame으로 변환. 기존 adj 테이블과 비교할 것임

            # df_adj에서 필요한 부분만 추출하여 job_task_adj와 비교할 것이다.
            df_adj = df_adj[['prd_cd', 'dept_cd', 'job_cd', 'duty_nm', 'task_nm',
                              'work_lv_imprt_adj', 'work_lv_dfclt_adj', 'work_lv_prfcn_adj', 'work_lv_sum_adj', 'work_grade_adj',
                                'prfrm_tm_ann_adj', 'adj_yn', 'job_seq', 'duty_seq', 'task_seq']]

            # df_adj를 job_seq, duty_seq, task_seq 순으로 정렬
            df_adj = df_adj.sort_values(['job_seq', 'duty_seq', 'task_seq']).reset_index(drop=True)

            # DB의 job_task_adj 테이블로부터 해당 회기의 해당 부서의 정보를 가져와서 dataframe을 생성한다.
            original_rows=JobTaskAdj.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
            data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                        'work_lv_imprt': rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                        'work_grade': rows.work_grade_id, 'prfrm_tm_ann': rows.prfrm_tm_ann, 'adj_yn': rows.adj_yn,
                          'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq
                          } for rows in original_rows]
            df1 = pd.DataFrame(data_list)

            # df1을 job_seq, duty_seq, task_seq 순으로 정렬
            df1 = df1.sort_values(['job_seq', 'duty_seq', 'task_seq']).reset_index(drop=True)

            # df1을 job_seq, duty_seq, task_seq 순으로 정렬

            # df_adj의 컬럼 이름 변경(_adj붙은 것들)
            df_adj.columns = ['prd_cd', 'dept_cd', 'job_cd', 'duty_nm', 'task_nm',
                               'work_lv_imprt', 'work_lv_dfclt', 'work_lv_prfcn', 'work_lv_sum', 'work_grade', 'prfrm_tm_ann', 'adj_yn', 'job_seq', 'duty_seq', 'task_seq']

            # df1과 df_adj를 비교하여 다른 행이 있으면 df_adj의 값들을 DB에 update
            for i in range(len(df_adj)):

                # 공통코드는 빼고 업데이트
                if df_adj.loc[i, 'job_cd'] != 'JC001': # 공통코드가 아닐 경우에만 업데이트
                    if df1.loc[i, 'work_lv_imprt'] != df_adj.loc[i, 'work_lv_imprt'] or df1.loc[i, 'work_lv_dfclt'] != df_adj.loc[i, 'work_lv_dfclt'] or df1.loc[i, 'work_lv_prfcn'] != df_adj.loc[i, 'work_lv_prfcn'] or df1.loc[i, 'prfrm_tm_ann'] != df_adj.loc[i, 'prfrm_tm_ann']:

                        JobTaskAdj.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected, job_cd=df_adj.loc[i, 'job_cd'], duty_nm=df_adj.loc[i, 'duty_nm'], task_nm=df_adj.loc[i, 'task_nm']).update(
                            work_lv_imprt=df_adj.loc[i, 'work_lv_imprt'], work_lv_dfclt=df_adj.loc[i, 'work_lv_dfclt'], work_lv_prfcn=df_adj.loc[i, 'work_lv_prfcn'],
                            work_lv_sum=df_adj.loc[i, 'work_lv_sum'], work_grade_id=df_adj.loc[i, 'work_grade'], prfrm_tm_ann=df_adj.loc[i, 'prfrm_tm_ann'], adj_yn=df_adj.loc[i, 'adj_yn'])


        # 다시 화면을 띄워준다. 취소일 때는 여기에 해당.
        # std_wrk_tm = BsStdWrkTm.objects.get(prd_cd=prd_selected).std_wrk_tm
        std_wrk_tm = float(BsStdWrkTm.objects.get(prd_cd=prd_selected).std_wrk_tm)

        # job_task 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
        original_rows=JobTask.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                    'work_lv_imprt': rows.work_lv_imprt, 'work_lv_dfclt': rows.work_lv_dfclt, 'work_lv_prfcn': rows.work_lv_prfcn, 'work_lv_sum': rows.work_lv_sum,
                        'work_grade': rows.work_grade_id, 'prfrm_tm_ann': rows.prfrm_tm_ann,
                          'job_seq': rows.job_seq, 'duty_seq': rows.duty_seq, 'task_seq': rows.task_seq } for rows in original_rows]
        df1 = pd.DataFrame(data_list)

        # df1에 prfrm_tm_ann_cal 열을 추가해준다. 환산 업무량이지. 초기값은 null이다.
        df1['prfrm_tm_ann_cal'] = None

        # df1에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
        # work_grade가 null일 경우 prfrm_tm_ann_cal열도 null값으로 한다.
        # df1의 각 행에 대하여 수행해준다. for문을 활용한다.
        for i in range(len(df1)):
            if df1['work_grade'][i] == 'G1':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G1').workload_wt)
            elif df1['work_grade'][i] == 'G2':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G2').workload_wt)
            elif df1['work_grade'][i] == 'G3':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G3').workload_wt)
            elif df1['work_grade'][i] == 'G4':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G4').workload_wt)
            elif df1['work_grade'][i] == 'G5':
                if df1['prfrm_tm_ann'][i] == None:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = None
                else:
                    df1.loc[i, 'prfrm_tm_ann_cal'] = float(df1.loc[i, 'prfrm_tm_ann']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G5').workload_wt)
            else:
                df1.loc[i, 'prfrm_tm_ann_cal'] = None

        # df1['prfrm_tm_ann_cal']의 자료형을 float으로 변경
        df1['prfrm_tm_ann_cal'] = df1['prfrm_tm_ann_cal'].astype(float)

        # job_task_adj 테이블에 접근하여 dataframe을 만들고, json을 만들어서 넘겨준다.
        original_rows_2=JobTaskAdj.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 나중에 prd_cd 바꿔줘야 함
        data_list_2 = [{'prd_cd' : rows.prd_cd_id, 'dept_cd' : rows.dept_cd_id, 'job_cd': rows.job_cd_id, 'duty_nm': rows.duty_nm, 'task_nm': rows.task_nm,
                        'work_lv_imprt_adj': rows.work_lv_imprt, 'work_lv_dfclt_adj': rows.work_lv_dfclt, 'work_lv_prfcn_adj': rows.work_lv_prfcn, 'work_lv_sum_adj': rows.work_lv_sum,
                        'work_grade_adj': rows.work_grade_id, 'prfrm_tm_ann_adj': rows.prfrm_tm_ann, 'adj_yn': rows.adj_yn} for rows in original_rows_2]
        df2 = pd.DataFrame(data_list_2)

        # df2에 prfrm_tm_ann_cal_adj열을 추가해준다. 환산 업무량의 adj인 것이다. 초기값은 null이다.
        df2['prfrm_tm_ann_cal_adj'] = None

        # df2에 연간 업무량에 가중치를 곱한 값을 열(prfrm_tm_ann_cal_adj)로 추가한다. 가중치는 각 행에 따라 다르며, work_grade_adj가 G1일 경우 1.25, ㅎ2이면 1.125, G3이면 1.0, G4이면 0.875, G5면 0.75이다.
        # work_grade_adj가 null일 경우 prfrm_tm_ann_cal_adj열도 null값으로 한다.
        # df2의 각 행에 대하여 수행해준다. for문을 활용한다.
        for i in range(len(df2)):
            if df2['work_grade_adj'][i] == 'G1':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G1').workload_wt)
            elif df2['work_grade_adj'][i] == 'G2':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G2').workload_wt)
            elif df2['work_grade_adj'][i] == 'G3':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G3').workload_wt)
            elif df2['work_grade_adj'][i] == 'G4':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G4').workload_wt)
            elif df2['work_grade_adj'][i] == 'G5':
                if df2['prfrm_tm_ann_adj'][i] == None:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None
                else:
                    df2.loc[i, 'prfrm_tm_ann_cal_adj'] = float(df2.loc[i, 'prfrm_tm_ann_adj']) * float(BsWorkGrade.objects.get(prd_cd_id=prd_selected, work_grade='G5').workload_wt)
            else:
                df2.loc[i, 'prfrm_tm_ann_cal_adj'] = None

        # df2['prfrm_tm_ann_cal_adj']의 자료형을 float으로 변경
        df2['prfrm_tm_ann_cal_adj'] = df2['prfrm_tm_ann_cal_adj'].astype(float)

        df3 = pd.merge(df1, df2).sort_values(['job_seq', 'duty_seq', 'task_seq']) # job_task와 job_activity merge, 순서는 job_seq, duty_seq, task_seq, act_seq 순

        # job_nm 찾기
        original_rows_3 = BsJob.objects.filter(prd_cd=prd_selected)
        data_list_3 = [{'prd_cd' : rows.prd_cd_id, 'job_cd': rows.job_cd, 'job_nm': rows.job_nm} for rows in original_rows_3]
        df4 = pd.DataFrame(data_list_3)

        df3 = pd.merge(df3, df4) # job_nm 추가. job_cd로 merge, 없는 부분은 뺀다. job_nm을 job_cd 뒤로 보낸다.

        # df3.to_excel('df3.xlsx')
        df_json = df3.to_json(orient='records')

        # BsWorkGrade에서 데이터를 가져와서, json으로 만들어서 넘겨준다.
        original_rows_4 = BsWorkGrade.objects.filter(prd_cd=prd_selected)
        data_list_4 = [{'work_grade': rows.work_grade, 'work_lv_min': rows.work_lv_min, 'work_lv_max': rows.work_lv_max, 'workload_wt': rows.workload_wt} for rows in original_rows_4]
        df5 = pd.DataFrame(data_list_4)
        df_json_2 = df5.to_json(orient='records')

        # 환산 업무량 합 구하기 #####################
        work_grade_list = ['G1', 'G2', 'G3', 'G4', 'G5'] # 업무등급 리스트. 이 리스트는 변하지 않는다.
        mbr_of_dept = BsMbr.objects.filter(prd_cd=prd_selected, dept_cd=dept_selected) # 해당 부서의 부서원 목록
        mbr_of_dept = mbr_of_dept.exclude(ttl_nm='팀장')

        analysis = pd.DataFrame({'work_grade':work_grade_list}) # 분석 결과를 담을 데이터프레임
        prfrm_tm_ann = [] # 각 업무등급별 총 업무량을 담을 리스트
        prfrm_tm_ann_adj = [] # 각 업무등급별 총 조정업무량을 담을 리스트

        for grade in work_grade_list: # 각 업무 등급에 대해서

            df6 = df3[df3['work_grade'] == grade] # JobTask 이용해 만든 df1에서 그 grade에 해당하는 task만 추출하여 df2 만들기
            df7 = df3[df3['work_grade_adj'] == grade] # JobTaskAdj에서 그 grade에 해당하는 task만 추출하여 df2 만들기

            prfrm_tm_ann.append(float(df6['prfrm_tm_ann'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann의 합을 구한다.
            prfrm_tm_ann_adj.append(float(df7['prfrm_tm_ann_adj'].sum())) # 그 grade에 해당하는 task의 prfrm_tm_ann_adj의 합을 구한다.

        analysis['prfrm_tm_ann'] = prfrm_tm_ann # 분석 결과 데이터프레임에 prfrm_tm_ann 열 추가
        analysis['prfrm_tm_ann_adj'] = prfrm_tm_ann_adj # 분석 결과 데이터프레임에 prfrm_tm_ann_adj 열 추가

        workload = BsWorkGrade.objects.filter(prd_cd=prd_selected) # 업무량 가중치 object
        workload_wt = [float(n.workload_wt) for n in workload ] # 업무량 가중치 list(float)

        analysis['workload_wt'] = workload_wt # 업무량 가중치

        analysis['prfrm_tm_ann_cal'] = analysis['prfrm_tm_ann'] * analysis['workload_wt']
        analysis['prfrm_tm_ann_cal_adj'] = analysis['prfrm_tm_ann_adj'] * analysis['workload_wt']

        sum_prfrm_tm_ann_cal = round(analysis['prfrm_tm_ann_cal'].sum(), 1)
        sum_prfrm_tm_ann_cal_adj = round(analysis['prfrm_tm_ann_cal_adj'].sum(), 1)

        analysis['po_right'] = round(analysis['prfrm_tm_ann_cal']/std_wrk_tm, 1) # 적정 인력 산정
        analysis['po_right_adj'] = round(analysis['prfrm_tm_ann_cal_adj']/std_wrk_tm, 1) # 조정 적정 인력 산정

        po_right = round(analysis['po_right'].sum(), 1)
        po_right_adj = round(analysis['po_right_adj'].sum(), 1)

        context.update({
                'data' : df_json,
                'standard': df_json_2,
                'std_wrk_tm': std_wrk_tm,
                'sum_prfrm_tm_ann_cal': sum_prfrm_tm_ann_cal,
                'po_right': po_right,
                'sum_prfrm_tm_ann_cal_adj': sum_prfrm_tm_ann_cal_adj,
                'po_right_adj': po_right_adj,
            })

    return render(request, 'jobs/JB200.html', context)


def JB300_1(request): # 직무 분류 체계에서 버튼 클릭 시

    if request.method == 'POST':

        dept_domain_selected = request.POST['dept_domain_selected']
        prd_cd_selected = request.POST['prd_cd_selected']

        context = {
            'title' : '직무 분류 체계', # 제목
            'prd' : BsPrd.objects.all(),
            'prd_cd_selected' : prd_cd_selected,
            'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
        }

        action = request.POST['action'] # 저장/취소 버튼 중 어떤 것을 눌렀는지

        if action == 'action1': # 직무 분류 체계 눌렀을 때

            # 조직 그룹
            # org_grp는 BsDeptGrpDomain 테이블에서 그 회기의 그 dept_domain에 해당하는 dept_grp_nm의 리스트이다.
            org_grp = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected,
                                                      dept_domain=dept_domain_selected).order_by('domain_seq', 'grp_seq').values_list('dept_grp_nm', flat=True).distinct()
            org_grp = list(org_grp) # 리스트로 변환

            # 그룹별 부서 목록
            # 부서 그룹 org_grp_dept는, BsDeptGrp 테이블에서 각 dept_grp_nm에 대해 속해 있는 dept_cd의 리스트들을 나열한 dictionary 형태의 객체이다.
            org_grp_dept = {}
            for grp in org_grp:
                dept_list = BsDeptGrp.objects.filter(prd_cd=prd_cd_selected, dept_grp_nm=grp).order_by('dept_seq').values_list('dept_cd', flat=True)
                dept_list = list(dept_list)
                org_grp_dept[grp] = dept_list

            # BsDept 테이블을 이용하여 dept_cd를 dept_nm으로 치환한다.
            for grp in org_grp:
                dept_list = []
                for dept_cd in org_grp_dept[grp]:
                    dept_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_cd).dept_nm
                    dept_list.append(dept_nm)
                org_grp_dept[grp] = dept_list

            # 부서 현인원
            # 부서 현인원 dept_emp_to는, BsMbr 테이블에서 각 dept_cd에 대해 그 dept_cd에 속한 부서원의 수를 나열한 dictionary 형태의 객체이다. 위의 org_grp와 상관없다.
            # 해당 회기의 dept_cd 목록을 만든다. 우선은 BsDeptGrpDomain으로부터 dept_grp_nm 리스트를 만든다.
            dept_list = BsDeptGrpDomain.objects.filter(prd_cd=prd_cd_selected,
                                                        dept_domain=dept_domain_selected).order_by('domain_seq', 'grp_seq').values_list('dept_grp_nm', flat=True).distinct()
            # 위의 리스트를 이용해 dept_cd 리스트를 만드는데, 이는 BsDeptGrp 테이블에서 dept_grp_nm에 해당하는 dept_cd를 가져와서 리스트로 만든다.
            dept_cd_list = []
            for dept in dept_list:
                dept_cd = BsDeptGrp.objects.filter(prd_cd=prd_cd_selected, dept_grp_nm=dept).order_by('dept_seq').values_list('dept_cd', flat=True)
                dept_cd_list.extend(dept_cd)

            # dept_cd_list를 이용하여 각 부서의 부서원 수를 구한다.
            dept_emp_to = {}
            for dept_cd in dept_cd_list:
                dept_emp_to[dept_cd] = BsMbr.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept_cd).count()

            # BsDept 테이블을 이용하여 dept_cd를 dept_nm으로 치환한다.
            dept_emp_to_nm = {}
            for dept_cd in dept_cd_list:
                dept_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept_cd).dept_nm
                dept_emp_to_nm[dept_nm] = dept_emp_to[dept_cd]

            dept_emp_to = dept_emp_to_nm

            # 부서 직무 목록
            # 부서 직무 목록 dept_jobs는, BsJobDept 테이블에서 각 dept_cd에 해당하는 job_cd를 가져와서 그 job_cd에 해당하는 job_nm을 나열한 dictionary 형태의 객체이다.
            dept_jobs = {}
            for dept in dept_cd_list:
                job_list = JobTask.objects.filter(prd_cd=prd_cd_selected, dept_cd=dept).order_by('job_seq').values_list('job_cd', flat=True).distinct()
                job_list = list(job_list)
                job_nm_list = []
                for job in job_list:
                    job_nm = BsJob.objects.get(prd_cd=prd_cd_selected, job_cd=job).job_nm
                    job_nm_list.append(job_nm)
                dept_jobs[dept] = job_nm_list

            # 부서 이름으로 치환
            for dept in dept_cd_list:
                dept_nm = BsDept.objects.get(prd_cd=prd_cd_selected, dept_cd=dept).dept_nm
                dept_jobs[dept_nm] = dept_jobs.pop(dept)

            # # 공통직무 개수
            # 공통직무 개수는 해당 회기의 BsJob 테이블에서 job_cd가 "JC"로 시작하는 job_cd의 개수이다.
            # commont_job = 4
            commont_job = BsJob.objects.filter(prd_cd=prd_cd_selected, job_cd__startswith='JC').count()

            # org_job_data 구성
            org_job_data = {}
            for grp in org_grp:

                dept_data = {}

                for dept in org_grp_dept[grp]:

                    dept_job_data = {}

                    dept_job_data['현인원'] = dept_emp_to[dept]
                    dept_job_data['직무수'] = len(dept_jobs[dept])
                    dept_job_data['직무'] = dept_jobs[dept]
                    dept_data[dept] = dept_job_data

                org_job_data[grp] = dept_data

            # 본부, 부서, 직무
            cnt_all_div = 0
            cnt_all_dept = 0
            cnt_all_job = 0
            for dk, dv in org_job_data.items(): # 본부 - Dictionary
                cnt_all_div += 1
                for tk, tv in dv.items():   # 부서 - Dictionary
                    cnt_all_dept += 1
                    for job in tv["직무"]:  # 직무 - List
                        cnt_all_job += 1

            wb = Workbook()
            ws = wb.active  # 현재 활성화된 sheet 가져옴
            ws.title = "직무분류체계"

            TITLE_ROW = 1 # 첫번째 행 "제목"
            DIV_ROW = 3 # 본부명 Row
            DEP_ROW = 5 # 부서명 Row
            EMP_ROW = 6 # 현인원 Row
            JOB_CNT_ROW = 7 # 직무수 Row
            JOB_START_ROW = 8 # 직무명 Start Row

            COMMON_JOB_CNT = commont_job  # 공통직무 개수

            TAG_COL = 1     # 태그 컬럼
            START_COL = 3

            JOB_ROW_HEIGHT = 70 # 직무명 행 높이

            COL_WIDTH = 15      # 컬럼 기본 크기
            COL_INTERVAL = 2    # 열과 열의 구분 간격

            # 행 높이 지정
            for c in range(1,30):
                if c == TITLE_ROW:
                    ws.row_dimensions[c].height = 70
                elif c == DIV_ROW:
                    ws.row_dimensions[c].height = 50
                elif c == DIV_ROW+1:
                    ws.row_dimensions[c].height = 10
                elif c == DEP_ROW:
                    ws.row_dimensions[c].height = 40
                elif c == EMP_ROW or c == JOB_CNT_ROW:
                    ws.row_dimensions[c].height = 30
                else:
                    ws.row_dimensions[c].height = JOB_ROW_HEIGHT

            # 열 너비 지정(A~Z열)
            alphabet_list = list(ascii_uppercase)
            for c in alphabet_list:
                if c == "A":    # 설명 Tag
                    ws.column_dimensions[c].width = COL_WIDTH
                elif c == "B":  # 간격
                    ws.column_dimensions[c].width = COL_INTERVAL
                else:
                    ws.column_dimensions[c].width = COL_WIDTH
            # 열 너비 지정(AA~AZ열)
            for c in alphabet_list:
                ws.column_dimensions["A"+c].width = COL_WIDTH

            # 테두리 적용
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            border_medium = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
            border_thick = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

            """
            설명 Tag 표시
            """
            EMP_ROW = 6     # 현인원 Row
            JOB_CNT_ROW = 7 # 직무수 Row
            JOB_START_ROW = 8   # 직무명 Start Row
            tag_emp = ws.cell(row=EMP_ROW, column=TAG_COL)
            tag_emp.value = "현 인 원 :"
            tag_emp.alignment = Alignment(horizontal="center", vertical="center")
            tag_emp.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")
            tag_job_cnt = ws.cell(row=JOB_CNT_ROW, column=TAG_COL)
            tag_job_cnt.value = "직 무 수 :"
            tag_job_cnt.alignment = Alignment(horizontal="center", vertical="center")
            tag_job_cnt.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")
            tag_common_job = ws.cell(row=JOB_START_ROW, column=TAG_COL)
            tag_common_job.value = "공통직무 :"
            tag_common_job.alignment = Alignment(horizontal="center", vertical="center")
            tag_common_job.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")
            tag_specific_job = ws.cell(row=JOB_START_ROW+COMMON_JOB_CNT, column=TAG_COL)
            tag_specific_job.value = "고유직무 :"
            tag_specific_job.alignment = Alignment(horizontal="center", vertical="center")
            tag_specific_job.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")

            """
            본부, 부서, 직무 데이터 표시
            """
            cnt_dep = 0
            col_div = START_COL
            last_col = 0
            for dk, dv in org_job_data.items(): # 본부 - Dictionary
                col_div += cnt_dep
                # print(dk, col_div)
                # 본부명
                div_name = ws.cell(row=DIV_ROW, column=col_div)
                div_name.value = dk
                div_name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                div_name.font = Font(color="0000FF", size=13, bold=True)
                div_name.fill = PatternFill(fgColor="99ccff", fill_type="solid")
                cnt_dep = 0
                for tk, tv in dv.items():   # 부서 - Dictionary
                    # 부서명
                    dep_name = ws.cell(row=DEP_ROW, column=col_div+cnt_dep)
                    dep_name.value = tk
                    dep_name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    dep_name.border = border_medium
                    # 현인원
                    employee = ws.cell(row=EMP_ROW, column=col_div+cnt_dep)
                    employee.value = tv["현인원"]
                    employee.alignment = Alignment(horizontal="center", vertical="center")
                    employee.border = border_thin
                    # 직무수
                    job_cnt = ws.cell(row=JOB_CNT_ROW, column=col_div+cnt_dep)
                    job_cnt.value = tv["직무수"]
                    job_cnt.alignment = Alignment(horizontal="center", vertical="center")
                    job_cnt.border = border_thin
                    cnt_job = 0
                    for job in tv["직무"]:  # 직무 - List
                        # 직무명
                        job_name = ws.cell(row=JOB_START_ROW+cnt_job, column=col_div+cnt_dep)
                        job_name.value = job
                        job_name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        job_name.border = border_thin
                        # 공통직무는 셀 색상을 다르게 설정
                        if job_name.row < JOB_START_ROW + COMMON_JOB_CNT:
                            job_name.fill = PatternFill(fgColor="edf0f3", fill_type="lightGray")
                        cnt_job += 1
                        last_col = col_div+cnt_dep
                    cnt_dep += 1

                # 본부명 셀 병합
                if cnt_dep > 0:
                    ws.merge_cells(start_row=DIV_ROW, start_column=col_div, end_row=DIV_ROW, end_column=col_div+cnt_dep-1)

                # 본부와 본부 사이를 구분하는 열 추가
                col_letter = get_column_letter(col_div+cnt_dep)
                ws.column_dimensions[col_letter].width = COL_INTERVAL
                cnt_dep += 1

            """
            제목 표시
            """
            title = ws.cell(row=TITLE_ROW, column=1)
            title.value = "대성에너지 직무분류체계"
            title.font = Font(color="0000FF", size=25, bold=True)   # Font 스타일
            title.alignment = Alignment(horizontal="center", vertical="center")
            # 제목 셀 병합
            if last_col > 0:
                ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=last_col)

            """
            페이지 설정 및 인쇄 옵션
            """
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = False

            """
            엑셀 파일 저장
            """

            now = dt.datetime.now()

            # 엑셀 파일을 BytesIO 객체에 저장
            excel_buffer = BytesIO()
            excel_file = "org_job_" + str(now) + ".xlsx"
            wb.save(excel_buffer)
            wb.close()
            excel_buffer.seek(0)

            encoded_filename = urllib.parse.quote(excel_file)

            # HttpResponse로 파일 전송
            response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

            return response # 엑셀 파일 다운로드

        elif action == 'action2': # 업무 분장표 눌렀을 때

            # SQLAlchemy 엔진을 사용하여 데이터베이스에 연결
            engine = create_engine('mysql+pymysql://jyj:1234@218.159.103.251/jobdb')
            db_name = 'jobdb'

            # prd_cd = '2022A'
            prd_cd = prd_cd_selected
            dept_domain = dept_domain_selected

            def get_data_from_db(engine, db_name, table_name, where_clause):
                query = f"SELECT * FROM {db_name}.{table_name} WHERE {where_clause}"
                df = pd.read_sql(query, engine)
                return df

            """ 직무정보 """
            # 직무코드 데이터
            where_clause = f"prd_cd = '{prd_cd_selected}'"
            df_job_code = get_data_from_db(engine, db_name, "bs_job", where_clause)

            # 직무-책무-과업-활동 데이터
            df1 = get_data_from_db(engine, db_name, "job_task", where_clause)
            df2 = get_data_from_db(engine, db_name, "job_activity", where_clause)

            """ 부서 및 조직 그룹 정보 """
            # 부서 데이터
            df_dept = get_data_from_db(engine, db_name, "bs_dept", where_clause)

            # 조직 그룹 도메인
            where_clause = f"prd_cd = '{prd_cd_selected}' and dept_domain = '{dept_domain}'"
            df_dept_grp_domain = get_data_from_db(engine, db_name, "bs_dept_grp_domain", where_clause)

            # 조직 그룹
            df_dept_grp = get_data_from_db(engine, db_name, "bs_dept_grp", where_clause)

            engine.dispose()

            # 결합 Key : prd_cd, dept_cd, job_cd, duty_nm, task_nm
            data = pd.merge(df1, df2, how='right', on=['prd_cd', 'dept_cd', 'job_cd', 'duty_nm', 'task_nm'], suffixes=('_left', '_right'))
            # 중복 컬럼 데이터 삭제
            data.drop(data.filter(regex='_left'), axis=1, inplace=True)
            # 직무명 추가
            data = pd.merge(data, df_job_code[['prd_cd', 'job_cd', 'job_nm']], how='left', on=['prd_cd', 'job_cd'])
            # NaN을 None으로 변환
            data = data.replace({np.nan: None})
            # 인덱스 초기화
            data = data.reset_index(drop=True)
            assert df2.shape[0] == data.shape[0]    # 전체 데이터 건수가 df2 데이터 건수와 같아야 정상

            # 조직 그룹 및 부서
            dept_all = pd.merge(df_dept_grp_domain, df_dept_grp, how='left', on=['prd_cd', 'dept_domain', 'dept_grp_nm'])
            # 부서명 추가
            dept_all = pd.merge(dept_all, df_dept[['prd_cd', 'dept_cd', 'dept_nm', 'dept_po']], how='left', on=['prd_cd', 'dept_cd'])
            # 부서 순서 정렬
            dept_all = dept_all.sort_values(by=['grp_seq', 'dept_seq']).reset_index(drop=True)

            wb = Workbook()

            # 테두리 적용
            BORDER_THIN_UP = Border(top=Side(style='thin'))
            BORDER_THIN_ALL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            """
            첫번째 Sheet : 표지
            """
            ws = wb.active  # 현재 활성화된 sheet 가져옴
            ws.title = "표지"

            # 제목
            row_no = 2
            title = ws.cell(row=row_no, column=1)
            title.value = "팀별 업무분장표"
            title.font = Font(color="000000", size=40, bold=True)
            title.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=7)
            ws.row_dimensions[row_no].height = 100

            # 년월
            row_no = 17
            ym = ws.cell(row=row_no, column=1)
            now = dt.datetime.now()
            ym.value = str(now.year) + "년 " + str(now.month) + "월"
            ym.font = Font(color="000000", size=15, bold=True)
            ym.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=7)
            ws.row_dimensions[row_no].height = 30

            # 회사
            row_no = 33
            company = ws.cell(row=row_no, column=1)
            company.value = "대성에너지(주)"
            company.font = Font(color="000000", size=25, bold=True)
            company.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=7)
            ws.row_dimensions[row_no].height = 50

            # Page Setup
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            ws.print_options.horizontalCentered = True

            """
            두번째 Sheet : 목차
            """
            ws_toc = wb.create_sheet("목차", 2)        # 2번째 index에 sheet 생성

            # Title
            title = ws_toc.cell(row=1, column=1)
            title.value = " < 업 무 분 장 표 목 차 > "
            title.font = Font(color="000000", size=16, bold=False)
            title.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws_toc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            ws_toc.row_dimensions[1].height = 40

            # 데이터 항목 개수
            DATA_COLS = 5

            # 컬럼 열 너비
            ws_toc.column_dimensions["A"].width = 5
            ws_toc.column_dimensions["B"].width = 20
            ws_toc.column_dimensions["C"].width = 20
            ws_toc.column_dimensions["D"].width = 20
            ws_toc.column_dimensions["E"].width = 5

            # 헤더: 항목 명칭 및 속성
            header_cols = ["조직", "", "팀"]
            row_no = 4
            for i, header_name in enumerate(header_cols):
                header = ws_toc.cell(row=row_no, column=i+2)
                header.value = header_name
                header.alignment = Alignment(horizontal="center", vertical="center", wrapText=False)
                header.fill = PatternFill(fgColor="D0FA58", fill_type="solid")

            ws_toc.row_dimensions[3].height = 15

            """
            조직그룹 및 부서 순서대로 표시
            """

            prev_dept_grp_nm = None

            for idx, row in dept_all.iterrows():
                dept_grp_nm = row['dept_grp_nm']
                dept_nm = row['dept_nm']

                row_no += 1
                # 조직 그룹
                if dept_grp_nm != prev_dept_grp_nm:
                    dept_grp = ws_toc.cell(row=row_no, column=2)
                    dept_grp.value = dept_grp_nm
                    dept_grp.alignment = Alignment(horizontal="center", vertical="center", wrapText=False)
                    prev_dept_grp_nm = dept_grp_nm
                    for c in [2, 3, 4]:
                        dept_grp_cells = ws_toc.cell(row=row_no, column=c)
                        dept_grp_cells.border = BORDER_THIN_UP

                # 부서명 및 총 인원
                dept = ws_toc.cell(row=row_no, column=4)
                dept.value = dept_nm
                dept.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

            row_no += 1
            for c in [2, 3, 4]:
                last_cell = ws_toc.cell(row=row_no, column=c)
                last_cell.border = BORDER_THIN_UP

            """
            페이지 설정 및 인쇄 옵션
            """
            ws_toc.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws_toc.page_setup.orientation = ws_toc.ORIENTATION_PORTRAIT
            ws_toc.page_setup.paperSize = ws_toc.PAPERSIZE_A4
            ws_toc.page_setup.fitToHeight = 0
            ws_toc.page_setup.fitToWidth = 1
            ws_toc.print_options.horizontalCentered = True

            """
            세번째 Sheet : 업무 분장표(과업 단위)
            """
            ws_task = wb.create_sheet("업무분장표_과업", 3)        # 3번째 index에 sheet 생성

            # TITLE_ROW = 1   # 첫번째 행 "제목"

            # Title
            title = ws_task.cell(row=1, column=1)
            title.value = prd_cd[:4] + "년도 업무분장표"
            title.font = Font(color="0000FF", size=25, bold=True)
            title.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws_task.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            ws_task.row_dimensions[1].height = 75


            # 데이터 항목 개수
            DATA_COLS = 4

            # 과업 항목 열 위치
            TASK_START_COL = 3

            # 컬럼 열 너비
            ws_task.column_dimensions["A"].width = 15
            ws_task.column_dimensions["B"].width = 20
            ws_task.column_dimensions["C"].width = 30
            ws_task.column_dimensions["D"].width = 20

            """
            조직그룹 및 부서별로 직무정보 표시
            """

            row_no = 1     # 직무정보 데이터 시작 위치: 조직그룹/부서/직무정보

            prev_dept_grp_nm = None

            for idx, row in dept_all.iterrows():
                dept_grp_nm = row['dept_grp_nm']
                dept_cd = row['dept_cd']
                dept_nm = row['dept_nm']
                dept_po = row['dept_po']

                # 조직 그룹
                if dept_grp_nm != prev_dept_grp_nm:
                    row_no += 3
                    dept_grp = ws_task.cell(row=row_no, column=1)
                    dept_grp.value = " [ " + dept_grp_nm + " ]"
                    dept_grp.font = Font(color="000000", size=15, bold=True)
                    dept_grp.alignment = Alignment(horizontal="left", vertical="center", wrapText=False)
                    prev_dept_grp_nm = dept_grp_nm
                    ws_task.row_dimensions[row_no].height = 25

                # 부서명 및 총 인원
                row_no += 2
                bullet1 = ws_task.cell(row=row_no, column=1)
                bullet1.value = "  ■ 팀명 :"
                bullet1.font = Font(color="000000", size=13, bold=True)
                bullet1.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                dept = ws_task.cell(row=row_no, column=2)
                dept.value = dept_nm
                dept.font = Font(color="000000", size=13, bold=True)
                dept.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                bullet2 = ws_task.cell(row=row_no, column=3)
                bullet2.value = "■ 총 인원(직책자포함) : " + str(dept_po) + "명"
                bullet2.font = Font(color="000000", size=13, bold=True)
                bullet2.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                ws_task.row_dimensions[row_no].height = 20

                # 헤더: 항목 명칭 및 속성
                header_cols = ["직무\n(Job)", "책무\n(Duty)", "과업\n(Task)", "과업 담당자"]
                row_no += 1
                for i, header_name in enumerate(header_cols):
                    header = ws_task.cell(row=row_no, column=i+1)
                    header.value = header_name
                    header.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                    header.fill = PatternFill(fgColor="D0FA58", fill_type="solid")

                # 직무정보 : 1줄씩 데이터 추가
                job_data = data[data['dept_cd'] == dept_cd]
                job_data = job_data.sort_values(by=['job_seq_right', 'duty_seq_right', 'task_seq_right', 'act_seq']).reset_index(drop=True)
                prev_job_nm = prev_duty_nm = prev_task_nm = None
                # row_no += 1
                for i, r in job_data.iterrows():
                    row_no += 1
                    # 직무명
                    job_nm = ws_task.cell(row=row_no, column=1)
                    if r['job_nm'] == prev_job_nm:  # 동일 데이터 반복 제거
                        job_nm.value = ""
                    else:
                        job_nm.value = prev_job_nm = r['job_nm']
                        job_nm.border = BORDER_THIN_UP
                        job_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                    # 책무
                    duty_nm = ws_task.cell(row=row_no, column=2)
                    if r['duty_nm'] == prev_duty_nm:  # 동일 데이터 반복 제거
                        duty_nm.value = ""
                    else:
                        duty_nm.value = prev_duty_nm = r['duty_nm']
                        duty_nm.border = BORDER_THIN_UP
                        duty_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                    """ 과업 데이터 """
                    # 과업, 과업 담당자, 과업 업무특성
                    task_nm = ws_task.cell(row=row_no, column=3)
                    if r['task_nm'] == prev_task_nm:  # 동일 데이터 반복 제거
                        task_nm.value = ""
                    else:
                        task_nm.value = prev_task_nm = r['task_nm']
                        task_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                        for c in range(TASK_START_COL, DATA_COLS+1):
                            task_related_cell = ws_task.cell(row=row_no, column=c)
                            task_related_cell.border = BORDER_THIN_UP

                    """ 과업 담당자 """
                    task_prsn_chrg = ws_task.cell(row=row_no, column=4)
                    task_prsn_chrg.value = r['task_prsn_chrg']
                    task_prsn_chrg.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

            """
            페이지 설정 및 인쇄 옵션
            """
            ws_task.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws_task.page_setup.orientation = ws_task.ORIENTATION_PORTRAIT
            ws_task.page_setup.paperSize = ws_task.PAPERSIZE_A4
            ws_task.page_setup.fitToHeight = 0
            ws_task.page_setup.fitToWidth = 1
            ws_task.print_options.horizontalCentered = True

            """
            네번째 Sheet : 업무 분장표(활동 단위)
            """
            ws_data = wb.create_sheet("업무분장표_활동", 4)        # 3번째 index에 sheet 생성

            # Title
            title = ws_data.cell(row=1, column=1)
            title.value = prd_cd[:4] + "년도 업무분장표"
            title.font = Font(color="0000FF", size=25, bold=True)
            title.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws_data.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            ws_data.row_dimensions[1].height = 75

            # 데이터 항목 개수
            DATA_COLS = 5

            # 과업 항목 열 위치
            TASK_START_COL = 3

            # 컬럼 열 너비
            ws_data.column_dimensions["A"].width = 15
            ws_data.column_dimensions["B"].width = 20
            ws_data.column_dimensions["C"].width = 30
            ws_data.column_dimensions["D"].width = 75
            ws_data.column_dimensions["E"].width = 20

            """
            조직그룹 및 부서별로 직무정보 표시
            """

            row_no = 1     # 직무정보 데이터 시작 위치: 조직그룹/부서/직무정보

            prev_dept_grp_nm = None

            for idx, row in dept_all.iterrows():
                dept_grp_nm = row['dept_grp_nm']
                dept_cd = row['dept_cd']
                dept_nm = row['dept_nm']
                dept_po = row['dept_po']

                # 조직 그룹
                if dept_grp_nm != prev_dept_grp_nm:
                    row_no += 3
                    dept_grp = ws_data.cell(row=row_no, column=1)
                    dept_grp.value = " [ " + dept_grp_nm + " ]"
                    dept_grp.font = Font(color="000000", size=15, bold=True)
                    dept_grp.alignment = Alignment(horizontal="left", vertical="center", wrapText=False)
                    prev_dept_grp_nm = dept_grp_nm
                    ws_data.row_dimensions[row_no].height = 25

                # 부서명 및 총 인원
                row_no += 2
                bullet1 = ws_data.cell(row=row_no, column=1)
                bullet1.value = "  ■ 팀명 :"
                bullet1.font = Font(color="000000", size=13, bold=True)
                bullet1.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                dept = ws_data.cell(row=row_no, column=2)
                dept.value = dept_nm
                dept.font = Font(color="000000", size=13, bold=True)
                dept.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                bullet2 = ws_data.cell(row=row_no, column=3)
                bullet2.value = "■ 총 인원(직책자포함) : " + str(dept_po) + "명"
                bullet2.font = Font(color="000000", size=13, bold=True)
                bullet2.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
                ws_data.row_dimensions[row_no].height = 20

                # 헤더: 항목 명칭 및 속성
                header_cols = ["직무\n(Job)", "책무\n(Duty)", "과업\n(Task)", "활동\n(Activity)", "활동 담당자"]
                row_no += 1
                for i, header_name in enumerate(header_cols):
                    header = ws_data.cell(row=row_no, column=i+1)
                    header.value = header_name
                    header.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                    header.fill = PatternFill(fgColor="D0FA58", fill_type="solid")

                # 직무정보 : 1줄씩 데이터 추가
                job_data = data[data['dept_cd'] == dept_cd]
                job_data = job_data.sort_values(by=['job_seq_right', 'duty_seq_right', 'task_seq_right', 'act_seq']).reset_index(drop=True)
                prev_job_nm = prev_duty_nm = prev_task_nm = None
                # row_no += 1
                for i, r in job_data.iterrows():
                    row_no += 1
                    # 직무명
                    job_nm = ws_data.cell(row=row_no, column=1)
                    if r['job_nm'] == prev_job_nm:  # 동일 데이터 반복 제거
                        job_nm.value = ""
                    else:
                        job_nm.value = prev_job_nm = r['job_nm']
                        job_nm.border = BORDER_THIN_UP
                        job_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                    # 책무
                    duty_nm = ws_data.cell(row=row_no, column=2)
                    if r['duty_nm'] == prev_duty_nm:  # 동일 데이터 반복 제거
                        duty_nm.value = ""
                    else:
                        duty_nm.value = prev_duty_nm = r['duty_nm']
                        duty_nm.border = BORDER_THIN_UP
                        duty_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                    """ 과업 데이터 """
                    # 과업, 과업 담당자, 과업 업무특성
                    task_nm = ws_data.cell(row=row_no, column=3)
                    if r['task_nm'] == prev_task_nm:  # 동일 데이터 반복 제거
                        task_nm.value = ""
                    else:
                        task_nm.value = prev_task_nm = r['task_nm']
                        task_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                        for c in range(TASK_START_COL, DATA_COLS+1):
                            task_related_cell = ws_data.cell(row=row_no, column=c)
                            task_related_cell.border = BORDER_THIN_UP

                    """ 활동 데이터 """
                    act_nm = ws_data.cell(row=row_no, column=4)
                    act_nm.value = r['act_nm']
                    act_nm.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

                    """ 활동 담당자 """
                    act_prsn_chrg = ws_data.cell(row=row_no, column=5)
                    act_prsn_chrg.value = r['act_prsn_chrg']
                    act_prsn_chrg.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

            """
            페이지 설정 및 인쇄 옵션
            """
            ws_data.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=True)
            ws_data.page_setup.orientation = ws_data.ORIENTATION_LANDSCAPE
            ws_data.page_setup.paperSize = ws_data.PAPERSIZE_A4
            ws_data.page_setup.fitToHeight = 0
            ws_data.page_setup.fitToWidth = 1
            ws_data.print_options.horizontalCentered = True

            """
            엑셀 파일 저장
            """
            # 엑셀 파일을 BytesIO 객체에 저장
            excel_buffer = BytesIO()
            excel_file = "org_job_" + str(now) + ".xlsx"
            wb.save(excel_buffer)
            wb.close()
            excel_buffer.seek(0)

            encoded_filename = urllib.parse.quote(excel_file)

            # HttpResponse로 파일 전송
            response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

            return response # 엑셀 파일 다운로드

    return render(request, 'jobs/JB300.html', context)


def main(request):

    context = {
        'dept_mgr_yn' : get_dept_mgr_yn(request.user.username),
    }

    return render(request, 'jobs/main.html', context)


############################################ 함수 ############################################

def BsMbrArrange(prd, dept): # 부서원 표시 함수 - 수정해야함

    ttl_nm_list = BsTtlList.objects.filter(prd_cd_id=prd).order_by('ttl_ordr').values_list('ttl_nm', flat=True)
    pos_nm_list = BsPosList.objects.filter(prd_cd_id=prd).order_by('pos_ordr').values_list('pos_nm', flat=True)

    result = BsMbr.objects.filter(prd_cd_id=prd, dept_cd_id = dept).order_by(
                Case(
                    When(ttl_nm='팀장', then=Value(0)),
                    When(ttl_nm='부팀장', then=Value(1)),
                    When(ttl_nm='팀원', then=Value(2)),
                    default=Value(3),  # 다른 값들은 default로 처리
                    output_field=CharField(),
                ),
                Case(
                    When(pos_nm='수석부장', then=Value(0)),
                    When(pos_nm='부장', then=Value(1)),
                    When(pos_nm='차장', then=Value(2)),
                    When(pos_nm='과장', then=Value(3)),
                    When(pos_nm='대리', then=Value(4)),
                    When(pos_nm='사원', then=Value(5)),
                    default=Value(6),  # 다른 값들은 default로 처리
                    output_field=CharField(),
                )
                ,'mbr_nm'  # pos_nm이 같을 경우 mbr_nm으로 추가 정렬
                )
    return result


def copy_period_data(period_old, period_new):
    # 데이터베이스 연결 파라미터
    user_id = 'jyj'  # 사용자 이름
    pwd = '1234'  # 비밀번호
    db_host = '218.159.103.251'  # 호스트명/IP
    db_port = 3306  # 포트번호 (고정값)
    db_name = "jobdb"  # 사용할 데이터베이스 jobdb

    dict_table = {  # 테이블 목록
        'bs_prd': '회기',
        'bs_std_wrk_tm': '표준근무시간',
        'bs_wl_ov_sht': '업무량과부족 산정 기준',
        'bs_work_grade': '업무 등급',
        'bs_ttl_list': '직책 리스트',
        'bs_pos_list': '직위 리스트',
        'bs_pos_grade': '업무 등급별 직위',
        'bs_dept': '부서',
        'bs_dept_resp': '부서 성과책임',
        'bs_dept_grp_domain': '부서 그룹 도메인',
        'bs_dept_grp': '부서 그룹',
        'bs_mbr': '부서원',
        'bs_ttl_cnt': '직책별 부서원수',
        'bs_mbr_grp_nm': '부서원 그룹명',
        'bs_mbr_grp': '부서원 그룹',
        'bs_acnt': '계정',
        'bs_job': '직무 리스트',
        'bs_job_resp': '직무 성과책임',
        'bs_job_dept': '부서별 직무',
        'job_task': '직무 상세_과업',
        'job_activity': '직무 상세_활동',
        'job_spcfc': '직무명세서'
    }

    conn = pymysql.connect(host=db_host, user=user_id, password=pwd, db=db_name, charset='utf8mb4')
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    messages = []  # 메시지를 수집할 리스트
    messages.append(f"{period_old} 회기 정보를 {period_new} 회기 정보로 복제합니다\n")

    result = None
    for key, value in dict_table.items():
        try:
            # 데이터 선택
            sql_str = f"SELECT * FROM {key} WHERE prd_cd = %s"
            cursor.execute(sql_str, (period_old,))
            rows = cursor.fetchall()

            # 데이터 삽입
            for row in rows:
                row['prd_cd'] = period_new
                columns = ', '.join(row.keys())
                placeholders = ', '.join(['%s'] * len(row))
                insert_sql = f"INSERT INTO {key} ({columns}) VALUES ({placeholders})"
                cursor.execute(insert_sql, tuple(row.values()))

            conn.commit()
            result = True
            messages.append(f"{value} 정보 복제 완료")

        except Exception as e:
            conn.rollback()
            result = False
            messages.append(f"{value} 정보 복제 오류")
            messages.append(traceback.format_exc())
            break

    cursor.close()
    conn.close()

    if result:
        messages.append(f"\n{period_new} 회기 정보가 생성 완료되었습니다.")

    return messages


def delete_period_data(period):
    # 데이터베이스 연결 파라미터
    user_id = 'jyj'  # 사용자 이름
    pwd = '1234'  # 비밀번호
    db_host = '218.159.103.251'  # 호스트명/IP
    db_port = 3306  # 포트번호 (고정값)
    db_name = "jobdb"  # 사용할 데이터베이스 jobdb

    dict_table = {  # 테이블 목록
        'job_spcfc': '직무명세서',
        'job_activity': '직무 상세_활동',
        'job_task': '직무 상세_과업',
        'bs_job_dept': '부서별 직무',
        'bs_job_resp': '직무 성과책임',
        'bs_job': '직무 리스트',
        'bs_acnt': '계정',
        'bs_mbr_grp': '부서원 그룹',
        'bs_mbr_grp_nm': '부서원 그룹명',
        'bs_ttl_cnt': '직책별 부서원수',
        'bs_mbr': '부서원',
        'bs_dept_grp': '부서 그룹',
        'bs_dept_grp_domain': '부서 그룹 도메인',
        'bs_dept_resp': '부서 성과책임',
        'bs_dept': '부서',
        'bs_pos_grade': '업무 등급별 직위',
        'bs_pos_list': '직위 리스트',
        'bs_ttl_list': '직책 리스트',
        'bs_work_grade': '업무 등급',
        'bs_wl_ov_sht': '업무량과부족 산정 기준',
        'bs_std_wrk_tm': '표준근무시간',
        'bs_prd': '회기'
    }

    conn = pymysql.connect(host=db_host, user=user_id, password=pwd, db=db_name, charset='utf8mb4')
    cursor = conn.cursor()

    messages = []  # 메시지를 수집할 리스트
    messages.append(f"{period} 회기 정보를 삭제합니다\n")
    result = None
    for key, value in dict_table.items():
        try:
            sql_str = f"DELETE FROM {key} WHERE prd_cd = %s"
            cursor.execute(sql_str, (period,))
            conn.commit()
            result = True
            messages.append(f"{value} 정보가 삭제되었습니다")

        except Exception as e:
            conn.rollback()
            result = False
            messages.append("... 삭제 오류. 삭제를 중단합니다.")
            messages.append(traceback.format_exc())
            break

    cursor.close()
    conn.close()

    if result:
        messages.append(f"\n{period} 회기 정보 삭제가 완료되었습니다.")

    return messages


def get_dept_code(user_id):
    prd_cd_id = "2024A"  # 상수로 지정
    try:
        account = BsAcnt.objects.get(dept_id=user_id, prd_cd_id=prd_cd_id)
        return account.dept_cd_id
    except:
        return None  # 부서 코드가 없는 경우에 대한 처리


def get_dept_mgr_yn(user_id):
    prd_cd_id = "2024A"
    try:
        account_cd = BsAcnt.objects.get(dept_id=user_id, prd_cd_id=prd_cd_id).dept_cd_id
        dept_mgr_yn = BsDept.objects.get(prd_cd_id=prd_cd_id, dept_cd=account_cd).dept_mgr_yn
        return dept_mgr_yn
    except:
        return None